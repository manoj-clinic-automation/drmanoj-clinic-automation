#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
portal_followups.py — Follow-up batch WhatsApp (S172, Phase A.2)
================================================================
Reads the daily "Staff_Action_Today_YYYY-MM-DD.xlsx" that the follow-up tracker
produces (today: written on the clinic PC and pushed to the VPS; when the tracker
moves to the VPS it writes the SAME path and nothing here changes). The batch
screen never cares who wrote the file — it reads a FILE at a fixed dir. (D-stable)

It parses the "Call Sheet", turns each follow-up row into a ready-to-send item
(template auto-picked from OD = overdue days), and lets the doctor/agent
checkbox rows section-wise + patient-wise and fire them through the ONE canonical
sender (portal_wa.send) — same log, same DRY-RUN switch, same family-correct keys.

Template ladder by OD (overdue days), matching WABA_Approved_Templates (S137):
    OD  < 0   -> drmanoj_followup_tomorrow   (due tomorrow)
    0 <= OD<=3 -> drmanoj_followup_due        (due today / grace)
    4 <= OD<=10-> drmanoj_followup_missed     (missed)
    OD  > 10   -> drmanoj_followup_dropout    (dropout; var 3 = days overdue)

Only rows in a FOLLOW-UP section with a valid 10-digit mobile + name are marked
sendable. Procedure call-backs and other sections are shown but not WABA-sendable
(use the single-send widget / wa.me fallback for those).
"""

import os, re, glob
import openpyxl

FOLLOWUP_DIR = os.environ.get("PORTAL_FOLLOWUP_DIR", "/root/wa/followups")
FILE_GLOB    = os.environ.get("PORTAL_FOLLOWUP_GLOB", "Staff_Action_Today*.xlsx")
SHEET_NAME   = "Call Sheet"

SECTION_RE = re.compile(r"^\s*\d+\.\s+(.+?)\s*(?:\u2014|\u2013|\s-\s|$)")


def find_latest(dirpath=None):
    d = dirpath or FOLLOWUP_DIR
    files = glob.glob(os.path.join(d, FILE_GLOB))
    return max(files, key=os.path.getmtime) if files else None


def od_template(od):
    """(template_name, ok) from overdue-days integer."""
    if od is None:
        return (None, False)
    if od < 0:
        return ("drmanoj_followup_tomorrow", True)
    if od <= 3:
        return ("drmanoj_followup_due", True)
    if od <= 10:
        return ("drmanoj_followup_missed", True)
    return ("drmanoj_followup_dropout", True)


def _fmt_date(raw, year):
    if not raw:
        return ""
    s = str(raw).strip()
    m = re.match(r"(\d{1,2})[-\s/]([A-Za-z]{3})", s)
    if m:
        return "%02d %s %s" % (int(m.group(1)), m.group(2).title(), year)
    return s


def parse(path):
    """Read the Call Sheet -> {file, year, sections:[{name, rows:[...]}]}.
    Read-only; never writes. Rows carry a full phone (doctor-session use) plus a
    masked phone for display."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError("sheet '%s' not found in %s" % (SHEET_NAME, os.path.basename(path)))
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    year = "2026"
    for r in rows[:3]:
        for c in (r or []):
            m = re.search(r"(20\d{2})", str(c or ""))
            if m:
                year = m.group(1); break

    sections = []; cur = None; in_table = False
    for r in rows:
        a = str(r[0]).strip() if (r and r[0] is not None) else ""
        sm = SECTION_RE.match(a)
        if sm:
            cur = {"name": sm.group(1).strip(), "rows": []}
            sections.append(cur); in_table = False
            continue
        if a.upper() == "S.N":
            in_table = True; continue
        if in_table and cur is not None and a.isdigit():
            def g(i):
                return r[i] if (len(r) > i and r[i] is not None) else ""
            name = str(g(2)).strip()
            phone = re.sub(r"\D", "", str(g(3)))
            if len(phone) == 11 and phone[0] == "0":
                phone = phone[1:]
            date_raw = str(g(5)).strip()
            od = None
            if str(g(6)).strip() != "":
                try:
                    od = int(float(g(6)))
                except Exception:
                    od = None
            status = str(g(7)).strip()
            tpl, ok = od_template(od)
            sec_is_fu = "FOLLOW" in (cur["name"] or "").upper()
            sendable = bool(sec_is_fu and ok and len(phone) == 10 and name)
            cur["rows"].append({
                "sn": int(a), "name": name,
                "phone": phone if len(phone) == 10 else "",
                "phone_mask": ("\u2022\u2022\u2022\u2022" + phone[-4:]) if len(phone) >= 4 else "",
                "od": od, "date_raw": date_raw, "date_fmt": _fmt_date(date_raw, year),
                "status": status, "diagnosis": str(g(4)).strip(),
                "template": tpl if (ok and sec_is_fu) else "",
                "sendable": sendable})
    return {"file": os.path.basename(path), "year": year, "sections": sections}


def values_for(template_name, row):
    """Build the WABA body values dict for a follow-up row + chosen template."""
    name = row.get("name", ""); date = row.get("date_fmt", ""); od = row.get("od")
    if template_name == "drmanoj_post_visit":
        return {"1": name}
    if template_name == "drmanoj_followup_dropout":
        return {"1": name, "2": date, "3": str(od if od is not None else "")}
    if template_name in ("drmanoj_followup_tomorrow", "drmanoj_followup_due",
                         "drmanoj_followup_missed"):
        return {"1": name, "2": date}
    return {"1": name, "2": date}


# --------------------------- portal wiring ---------------------------------- #
def register(app, guard, get_user, cfg_get, wa_send):
    """wa_send = portal_wa.send  (the ONE canonical sender). cfg_get = portal config."""
    from flask import request, jsonify, Response

    def _token(): return cfg_get("MYOP_AUTH_TOKEN", "") or os.environ.get("MYOP_AUTH_TOKEN", "")
    def _dry():   return (cfg_get("PORTAL_WA_DRYRUN", "1") or "1") != "0"
    WA_DIR = os.environ.get("PORTAL_WA_DIR", "/root/wa/wa_portal")

    @app.route("/portal/wa/followups")
    @guard
    def fu_page():
        fp = os.path.join(WA_DIR, "followups_page.html")
        if os.path.exists(fp):
            return Response(open(fp, encoding="utf-8").read(),
                            mimetype="text/html; charset=utf-8")
        return Response("followups_page.html not installed at " + fp, status=500,
                        mimetype="text/plain")

    @app.route("/portal/wa/followups/data")
    @guard
    def fu_data():
        try:
            path = find_latest()
            if not path:
                return jsonify({"ok": False,
                                "error": "no Staff_Action_Today file found in " + FOLLOWUP_DIR})
            d = parse(path)
            d["ok"] = True; d["dry"] = _dry()
            return jsonify(d)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/portal/wa/followups/send", methods=["POST"])
    @guard
    def fu_send():
        b = request.get_json(force=True, silent=True) or {}
        items = b.get("items") or []
        if not items:
            return jsonify({"ok": False, "error": "no rows selected"}), 400
        who = get_user() or "portal"
        tok = _token(); dry = _dry()
        results = []
        sent = 0
        for it in items[:500]:
            r = wa_send(it.get("phone", ""), it.get("template", ""),
                        it.get("values") or {}, who, tok, dry_run=dry)
            if r.get("ok"):
                sent += 1
            results.append({"sn": it.get("sn"), "name": it.get("name", ""),
                            "ok": r.get("ok"), "mode": r.get("mode"),
                            "error": r.get("error", "")})
        return jsonify({"ok": True, "sent": sent, "total": len(results),
                        "dry": dry, "results": results})

    return True
