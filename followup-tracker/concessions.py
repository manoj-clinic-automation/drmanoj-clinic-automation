"""
concessions.py
Dr. Manoj Agarwal Clinic — VIP & Concession List (read-only query layer)

Reads the concession profile that the diagnosis layer extracts into
    data/patient_diagnosis.csv
(columns Admin_CC, Admin_PD, Admin_BID, Is_VIP, Concession_Scheme) and exposes
it for the /concessions query page and an Excel export.

This module NEVER writes to patient_diagnosis.csv. It only reads. Refreshing the
profile is done by re-absorbing a Docterz patient export through the diagnosis
route, exactly as before.
"""

import json
import datetime as dt
from pathlib import Path
import pandas as pd

DATA_DIR  = Path(__file__).resolve().parent / "data"
DIAG_FILE = DATA_DIR / "patient_diagnosis.csv"
META_FILE = DATA_DIR / "diagnosis_source_meta.json"


def _is_set(v) -> bool:
    s = str(v).strip()
    return s not in ("", "nan", "none", "None", "False")


def _asof_label():
    """Return (label, is_blank). When the absorbed export filename carried no
    'UPTO DD MON YYYY' date, as_of_date is blank — we then fall back to the
    ingest timestamp and flag it so the page can warn the doctor."""
    try:
        m = json.loads(META_FILE.read_text(encoding="utf-8"))
        a = (m.get("as_of_date") or "").strip()
        if a:
            return a, False
        ing = (m.get("ingested_on") or "").strip()
        return (ing or "unknown"), True
    except Exception:
        return "unknown", True


def load_concessions():
    """Every patient carrying ANY concession marker (VIP, a CC/PD/BID code, or a
    named scheme), as a list of plain dicts ready for the table / export."""
    if not DIAG_FILE.exists():
        return []
    d = pd.read_csv(DIAG_FILE, dtype=str).fillna("")
    rows = []
    for _, r in d.iterrows():
        vip    = str(r.get("Is_VIP", "")).strip().lower() == "true"
        cc     = str(r.get("Admin_CC", "")).strip()
        pdv    = str(r.get("Admin_PD", "")).strip()
        bid    = str(r.get("Admin_BID", "")).strip()
        scheme = str(r.get("Concession_Scheme", "")).strip()
        has_cc, has_pd, has_bid = _is_set(cc), _is_set(pdv), _is_set(bid)
        if not (vip or has_cc or has_pd or has_bid or scheme):
            continue
        rows.append({
            "uid":       r.get("Patient_UID", ""),
            "clinic_id": r.get("Clinic_Specific_Id", ""),
            "name":      r.get("Patient_Name", "") or "(no name)",
            "mobile":    r.get("Mobile_Clean", ""),
            "age":       r.get("Age", ""),
            "sex":       r.get("Sex", ""),
            "vip":       vip,
            "scheme":    scheme,
            "cc":        cc  if has_cc  else "",
            "pd":        pdv if has_pd  else "",
            "bid":       bid if has_bid else "",
            "diagnosis": r.get("Standardized_Diagnosis", ""),
            "priority":  r.get("Diagnosis_Priority", ""),
            "updated":   r.get("Last_Updated", ""),
        })
    # VIPs first, then patients on a named scheme, then alphabetical
    rows.sort(key=lambda x: (not x["vip"], x["scheme"] == "", x["name"].lower()))
    return rows


def concession_summary():
    rows = load_concessions()
    asof, asof_blank = _asof_label()
    schemes = {}
    for r in rows:
        for s in str(r["scheme"]).split(";"):
            s = s.strip()
            if s:
                schemes[s] = schemes.get(s, 0) + 1
    return {
        "rows":     rows,
        "total":    len(rows),
        "vip":      sum(1 for r in rows if r["vip"]),
        "with_cc":  sum(1 for r in rows if r["cc"]  != ""),
        "with_pd":  sum(1 for r in rows if r["pd"]  != ""),
        "with_bid": sum(1 for r in rows if r["bid"] != ""),
        "schemes":  sorted(schemes.items(), key=lambda x: -x[1]),
        "asof":     asof,
        "asof_blank": asof_blank,
    }


def build_concession_xlsx(path):
    """Write the concession list to an .xlsx for download / staff reference."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    s = concession_summary()
    wb = Workbook()
    ws = wb.active
    ws.title = "VIP & Concessions"

    ws.append([f"VIP & Concession List — {s['total']} patients (as of {s['asof']})"])
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws.append([f"VIP {s['vip']}   ·   CC {s['with_cc']}   ·   PD {s['with_pd']}   ·   BID {s['with_bid']}"])
    ws.append([])

    hdr = ["Clinic ID", "Patient Name", "Mobile", "Age", "Sex", "VIP", "Scheme",
           "Cons. Charge (CC ₹)", "Pharmacy Disc % (PD)", "Blood-Inv Disc % (BID)",
           "Diagnosis", "Priority"]
    ws.append(hdr)
    hr = ws.max_row
    fill = PatternFill("solid", fgColor="1F4E79")
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for r in s["rows"]:
        ws.append([r["clinic_id"], r["name"], r["mobile"], r["age"], r["sex"],
                   "VIP" if r["vip"] else "", r["scheme"], r["cc"], r["pd"],
                   r["bid"], r["diagnosis"], r["priority"]])
        if r["vip"]:
            ws.cell(row=ws.max_row, column=6).font = Font(bold=True, color="C00000")

    widths = [10, 24, 13, 5, 5, 6, 16, 18, 18, 20, 30, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    wb.save(path)
    return path


if __name__ == "__main__":
    s = concession_summary()
    print(f"as of {s['asof']}{'  (no date in source filename)' if s['asof_blank'] else ''}")
    print(f"total tagged {s['total']}  | VIP {s['vip']}  CC {s['with_cc']}  "
          f"PD {s['with_pd']}  BID {s['with_bid']}")
    print("schemes:", s["schemes"])
    for r in s["rows"][:10]:
        print(f"  {r['clinic_id']:>6}  {r['name'][:22]:22}  VIP={r['vip']!s:5} "
              f"CC={r['cc']:>3} PD={r['pd']:>3} BID={r['bid']:>3}  {r['scheme']}")
