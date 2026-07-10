#!/usr/bin/env python3
# push_followups_today.py
# -----------------------------------------------------------------------------
# One-way mirror of the Follow-Up Tracker's daily "Staff_Action_Today_*.xlsx"
# into two tabs of the Clinic Callback Tracker Google Sheet, so the dashboard
# can show a live "Today's follow-up calls" worklist with one-tap calling.
#
#   Call Sheet  tab  ->  "Followups_Today"     (the worklist the staff call)
#   Settled     tab  ->  "Followups_Settled"   (returned / dropout history)
#
# It is a deliberate TWIN of push_patient_mirror.py — same safety model:
#   - Reads the workbook READ-ONLY. Never writes to your local file.
#   - Pushes UP only (local -> cloud), with ONE deliberate exception (D194):
#     during a LIVE push it READS the "Do_Not_Call" tab of the same sheet and
#     drops any worklist row whose mobile is listed there, BEFORE writing.
#     That tab is maintained by humans in the sheet; this script never writes
#     or creates it. If the tab is missing you get a loud warning every run;
#     if the tab exists but cannot be read, the push STOPS (never push a list
#     that skipped the do-not-call check).
#   - Replace-only: rewrites each tab fresh every run.
#   - Refuses to push an empty Call Sheet (so a bad export can't wipe the list).
#   - Console output MASKS phone numbers. The full 10-digit number is written
#     to the sheet (the dashboard needs it to place the call), exactly as the
#     patient mirror already does for Patient_Master.
#   - Outcomes (who called / what happened) are NOT touched here. They live in
#     a separate "Followup_Outcomes" tab the dashboard writes, so re-running
#     this in the morning never wipes a result.
#
# SESSION 29 CHANGE (follow-up sub-sections, Decision D47):
#   The "Section" column is now derived from the tracker's own STATUS value
#   (the part before the " · " hint), so the dashboard groups follow-ups into
#   Due Today / Grace Period / Actionable Missed Follow-Up / Probable Dropout
#   — EXACTLY the buckets on the staff Call Sheet. Names/periods can never
#   drift, because we reuse the tracker's status instead of recomputing days.
#   The full STATUS (with its "· said call later…" hint) is left untouched in
#   the Status column. The Procedure block keeps its own "Procedure" label.
#
# HOW TO RUN (staff PC, once each morning — same habit as the patient mirror):
#   python push_followups_today.py            <-- PREVIEW (writes nothing)
#   python push_followups_today.py --push     <-- LIVE push to the sheet
#   python push_followups_today.py --file "C:\path\Staff_Action_Today_2026-06-29.xlsx" --push
# -----------------------------------------------------------------------------

import os
import sys
import glob
import json
import datetime

# --- openpyxl is the only extra dependency (same one the tracker already uses)
try:
    import openpyxl
except ImportError:
    print("STOP: the 'openpyxl' library isn't installed. Run once:")
    print("   pip install --upgrade openpyxl")
    sys.exit(1)

# =============================== CONFIG ======================================
# Normally nothing here needs changing.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Where the tracker drops its daily workbook. Default: an "outputs" subfolder
# next to this script (the tracker's OUTPUTS_DIR). Override with --file.
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")

# Folder holding the service-account .json key (same key as the patient mirror).
KEY_DIR = BASE_DIR

# The Clinic Callback Tracker sheet + the two tabs the dashboard reads.
SHEET_ID       = "1USjArkqIdrE9hIqerghms76STatM5XTbSW_a9I3klo0"
TAB_TODAY      = "Followups_Today"
TAB_DNC        = "Do_Not_Call"     # human-maintained; this script READS it only
TAB_SETTLED    = "Followups_Settled"
# =============================================================================

TODAY_HEADERS = ["Key", "Section", "PR", "Patient Name", "Mobile",
                 "Diagnosis", "Due Date", "OD", "Status"]
SETTLED_HEADERS = ["Due", "Patient", "Mobile", "Clinic ID", "Outcome",
                   "Handled By", "When"]


# ------------------------------- helpers ------------------------------------
def norm(s):
    return "".join(ch for ch in str(s or "").lower() if ch.isalnum())


def mask_phone(p):
    p = "".join(ch for ch in str(p or "") if ch.isdigit())
    return ("..." + p[-4:]) if len(p) >= 4 else ("..." if p else "")


def normalize_phone(raw):
    """Return a clean 10-digit Indian mobile, or '' if not valid."""
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if len(digits) > 10:
        digits = digits[-10:]
    if len(digits) == 10 and digits[0] in "6789":
        return digits
    return ""


def bucket_from_status(status):
    """Return the clean follow-up bucket name from a STATUS cell.

    The tracker writes the bucket, then optionally a ' · ' and an actionable
    hint (e.g. 'Actionable Missed Follow-Up  ·  said call later — call now').
    We keep ONLY the bucket (the text before the first middot), so the
    dashboard groups on a stable label:
        'Due Today'                          -> 'Due Today'
        'Grace Period'                       -> 'Grace Period'
        'Actionable Missed Follow-Up  ·  ..' -> 'Actionable Missed Follow-Up'
        'Probable Dropout'                   -> 'Probable Dropout'
        'Invalid Mobile / No Contact'        -> 'Invalid Mobile / No Contact'
        'Identity Unresolved'                -> 'Identity Unresolved'
    The full STATUS text (with the hint) is preserved separately in the Status
    column, so the actionable detail is never lost.
    """
    s = "" if status is None else str(status).strip()
    if not s:
        return ""
    head = s.split("\u00b7")[0]     # \u00b7 = the ' · ' middot separator
    return head.strip()


# --- date normalizer --------------------------------------------------------
# The tracker writes the per-row date as yearless text (e.g. "30-Jun"). When the
# dashboard formats that, it can't find a year and falls back to "2001". We fix
# it HERE, at the source, so the sheet always carries a full "DD-Mon-YYYY" date.
# This is defensive: it accepts a real Excel date object, yearless text, a full
# date in several formats, or junk — and never raises.
_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def to_full_date(raw, default_year=None):
    """Normalize any date cell value to 'DD-Mon-YYYY' (e.g. '30-Jun-2026').

    Handles:
      - a real datetime/date object  -> formatted directly (year already right)
      - yearless text like '30-Jun'   -> current year attached
      - full text '30-Jun-2026', '30/06/2026', '2026-06-30' -> parsed & reformatted
      - anything unparseable           -> returned as a trimmed string, unchanged
    Never raises. Blank stays blank.
    """
    if default_year is None:
        default_year = datetime.date.today().year

    # 1) real date/datetime object straight from Excel
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return raw.strftime("%d-%b-%Y")

    s = "" if raw is None else str(raw).strip()
    if not s:
        return ""

    # 2) try a set of explicit formats (full dates first)
    fmts_full = ("%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%d-%m-%Y",
                 "%Y-%m-%d", "%Y/%m/%d", "%d %b %Y", "%d %B %Y")
    for f in fmts_full:
        try:
            return datetime.datetime.strptime(s, f).strftime("%d-%b-%Y")
        except ValueError:
            pass

    # 3) yearless forms like '30-Jun', '30 Jun', '30/06', '30-6'
    #    parse day + month ourselves, attach default_year.
    cleaned = s.replace("/", "-").replace(" ", "-")
    parts = [p for p in cleaned.split("-") if p != ""]
    if len(parts) == 2:
        d_str, m_str = parts[0], parts[1]
        try:
            day = int(d_str)
        except ValueError:
            day = None
        month = None
        ml = m_str.lower()[:3]
        if ml in _MONTHS:
            month = _MONTHS[ml]
        else:
            try:
                mi = int(m_str)
                if 1 <= mi <= 12:
                    month = mi
            except ValueError:
                month = None
        if day and month and 1 <= day <= 31:
            try:
                return datetime.date(default_year, month, day).strftime("%d-%b-%Y")
            except ValueError:
                pass

    # 4) give up safely: hand back the original trimmed text
    return s


def find_workbook():
    """Pick the workbook: --file if given, else the newest Staff_Action_Today_*."""
    if "--file" in sys.argv:
        i = sys.argv.index("--file")
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    cands = glob.glob(os.path.join(OUTPUTS_DIR, "Staff_Action_Today_*.xlsx"))
    if not cands:
        # also try this script's own folder, just in case
        cands = glob.glob(os.path.join(BASE_DIR, "Staff_Action_Today_*.xlsx"))
    if not cands:
        return None
    return max(cands, key=os.path.getmtime)   # newest by modified time


def header_index(ws, must_have, scan_rows=12):
    """Find the header row (1-based) containing a cell whose normalized text
    matches one of `must_have`. Return (row_index, {field: col_index})."""
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        cells = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(m in cells for m in must_have):
            return r, cells
    return None, None


def col_of(header_cells, *aliases):
    for a in aliases:
        a = norm(a)
        if a in header_cells:
            return header_cells.index(a)  # 0-based within row
    return -1


# ----------------------------- Call Sheet -----------------------------------
def read_call_sheet(wb):
    ws = wb["Call Sheet"]
    hrow, H = header_index(ws, ["patientname"])
    if not hrow:
        return []
    iSN  = col_of(H, "s.n", "sn")
    iPR  = col_of(H, "pr")
    iNm  = col_of(H, "patient name", "name")
    iMob = col_of(H, "mobile", "mobile no")
    iDx  = col_of(H, "diagnosis / info", "diagnosis", "info")
    iDt  = col_of(H, "date", "due", "due date")
    iOD  = col_of(H, "od")
    iSt  = col_of(H, "status")
    iKey = col_of(H, "key")

    def cell(row, idx):
        if idx < 0:
            return ""
        v = ws.cell(row, idx + 1).value
        return "" if v is None else str(v).strip()

    def raw_cell(row, idx):
        # untouched value (keeps real date objects intact for to_full_date)
        if idx < 0:
            return None
        return ws.cell(row, idx + 1).value

    out, seen = [], set()
    block = "Follow-up"           # which block of the sheet we're in
    for r in range(hrow + 1, ws.max_row + 1):
        a    = cell(r, iSN)
        name = cell(r, iNm)
        # ---- detect a section banner / note row (no patient name on it) ----
        if not name:
            banner = " ".join(cell(r, c) for c in range(0, 3) if cell(r, c)).lower()
            if "procedure" in banner:
                block = "Procedure"
            elif "follow" in banner:
                block = "Follow-up"
            continue                                   # skip all non-patient rows
        if norm(name) in ("total", "drmanojagarwalclinic"):
            continue
        key = cell(r, iKey) or ("ROW%d" % r)
        if key in seen:
            continue
        seen.add(key)

        status = cell(r, iSt)
        # ---- Section = the tracker's own bucket (Decision D47) --------------
        # Follow-up rows group by their STATUS bucket (Due Today / Grace Period
        # / Actionable Missed Follow-Up / Probable Dropout / etc.). The
        # Procedure block keeps its own label. Full STATUS stays in column 9.
        if block == "Procedure":
            section = "Procedure"
        else:
            section = bucket_from_status(status) or "Follow-up"

        out.append([
            key,
            section,
            cell(r, iPR),
            name,
            normalize_phone(cell(r, iMob)),            # '' if uncallable (dashboard hides Call)
            cell(r, iDx),
            to_full_date(raw_cell(r, iDt)),            # always full DD-Mon-YYYY (fixes "2001")
            cell(r, iOD),
            status,
        ])
    return out


# --------------------------- Settled Follow-Ups -----------------------------
def read_settled(wb):
    if "Settled Follow-Ups" not in wb.sheetnames:
        return []
    ws = wb["Settled Follow-Ups"]
    hrow, H = header_index(ws, ["outcome", "patient"])
    if not hrow:
        return []
    iDue = col_of(H, "due")
    iPt  = col_of(H, "patient")
    iMob = col_of(H, "mobile")
    iCid = col_of(H, "clinic id", "clinicid")
    iOut = col_of(H, "outcome")
    iBy  = col_of(H, "handled by", "handledby")
    iWhn = col_of(H, "when")

    def cell(row, idx):
        if idx < 0:
            return ""
        v = ws.cell(row, idx + 1).value
        return "" if v is None else str(v).strip()

    def raw_cell(row, idx):
        if idx < 0:
            return None
        return ws.cell(row, idx + 1).value

    out = []
    for r in range(hrow + 1, ws.max_row + 1):
        pt = cell(r, iPt)
        if not pt:
            continue
        out.append([
            to_full_date(raw_cell(r, iDue)), pt, normalize_phone(cell(r, iMob)),
            cell(r, iCid), cell(r, iOut), cell(r, iBy), to_full_date(raw_cell(r, iWhn)),
        ])
    return out


# ------------------------------ Do-Not-Call ----------------------------------
def fetch_dnc_set(sh):
    """Read the human-maintained Do_Not_Call tab. Return a set of clean
    10-digit numbers. Safety contract (Decision D194):
      - tab MISSING            -> loud warning, empty set (push continues)
      - tab present, read FAILS-> STOP the push (never skip the check silently)
      - a row whose Phone cell can't be normalized -> reported (masked), skipped
    This is the ONLY place the script reads the sheet, and it never writes
    or creates this tab."""
    import gspread
    try:
        ws = sh.worksheet(TAB_DNC)
    except gspread.WorksheetNotFound:
        print("\n*** WARNING: no '%s' tab exists in the sheet. ***" % TAB_DNC)
        print("*** The do-not-call filter is INACTIVE until you create it.  ***")
        return set()
    try:
        col = ws.col_values(1)                     # column A = Phone
    except Exception as e:
        print("\nSTOP: could not read the '%s' tab (%s)." % (TAB_DNC, type(e).__name__))
        print("Refusing to push a worklist that skipped the do-not-call check.")
        print("Fix the connection (or the tab) and run again.")
        sys.exit(1)
    dnc, bad = set(), 0
    for v in col[1:]:                              # skip the header row
        v = str(v or "").strip()
        if not v:
            continue
        p = normalize_phone(v)
        if p:
            dnc.add(p)
        else:
            bad += 1
            print("   Do_Not_Call: could not read a number ending '%s' - row skipped,"
                  " please fix it in the sheet." % mask_phone(v))
    print("Do-not-call list loaded  : %d number(s)%s"
          % (len(dnc), ("  (%d unreadable row(s) ignored)" % bad) if bad else ""))
    return dnc


# --------------------------------- write ------------------------------------
def write_tab(sh, title, headers, rows):
    import gspread
    try:
        ws = sh.worksheet(title)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=len(rows) + 10, cols=len(headers))
        print("Created new tab: " + title)
    values = [headers] + rows
    ws.resize(rows=max(len(values), 1), cols=len(headers))
    ws.update(values=values, range_name="A1", value_input_option="RAW")


def find_key():
    keys = []
    for path in glob.glob(os.path.join(KEY_DIR, "*.json")):
        try:
            with open(path, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if obj.get("type") == "service_account" and obj.get("client_email"):
                keys.append(path)
        except Exception:
            pass
    return keys


# --------------------------------- main -------------------------------------
def main():
    push = "--push" in sys.argv
    print("=" * 64)
    print("FOLLOW-UP INGEST  -  " + ("LIVE PUSH" if push else "PREVIEW (no write)"))
    print("=" * 64)

    wbpath = find_workbook()
    if not wbpath or not os.path.exists(wbpath):
        print("STOP: could not find a 'Staff_Action_Today_*.xlsx'.")
        print("Looked in: " + OUTPUTS_DIR)
        print("Pass one explicitly with:  --file \"<path to the workbook>\"")
        sys.exit(1)
    print("Workbook : " + os.path.basename(wbpath))

    wb = openpyxl.load_workbook(wbpath, data_only=True)
    today = read_call_sheet(wb)
    settled = read_settled(wb)

    callable_n = sum(1 for r in today if r[4])
    print("\nWorklist rows (Call Sheet) : %d   (%d callable, %d without a valid mobile)"
          % (len(today), callable_n, len(today) - callable_n))
    # quick section + status breakdown
    from collections import Counter
    sec = Counter(r[1] for r in today)
    print("By section : " + ", ".join("%s=%d" % (k, v) for k, v in sec.items()))
    print("Settled rows               : %d" % len(settled))

    print("\nSample worklist (phone masked):")
    print("   " + " | ".join(TODAY_HEADERS))
    for row in today[:4]:
        shown = row[:4] + [mask_phone(row[4])] + row[5:]
        print("   " + " | ".join(str(x) for x in shown))

    if not today:
        print("\nSTOP: the Call Sheet had zero patient rows. Refusing to push.")
        sys.exit(1)

    if not push:
        print("\nThis was a PREVIEW - nothing was written.")
        print("(The Do_Not_Call filter runs at push time; listed numbers")
        print(" will be removed from the worklist before it is written.)")
        print("If it looks right, run again with:  --push")
        return

    # ----------------------------- live push ---------------------------------
    try:
        import gspread  # noqa: F401
    except ImportError:
        print("\nSTOP: 'gspread' isn't installed. Run once:  pip install --upgrade gspread")
        sys.exit(1)
    keys = find_key()
    if len(keys) != 1:
        print("\nSTOP: need exactly one service-account .json in:\n   " + KEY_DIR)
        for k in keys:
            print("   found: " + os.path.basename(k))
        sys.exit(1)
    import gspread
    print("\nUsing key file: " + os.path.basename(keys[0]))
    gc = gspread.service_account(filename=keys[0])
    sh = gc.open_by_key(SHEET_ID)

    # ---- Do-not-call filter (Decision D194) --------------------------------
    dnc = fetch_dnc_set(sh)
    if dnc:
        kept = [r for r in today if r[4] not in dnc]
        dropped = [r for r in today if r[4] in dnc]
        if dropped:
            print("Do-not-call filter removed %d row(s):" % len(dropped))
            for r in dropped:
                print("   - %s (%s)" % (r[3], mask_phone(r[4])))
        today = kept

    write_tab(sh, TAB_TODAY, TODAY_HEADERS, today)
    write_tab(sh, TAB_SETTLED, SETTLED_HEADERS, settled)
    print("\nDONE. Wrote %d worklist rows to '%s' and %d to '%s'. (Replace-only.)"
          % (len(today), TAB_TODAY, len(settled), TAB_SETTLED))
    print("The dashboard will show today's follow-up calls on its next refresh.")


if __name__ == "__main__":
    main()
