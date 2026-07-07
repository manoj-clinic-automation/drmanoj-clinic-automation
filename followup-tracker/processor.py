"""
Docterz EMR Follow-Up Dropout Tracking System
Core Processing Engine
Dr. Manoj Agarwal Clinic, Bareilly
"""

import pandas as pd
import numpy as np
import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path

# ── Paths ────────────────────────────────────────────────────────────────────
DATA_DIR     = Path(__file__).parent / "data"
OUTPUTS_DIR  = Path(__file__).parent / "outputs"
BACKUP_DIR   = DATA_DIR / "backups"
MASTER_FILE  = DATA_DIR / "patient_master.csv"
VISITS_FILE  = DATA_DIR / "visit_ledger.csv"
FOLLOWUP_FILE= DATA_DIR / "followup_ledger.csv"
DIAGNOSIS_FILE = DATA_DIR / "patient_diagnosis.csv"   # built by seed_diagnosis.py
VACATION_FILE  = DATA_DIR / "vacation_calendar.csv"   # editable from admin page
CALL_LOG_FILE  = DATA_DIR / "call_log.csv"            # Phase 2: staff call outcomes

VACATION_COLS = ["Start_Date", "End_Date", "Slot", "Note"]

def load_vacations() -> list:
    """Return list of {'start': date, 'end': date, 'slot': str, 'note': str}.
    Slot is 'Full Day', 'Morning' or 'Evening' — informational, used to word
    the unavailability notice. Missing/invalid rows are skipped silently."""
    if not VACATION_FILE.exists():
        return []
    try:
        d = pd.read_csv(VACATION_FILE, dtype=str).fillna("")
    except Exception:
        return []
    out = []
    for _, r in d.iterrows():
        s = parse_date(r.get("Start_Date"))
        e = parse_date(r.get("End_Date")) or s
        if not s:
            continue
        if e < s:
            s, e = e, s
        out.append({
            "start": s, "end": e,
            "slot": (str(r.get("Slot", "")).strip() or "Full Day"),
            "note": str(r.get("Note", "")).strip(),
        })
    return out

def save_vacations(rows: list):
    """rows = list of dicts with Start_Date/End_Date/Slot/Note strings."""
    pd.DataFrame(rows, columns=VACATION_COLS).to_csv(VACATION_FILE, index=False)

def vacation_notice_for(due: date, vacations: list) -> str:
    """If due falls inside any vacation window, return a notice label like
    '12-06-2026 to 15-06-2026 (Full Day)'. Empty string otherwise."""
    if not due:
        return ""
    for v in vacations:
        if v["start"] <= due <= v["end"]:
            return f"{date_to_str(v['start'])} to {date_to_str(v['end'])} ({v['slot']})"
    return ""

DATA_DIR.mkdir(exist_ok=True)
OUTPUTS_DIR.mkdir(exist_ok=True)

# How many dated backups of the data/ folder to keep (oldest auto-pruned).
MAX_BACKUPS = 30


def _load_csv_with_schema(path: Path, cols: list) -> pd.DataFrame:
    """Load a ledger CSV and reconcile it to the CURRENT schema.

    This is what keeps pending follow-ups (and all ledger state) alive across
    version updates: if a new version ADDS a column, old rows get it as "" rather
    than crashing on save; if a column was removed, it's dropped harmlessly.
    Existing data is never altered — only the column set is reconciled."""
    if not path.exists():
        return pd.DataFrame(columns=cols)
    df = pd.read_csv(path, dtype=str).fillna("")
    added = [c for c in cols if c not in df.columns]
    for c in added:
        df[c] = ""
    if added:
        print(f"  [schema] {path.name}: added new column(s) {added} to existing ledger")
    return df[cols]


def backup_data(reason: str = "run"):
    """Snapshot every ledger CSV in data/ into a dated backup folder BEFORE a run.
    Pending follow-ups live only in these CSVs, so this is the safety net for
    version updates and bad inputs. Backups beyond MAX_BACKUPS are pruned."""
    import shutil
    csvs = sorted(p for p in DATA_DIR.glob("*.csv"))
    if not csvs:
        return None
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"{stamp}_{reason}"
    dest.mkdir(exist_ok=True)
    for f in csvs:
        shutil.copy2(f, dest / f.name)
    backups = sorted(d for d in BACKUP_DIR.iterdir() if d.is_dir())
    for old in backups[:-MAX_BACKUPS]:
        shutil.rmtree(old, ignore_errors=True)
    print(f"  [backup] data/ snapshot saved to backups/{dest.name} ({len(csvs)} files)")
    return dest


# ── Diagnosis taxonomy → message-category provisions (Phase-later WABA) ──────
# Maps a Standardized_Diagnosis / Taxonomy value to a message category tag.
# NO messages are sent by this system today — the tag simply travels on the
# action sheet so that, when WABA goes live, RA/AS dropouts can receive an
# inflammatory-arthritis contextual bilingual template, fracture patients a
# fracture-review template, etc. Extend this map; nothing else changes.
TAXONOMY_MESSAGE_MAP = {
    "RA":                     "inflammatory_arthritis_bilingual",
    "AS":                     "inflammatory_arthritis_bilingual",
    "RHEUMATOID ARTHRITIS":   "inflammatory_arthritis_bilingual",
    "ANKYLOSING SPONDYLITIS": "inflammatory_arthritis_bilingual",
    "INFLAMMATORY ARTHRITIS": "inflammatory_arthritis_bilingual",
    # "FRACTURE":             "fracture_review",
    # "OA KNEE":              "oa_knee_review",
}

def message_category_for(diag: str, taxonomy: str = "") -> str:
    """Return the WABA message-category tag for a diagnosis (provision only)."""
    for key in (str(taxonomy).strip().upper(), str(diag).strip().upper()):
        if key and key in TAXONOMY_MESSAGE_MAP:
            return TAXONOMY_MESSAGE_MAP[key]
    return ""

def load_diagnosis_lookup():
    """Return (by_uid, by_mobile) lookups from the enriched diagnosis master.

    P0-03 fix: Patient_UID is the clinical identity key. If the diagnosis CSV
    has a Patient_UID column, by_uid is the primary lookup. by_mobile remains
    only as a legacy fallback applied exclusively to rows whose identity was
    resolved with HIGH confidence (unique mobile) — never to shared/ambiguous
    mobiles, where it could attach a relative's diagnosis to the wrong patient.
    Degrades gracefully (empty dicts) when the file is absent."""
    if not DIAGNOSIS_FILE.exists():
        return {}, {}
    try:
        d = pd.read_csv(DIAGNOSIS_FILE, dtype=str).fillna("")
    except Exception:
        return {}, {}
    by_uid, by_mobile = {}, {}
    has_uid_col = "Patient_UID" in d.columns
    for _, r in d.iterrows():
        rec = {
            "diag":     r.get("Standardized_Diagnosis", ""),
            "priority": r.get("Diagnosis_Priority", ""),
            "vip":      str(r.get("Is_VIP", "")) == "True",
            "taxonomy": r.get("Diagnosis_Taxonomy", ""),
        }
        if has_uid_col:
            uid = str(r.get("Patient_UID", "")).strip()
            if uid:
                by_uid[uid] = rec
        mob = str(r.get("Mobile_Clean", "")).strip()
        if mob and mob != NO_MOBILE_PLACEHOLDER:
            # If two diagnosis rows share a mobile, the mobile fallback is
            # unsafe for that number — remove it from the mobile map entirely.
            by_mobile[mob] = None if mob in by_mobile else rec
    by_mobile = {m: rec for m, rec in by_mobile.items() if rec is not None}
    return by_uid, by_mobile

# ── Mobile cleaning ───────────────────────────────────────────────────────────
NO_MOBILE_PLACEHOLDER = "1111111111"  # Docterz placeholder for patients with no mobile

def clean_mobile(raw) -> str:
    """
    Normalize any mobile input to a clean 10-digit Indian number string.
    Returns the cleaned number or '' if invalid.
    1111111111 is a Docterz placeholder for no-mobile patients — treated as blank.
    """
    if pd.isna(raw) or raw is None:
        return ""
    s = str(raw).strip()
    # Remove float artifact (.0)
    s = re.sub(r'\.0+$', '', s)
    # Remove all non-digit characters
    s = re.sub(r'\D', '', s)
    # Strip country code prefix
    if s.startswith('91') and len(s) == 12:
        s = s[2:]
    if s.startswith('0') and len(s) == 11:
        s = s[1:]
    # Docterz no-mobile placeholder — treat as no mobile
    if s == NO_MOBILE_PLACEHOLDER:
        return ""
    # Validate: 10 digits, starts with 6-9
    if len(s) == 10 and s[0] in '6789':
        return s
    return ""

def mobile_status(cleaned: str) -> str:
    if not cleaned:
        return "Invalid Mobile"
    return "Valid"

# ── Patient Master ────────────────────────────────────────────────────────────
MASTER_COLS = [
    "Patient_UID", "Clinic_Specific_Id", "Patient_Name",
    "Mobile_Raw", "Mobile_Clean", "Mobile_Status",
    "First_Seen_Date", "Last_Seen_Date",
    "Mobile_Duplicate_Count", "Identity_Status",
    "Added_From", "Last_Updated"
]

def load_master() -> pd.DataFrame:
    if MASTER_FILE.exists():
        df = pd.read_csv(MASTER_FILE, dtype=str)
        # Ensure all expected columns exist
        for col in MASTER_COLS:
            if col not in df.columns:
                df[col] = ""
        return df[MASTER_COLS]
    return pd.DataFrame(columns=MASTER_COLS)

def save_master(df: pd.DataFrame):
    df[MASTER_COLS].to_csv(MASTER_FILE, index=False)

def recalculate_duplicate_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Recount how many Patient_UIDs share each Mobile_Clean and set Identity_Status."""
    df = df.copy()

    if len(df) == 0:
        return df

    # Ensure Mobile_Duplicate_Count is object dtype to accept string values
    df["Mobile_Duplicate_Count"] = df["Mobile_Duplicate_Count"].astype(object)

    valid = df["Mobile_Clean"].str.len() == 10
    counts = df[valid].groupby("Mobile_Clean")["Patient_UID"].transform("count")
    df.loc[valid, "Mobile_Duplicate_Count"] = counts.astype(str).values
    df.loc[~valid, "Mobile_Duplicate_Count"] = "0"

    def identity(row):
        if row["Mobile_Status"] == "Invalid Mobile" or not row["Mobile_Clean"]:
            return "No/Invalid Mobile"
        if int(row.get("Mobile_Duplicate_Count", 1) or 1) > 1:
            return "Shared Mobile"   # Family/attendant number — not a data error
        return "Unique Mobile"

    df["Identity_Status"] = df.apply(identity, axis=1, result_type="reduce")
    return df

# ── Visit Ledger ──────────────────────────────────────────────────────────────
VISIT_COLS = [
    "Visit_ID", "Visit_Date", "Patient_UID", "Clinic_Specific_Id",
    "Patient_Name", "Mobile_Raw", "Mobile_Clean",
    "Had_Procedure",
    "Source_File", "Processed_On"
]

def load_visits() -> pd.DataFrame:
    return _load_csv_with_schema(VISITS_FILE, VISIT_COLS)

def save_visits(df: pd.DataFrame):
    df[VISIT_COLS].to_csv(VISITS_FILE, index=False)

# ── Accidental-bulk-upload guard ──────────────────────────────────────────────
# Docterz's DEFAULT follow-up export covers ONE MONTH. This tracker is a
# one-day-ahead system: a normal file's follow-ups are all due on a single day.
# If an uploaded file's due dates span more than MAX_FOLLOWUP_SPAN_DAYS, we treat
# it as an accidental month/bulk export and load ONLY the earliest day's rows.
# The remaining days arrive naturally in their own daily exports, so nothing is
# lost. A human-readable notice is recorded in INGEST_NOTICES for the upload page.
MAX_FOLLOWUP_SPAN_DAYS = 3

# After this many days overdue with no return, a follow-up is archived as
# "Expired Unresolved": kept in the ledger (and can still reconcile if the
# patient returns very late), but hidden from Staff Action Today so the
# action sheet never accumulates stale cases.
EXPIRY_DAYS = 60

# Whether a visit dated exactly ON a follow-up's log date counts as a return.
# False (default) = treat it as the prescribing visit, not a return.
COUNT_LOG_DATE_VISIT_AS_RETURN = False
INGEST_NOTICES: list[str] = []

# ── Follow-Up Ledger ──────────────────────────────────────────────────────────
FU_COLS = [
    "Followup_ID", "Followup_Log_Date", "Due_Date",
    "FU_Name_Raw", "FU_Mobile_Raw", "FU_Mobile_Clean",
    "Appointment_ID",
    "Patient_UID_Resolved", "Clinic_Specific_Id_Resolved",
    "Standard_Name", "Identity_Confidence", "Identity_Issue",
    "Matched_Return_Visit_Date", "Matched_Visit_ID", "Return_Delay_Days",
    "Followup_Status", "Days_Overdue", "Suggested_Action",
    "Last_Status_Update",
    # ── Phase 2 call read-back summary (latest state; full history in call_log.csv) ──
    "Last_Response", "Last_Called_By", "Last_Call_Date",
    "Call_Attempts", "Call_Resolution",
]

def load_followups() -> pd.DataFrame:
    return _load_csv_with_schema(FOLLOWUP_FILE, FU_COLS)

def save_followups(df: pd.DataFrame):
    df[FU_COLS].to_csv(FOLLOWUP_FILE, index=False)

# ── Staff call-sheet vocabulary ───────────────────────────────────────────────
# RESPONSE dropdown matches the words staff already use in FOLLOWUP LIST PT.xlsx.
# CALLED BY dropdown is the four callers. PROCEDURE_CALLER is pre-filled on the
# procedure call-back tab. The call-log read-back (Phase 2) keys off Followup_ID
# (follow-up calls) and Visit_ID (procedure call-backs), carried as a hidden col.
CALL_RESPONSES = [
    "YES", "NO", "NOT PICK", "LATER", "SWITCH OFF",
    "DIKHA CHUKE", "MED HERE", "MED OUTSIDE", "Tt COMPLETE",
]
CALL_STAFF = ["Shivani", "Alisha", "Shavez", "Ranjeet"]
PROCEDURE_CALLER = "Shavez"
DEFAULT_CONSULT_FEE = 600   # standard consultation fee; flags ▲/▼ on Day Revenue tab

# ── Phase 2: staff call read-back (call_log) ──────────────────────────────────
# One row per call ATTEMPT (a follow-up may be called several times). Staff fill
# RESPONSE + CALLED BY on the Call Sheet; on upload we capture them here, keyed by
# the hidden KEY column — Followup_ID (F######) for follow-ups, Visit_ID (V######)
# for procedure call-backs. This ledger lives in data\ only — PHI never to Google.
#
# CAPTURE-ONLY at this stage: outcomes are recorded and the ledger summary columns
# are refreshed, but the GENERATED call sheet is not yet changed by them. The
# suppress/annotate/escalate overlay on the sheet is the next build step.
CALL_LOG_COLS = [
    "Call_ID", "Key_Type", "Obligation_Key", "Patient_UID",
    "Call_Date", "Response", "Called_By", "Call_Time", "Remarks",
    "Attempt_No", "Resolution_Class", "Source_Sheet", "Ingested_On",
]

# RESPONSE → what to do next. Reviewed & approved by Dr. Manoj, 22-Jun-2026.
RESOLUTION_MAP = {
    "YES":          "RESOLVE",     # reached, positive / coming — stop calling
    "Tt COMPLETE":  "RESOLVE",     # treatment complete
    "MED HERE":     "RESOLVE",     # bought medicine HERE — corroborable vs pharmacy
    "MED OUTSIDE":  "RESOLVE",     # bought medicine outside — self-owned, no check
    "MEDICIN LI H": "RESOLVE",     # legacy (pre-split) — taking medicine, self-managing
    "DIKHA CHUKE":  "RESOLVE",     # already consulted / seen
    "NO":           "DECLINE",     # declined / not coming — stop, reason logged
    "LATER":        "RESCHEDULE",  # call again after RESCHEDULE_DAYS
    "NOT PICK":     "RETRY",       # no answer — call again next call day
    "SWITCH OFF":   "RETRY",       # phone off — call again next call day
}
RESCHEDULE_DAYS = 2     # LATER → re-surface after this many days
ESCALATE_AFTER  = 3     # this many consecutive RETRYs (no contact) → unreachable
LATER_ESCALATE_AFTER = 3  # this many consecutive LATERs (perpetual deferral) → escalate
SETTLED_WINDOW_DAYS = 14   # 'Settled Follow-Ups' tab shows closes due within this window
PROMISE_GRACE_DAYS = 3  # a 'YES/coming' that hasn't produced a real return within
                        # this many days re-surfaces for a re-confirm call
EXPECTS_RETURN_RESPONSES = {"YES"}   # responses that promise a return visit and
                                     # must be verified against the consultation report

def resolution_class_for(response: str) -> str:
    return RESOLUTION_MAP.get(str(response).strip(), "")

def load_call_log() -> pd.DataFrame:
    return _load_csv_with_schema(CALL_LOG_FILE, CALL_LOG_COLS)

def save_call_log(df: pd.DataFrame):
    df[CALL_LOG_COLS].to_csv(CALL_LOG_FILE, index=False)

def _parse_callsheet_date(ws, fallback: date) -> date:
    """The call sheet's row-2 banner reads 'Date: 22-Jun-2026  Pick RESPONSE …'.
    Pull that date so every captured call is stamped with the clinic day it was
    made on, regardless of when the file is uploaded."""
    for r in range(1, 4):
        v = ws.cell(r, 1).value
        if v and "Date:" in str(v):
            m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", str(v))
            if m:
                try:
                    return datetime.strptime(m.group(1), "%d-%b-%Y").date()
                except ValueError:
                    pass
    return fallback

def _finalize_call_log(log: pd.DataFrame) -> pd.DataFrame:
    """Assign stable Call_IDs to new rows and (re)number Attempt_No per obligation
    in chronological order. Resolution_Class is filled from the response map."""
    log = log.copy()
    have = log["Call_ID"].astype(str).str.strip()
    nums = [int(x[2:]) for x in have if x.startswith("CL") and x[2:].isdigit()]
    nextn = (max(nums) + 1) if nums else 1
    new_ids = []
    for v in have:
        if v.startswith("CL"):
            new_ids.append(v)
        else:
            new_ids.append(f"CL{nextn:06d}"); nextn += 1
    log["Call_ID"] = new_ids
    log["_d"] = log["Call_Date"].map(lambda s: parse_date(s) or date(1900, 1, 1))
    log = log.sort_values(["Obligation_Key", "_d", "Call_ID"]).reset_index(drop=True)
    log["Attempt_No"] = (log.groupby("Obligation_Key").cumcount() + 1).astype(str)
    log["Resolution_Class"] = log.apply(
        lambda r: (str(r["Resolution_Class"]).strip() or resolution_class_for(r["Response"])),
        axis=1)
    return log.drop(columns=["_d"])

def ingest_call_sheet(filepath, fu_ledger=None, visits=None, source_name=None) -> dict:
    """Read a FILLED Staff Call Sheet (.xlsx) and capture RESPONSE / CALLED BY into
    call_log.csv.

    Robust to the sheet's two-section layout: it locates columns by their header
    labels and reads any row that has a KEY and a non-blank RESPONSE, so section
    titles, header rows and '— none —' lines are skipped automatically.

    Idempotent: a call is identified by Obligation_Key + Call_Date + Response +
    Called_By, so uploading the same sheet twice adds nothing.

    Returns {'read', 'new', 'duplicate', 'date'} (or {'error': ...})."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    sheet = "Call Sheet" if "Call Sheet" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    col = {}
    for r in range(1, min(ws.max_row, 12) + 1):
        labels = {str(ws.cell(r, c).value).strip().upper(): c
                  for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "KEY" in labels and "RESPONSE" in labels:
            for lab, k in (("KEY", "key"), ("RESPONSE", "resp"), ("CALLED BY", "by"),
                           ("CALL TIME", "time"), ("REMARKS", "rem")):
                if lab in labels:
                    col[k] = labels[lab]
            break
    if "key" not in col or "resp" not in col:
        return {"read": 0, "new": 0, "duplicate": 0, "date": None,
                "error": "could not find KEY/RESPONSE columns — is this a Call Sheet?"}

    call_date = _parse_callsheet_date(ws, date.today())

    fu_uid, v_uid = {}, {}
    if fu_ledger is not None and len(fu_ledger):
        fu_uid = dict(zip(fu_ledger["Followup_ID"].astype(str),
                          fu_ledger["Patient_UID_Resolved"].astype(str)))
    if visits is not None and len(visits):
        v_uid = dict(zip(visits["Visit_ID"].astype(str),
                         visits["Patient_UID"].astype(str)))

    def _cell(r, k):
        c = col.get(k)
        if not c:
            return ""
        v = ws.cell(r, c).value
        return "" if v is None else str(v).strip()

    rows = []
    for r in range(1, ws.max_row + 1):
        key = _cell(r, "key")
        if not key or key.upper() == "KEY":
            continue
        resp = _cell(r, "resp")
        if not resp:
            continue
        ktype = "Followup" if key.upper().startswith("F") else "Procedure"
        uid = fu_uid.get(key, "") if ktype == "Followup" else v_uid.get(key, "")
        rows.append({
            "Key_Type": ktype, "Obligation_Key": key, "Patient_UID": uid,
            "Call_Date": date_to_str(call_date), "Response": resp,
            "Called_By": _cell(r, "by"), "Call_Time": _cell(r, "time"),
            "Remarks": _cell(r, "rem"),
        })

    log = load_call_log()
    existing = set(zip(log["Obligation_Key"], log["Call_Date"],
                       log["Response"], log["Called_By"])) if len(log) else set()
    src = source_name or Path(filepath).name
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_rows, dup = [], 0
    for rec in rows:
        sig = (rec["Obligation_Key"], rec["Call_Date"], rec["Response"], rec["Called_By"])
        if sig in existing:
            dup += 1
            continue
        existing.add(sig)
        rec.update({"Call_ID": "", "Attempt_No": "",
                    "Resolution_Class": resolution_class_for(rec["Response"]),
                    "Source_Sheet": src, "Ingested_On": now})
        new_rows.append(rec)

    if new_rows:
        log = pd.concat([log, pd.DataFrame(new_rows)], ignore_index=True, sort=False)
        log = _finalize_call_log(log)
        save_call_log(log)

    return {"read": len(rows), "new": len(new_rows), "duplicate": dup,
            "date": date_to_str(call_date)}

def recompute_call_summary(fu_ledger: pd.DataFrame, call_log: pd.DataFrame) -> pd.DataFrame:
    """Fold the latest call outcome per follow-up into the ledger's summary columns
    (latest state only; full history stays in call_log). Escalation: ESCALATE_AFTER
    consecutive trailing RETRYs flips Call_Resolution to 'ESCALATE'. Follow-ups with
    no calls keep whatever they had. Procedure (V-keyed) calls are ignored here."""
    fu = fu_ledger.copy()
    summary_cols = ["Last_Response", "Last_Called_By", "Last_Call_Date",
                    "Call_Attempts", "Call_Resolution"]
    for c in summary_cols:
        if c not in fu.columns:
            fu[c] = ""
    if call_log is None or not len(call_log):
        return fu
    cl = call_log[call_log["Key_Type"].astype(str) == "Followup"].copy()
    if not len(cl):
        return fu
    cl["_d"] = cl["Call_Date"].map(lambda s: parse_date(s) or date(1900, 1, 1))
    cl = cl.sort_values(["Obligation_Key", "_d", "Call_ID"])
    summ = {}
    for key, grp in cl.groupby("Obligation_Key"):
        last = grp.iloc[-1]
        run = 0
        for cls in reversed(list(grp["Resolution_Class"])):
            if str(cls) == "RETRY":
                run += 1
            else:
                break
        later_run = 0
        for cls in reversed(list(grp["Resolution_Class"])):
            if str(cls) == "RESCHEDULE":
                later_run += 1
            else:
                break
        res = str(last["Resolution_Class"])
        if res == "RETRY" and run >= ESCALATE_AFTER:
            res = "ESCALATE"
        elif res == "RESCHEDULE" and later_run >= LATER_ESCALATE_AFTER:
            res = "ESCALATE"   # perpetual 'call me later' — stop deferring, escalate
        summ[str(key)] = [str(last["Response"]), str(last["Called_By"]),
                          str(last["Call_Date"]), str(len(grp)), res]
    keys = fu["Followup_ID"].astype(str)
    for i, name in enumerate(summary_cols):
        fu[name] = [summ[k][i] if k in summ else fu.at[idx, name]
                    for idx, k in zip(fu.index, keys)]
    return fu

def call_monitor_stats(call_date: str | None = None) -> dict:
    """Summarise the call log for one clinic day (default: the most recent day
    with calls). 'Reached' = any response other than NOT PICK / SWITCH OFF.
    'Pending' = actionable follow-ups with no captured outcome yet."""
    NO_CONTACT = {"NOT PICK", "SWITCH OFF"}
    out = {"date": call_date or "", "total": 0, "reached": 0, "not_reached": 0,
           "contact_rate": 0, "by_response": [], "by_staff": [], "pending": 0,
           "all_dates": []}
    log = load_call_log()
    if len(log):
        dates = sorted({d for d in log["Call_Date"].astype(str) if d}, reverse=True)
        out["all_dates"] = dates
        if not call_date:
            call_date = dates[0] if dates else ""
        out["date"] = call_date
        day = log[log["Call_Date"].astype(str) == call_date]
        out["total"] = len(day)
        if len(day):
            up = day["Response"].astype(str).str.upper()
            reached_mask = ~up.isin({s.upper() for s in NO_CONTACT})
            out["reached"] = int(reached_mask.sum())
            out["not_reached"] = out["total"] - out["reached"]
            out["contact_rate"] = round(100 * out["reached"] / out["total"])
            out["by_response"] = [(str(k), int(v)) for k, v in day["Response"].value_counts().items()]
            bs = []
            for staff, g in day.groupby("Called_By"):
                gu = g["Response"].astype(str).str.upper()
                bs.append((str(staff) or "(blank)", len(g),
                           int((~gu.isin({s.upper() for s in NO_CONTACT})).sum())))
            out["by_staff"] = sorted(bs, key=lambda x: -x[1])
    try:
        fu = load_followups()
        if len(fu):
            actionable = fu[fu["Followup_Status"].isin(
                ["Probable Dropout", "Actionable Missed Follow-Up", "Grace Period", "Due Today"])]
            out["pending"] = int((actionable["Call_Resolution"].astype(str).str.strip() == "").sum())
    except Exception:
        pass
    return out

def call_reconciliation_stats() -> dict:
    """Cross-verify recorded call responses against ACTUAL return visits.

    The follow-up ledger now holds both the phone claim (Last_Response /
    Call_Resolution) and the real return match (Followup_Status 'Returned*',
    written by the visit engine from the consultation report). Joining them turns
    'we made calls' into 'calls that produced real returns'."""
    out = {"called": 0, "called_returned": 0, "call_return_rate": 0,
           "unfulfilled_promise": 0, "self_resolved_no_return": 0,
           "declined": 0, "unreached": 0, "organic_return": 0, "total_returned": 0}
    fu = load_followups()
    if not len(fu):
        return out
    resp = fu["Last_Response"].astype(str).str.strip()
    rescls = fu["Call_Resolution"].astype(str).str.strip()
    returned = fu["Followup_Status"].astype(str).str.startswith("Returned")
    called = resp != ""
    SELF = {"Tt COMPLETE", "MEDICIN LI H", "DIKHA CHUKE"}
    out["total_returned"] = int(returned.sum())
    out["called"] = int(called.sum())
    out["called_returned"] = int((called & returned).sum())
    out["call_return_rate"] = (round(100 * out["called_returned"] / out["called"])
                               if out["called"] else 0)
    out["organic_return"] = int((returned & ~called).sum())
    out["unfulfilled_promise"] = int(((resp == "YES") & ~returned).sum())
    out["self_resolved_no_return"] = int((resp.isin(SELF) & ~returned).sum())
    out["declined"] = int((resp == "NO").sum())
    out["unreached"] = int((rescls == "ESCALATE").sum())
    return out

# ═══════════════════════════════════════════════════════════════════════════
# IVR TELEPHONY RECONCILIATION (Feature A)  ·  added 22-Jun-2026
# Cross-check logged call outcomes (call_log) against MyOperator outbound calls
# (the Outbound_Log tab, exported to CSV/xlsx). ADVISORY ONLY — never changes the
# call sheet or ledger. The join happens locally, so PHI never leaves the clinic.
#
# Agent model (confirmed w/ Dr. Manoj 22-Jun-2026):
#   Shivani / Alisha / Shavez  -> named on IVR            -> claims are verifiable
#   Ranjeet                    -> NOT on IVR (backs up from the reception mobile);
#                                 his claims route via "Reception Mobile" -> the
#                                 agent-identity check is skipped for Ranjeet
#   Manoj Bhati                -> real staff on IVR (off-time); not a CALLED BY option
#   "Reception Mobile"         -> shared handset; confirms a call, caller unidentifiable
# Contact is judged by Status (connected/missed), NEVER by duration (ring time).
# ═══════════════════════════════════════════════════════════════════════════
OUTBOUND_LOG_FILE = DATA_DIR / "outbound_log.csv"
OUTBOUND_LOG_COLS = ["Date", "Time", "Phone10", "Direction", "Agent",
                     "Duration_s", "Status", "Start_Unix"]
FEED_URL_FILE = DATA_DIR / "feed_url.txt"   # one-time: paste the published Call_Feed CSV link here

# Follow-up statuses that mean "this patient is genuinely due / overdue" — used by
# the Called-In worklist (Feature B). Identity-problem statuses are excluded so the
# worklist stays high-confidence.
CALLIN_DUE_STATUSES = {"Probable Dropout", "Actionable Missed Follow-Up",
                       "Grace Period", "Due Today"}

SHARED_AGENT_LABELS  = {"reception mobile"}   # call real, caller not identifiable
CALLERS_NOT_ON_IVR   = {"ranjeet"}            # CALLED BY names absent from the IVR
AGENT_ALIASES        = {}                     # ivr_agent_lower -> caller firstname (exceptions only)
RECONCILE_WINDOW_DAYS = 1                     # match same day ± this many days

# RESPONSE → does the claim imply the patient was actually reached?
REACHED_CLAIM_CLASSES  = {"RESOLVE", "DECLINE"}  # spoke to patient (positive or declined)
NOANSWER_CLAIM_CLASSES = {"RETRY"}               # NOT PICK / SWITCH OFF — no contact claimed

def load_outbound_log() -> pd.DataFrame:
    return _load_csv_with_schema(OUTBOUND_LOG_FILE, OUTBOUND_LOG_COLS)

def save_outbound_log(df: pd.DataFrame):
    df[OUTBOUND_LOG_COLS].to_csv(OUTBOUND_LOG_FILE, index=False)

def _agent_token(ivr_agent) -> str:
    """First-name token for an IVR Agent ('Shivani Srivastava' -> 'shivani').
    Returns '' for a shared/blank label (caller not identifiable)."""
    a = str(ivr_agent or "").strip().lower()
    if not a or a in SHARED_AGENT_LABELS:
        return ""
    if a in AGENT_ALIASES:
        return AGENT_ALIASES[a]
    return a.split()[0]

def _is_shared_agent(ivr_agent) -> bool:
    return str(ivr_agent or "").strip().lower() in SHARED_AGENT_LABELS

def ingest_outbound_log(filepath: str, source_name: str = "") -> dict:
    """Read an Outbound_Log export (.xlsx or .csv) into data/outbound_log.csv.
    Idempotent on Start_Unix (one call = one unix start); re-uploading adds nothing
    new. Returns a small report dict."""
    p = str(filepath)
    try:
        if p.lower().endswith((".xlsx", ".xlsm")):
            new = pd.read_excel(p, dtype=str)
        else:
            new = pd.read_csv(p, dtype=str)
    except Exception as e:
        return {"ok": False, "error": f"could not read file: {e}"}

    new.columns = [str(c).strip() for c in new.columns]
    rename = {}
    for want in OUTBOUND_LOG_COLS:
        for c in new.columns:
            if c.lower() == want.lower():
                rename[c] = want
    new = new.rename(columns=rename)
    miss = [c for c in ["Phone10", "Status"] if c not in new.columns]
    if miss:
        return {"ok": False,
                "error": f"missing column(s): {', '.join(miss)} — is this an Outbound_Log export?"}
    for c in OUTBOUND_LOG_COLS:
        if c not in new.columns:
            new[c] = ""
    new = new[OUTBOUND_LOG_COLS].copy()

    new["Phone10"]    = new["Phone10"].map(clean_mobile)
    new["Status"]     = new["Status"].astype(str).str.strip().str.lower()
    new["Agent"]      = new["Agent"].astype(str).str.strip()
    new["Direction"]  = (new["Direction"].astype(str).str.strip().str.lower()
                         .replace({"nan": "outgoing", "": "outgoing",
                                   "out": "outgoing", "in": "incoming"}))
    new["Date"]       = new["Date"].map(lambda s: date_to_str(parse_date(s)))
    new["Duration_s"] = pd.to_numeric(new["Duration_s"], errors="coerce").fillna(0).astype(int).astype(str)
    new["Start_Unix"] = new["Start_Unix"].astype(str).str.strip().replace({"nan": ""})
    new = new[new["Phone10"].str.len() == 10].reset_index(drop=True)

    def _sig(r):
        su = str(r.get("Start_Unix", "")).strip()
        if su and su.lower() != "nan":
            return ("u", su)
        return ("dtp", r.get("Date", ""), str(r.get("Time", "")), r.get("Phone10", ""))

    cur = load_outbound_log()
    have = set(_sig(r) for _, r in cur.iterrows()) if len(cur) else set()
    add = [r for _, r in new.iterrows() if _sig(r) not in have]
    added = pd.DataFrame(add, columns=OUTBOUND_LOG_COLS) if add else pd.DataFrame(columns=OUTBOUND_LOG_COLS)
    combined = pd.concat([cur, added], ignore_index=True) if len(added) else cur
    save_outbound_log(combined)
    return {"ok": True, "rows_in_file": int(len(new)), "added": int(len(added)),
            "total": int(len(combined)), "dates": sorted(set(new["Date"]) - {""})}

def _mobile_for_obligation(key, fu_ledger, visits) -> str:
    """Resolve a call_log Obligation_Key (F##### / V#####) to the patient's mobile."""
    k = str(key).strip()
    if k.startswith("F"):
        row = fu_ledger[fu_ledger["Followup_ID"].astype(str) == k]
        if len(row):
            return str(row.iloc[0]["FU_Mobile_Clean"]).strip()
    elif k.startswith("V"):
        row = visits[visits["Visit_ID"].astype(str) == k]
        if len(row):
            return str(row.iloc[0]["Mobile_Clean"]).strip()
    return ""

def _agent_cross_check(call_log, ob, fu_mobiles) -> list:
    """Per named caller: outcomes LOGGED vs IVR follow-up calls PLACED (and how many
    connected). IVR side counts only calls to follow-up-patient numbers, so it is
    comparable with the logged outcomes. Reception Mobile is its own bucket."""
    logged = {}
    for _, c in call_log.iterrows():
        t = str(c["Called_By"]).strip().lower()
        if not t:
            continue
        t = t.split()[0]
        logged[t] = logged.get(t, 0) + 1
    ivr = {}; ivr_conn = {}; shared_n = 0; shared_conn = 0
    for _, o in ob.iterrows():
        if str(o["Phone10"]) not in fu_mobiles:
            continue
        conn = str(o["Status"]).lower() == "connected"
        if _is_shared_agent(o["Agent"]):
            shared_n += 1; shared_conn += 1 if conn else 0
            continue
        t = _agent_token(o["Agent"])
        if not t:
            continue
        ivr[t] = ivr.get(t, 0) + 1
        if conn:
            ivr_conn[t] = ivr_conn.get(t, 0) + 1
    out = []
    for n in sorted(set(list(logged.keys()) + list(ivr.keys()))):
        out.append({"agent": n.title(), "logged": logged.get(n, 0),
                    "ivr_calls": ivr.get(n, 0), "ivr_connected": ivr_conn.get(n, 0),
                    "on_ivr": (n not in CALLERS_NOT_ON_IVR)})
    if shared_n:
        out.append({"agent": "Reception Mobile (shared)", "logged": "—",
                    "ivr_calls": shared_n, "ivr_connected": shared_conn, "on_ivr": True})
    return out

def reconcile_calls(call_log=None, fu_ledger=None, visits=None, outbound=None) -> dict:
    """Match each logged call outcome against IVR outbound calls and classify it.
    Returns {'rows', 'counts', 'agent', 'meta'}. Reads the data\\ ledgers when args
    are omitted. Advisory only — nothing is written back."""
    if call_log is None:  call_log = load_call_log()
    if fu_ledger is None: fu_ledger = load_followups()
    if visits is None:    visits = load_visits()
    if outbound is None:  outbound = load_outbound_log()

    FLAGS = ["logged_no_call", "claims_reached_no_connect",
             "claims_noanswer_but_connected", "agent_mismatch",
             "call_not_logged", "consistent"]
    counts = {f: 0 for f in FLAGS}

    ob = outbound.copy()
    ob = ob[ob["Phone10"].astype(str).str.len() == 10]
    if "Direction" in ob.columns:
        dirn = ob["Direction"].fillna("").astype(str).str.strip().str.lower()
        ob = ob[dirn.isin(["", "nan", "outgoing", "out"])]
    ob = ob.reset_index(drop=True)
    if len(ob):
        ob["_d"] = ob["Date"].map(parse_date)
    else:
        ob["_d"] = pd.Series([], dtype=object)

    if not len(call_log):
        return {"rows": [], "counts": counts, "agent": [],
                "meta": {"calls": 0, "outbound": int(len(ob))}}

    if not len(ob):
        # No telephony loaded yet — don't flag every logged call as "no IVR call".
        return {"rows": [], "counts": counts, "agent": [],
                "meta": {"calls": int(len(call_log)), "outbound": 0, "no_outbound": True}}

    by_phone = {}
    for _, r in ob.iterrows():
        by_phone.setdefault(str(r["Phone10"]), []).append(
            {"d": r["_d"], "status": str(r["Status"]).lower(),
             "agent": str(r["Agent"]), "time": str(r["Time"]),
             "dur": str(r.get("Duration_s", "")).strip()})

    # name lookups
    fu_name = {}
    fu_mobiles = set()
    for _, r in fu_ledger.iterrows():
        m = str(r.get("FU_Mobile_Clean", "")).strip()
        if len(m) == 10:
            fu_mobiles.add(m)
            fu_name.setdefault(m, str(r.get("Standard_Name") or r.get("FU_Name_Raw") or ""))

    rows = []
    covered = set()   # (mobile, day) already accounted for by a logged outcome

    for _, c in call_log.iterrows():
        key = str(c["Obligation_Key"]).strip()
        mob = _mobile_for_obligation(key, fu_ledger, visits)
        cdate = parse_date(c["Call_Date"])
        resp = str(c["Response"]).strip()
        cls = str(c.get("Resolution_Class", "")).strip() or resolution_class_for(resp)
        caller = str(c["Called_By"]).strip()
        caller_l = caller.lower().split()[0] if caller else ""

        matched = []
        if mob and cdate:
            for o in by_phone.get(mob, []):
                if o["d"] and abs((o["d"] - cdate).days) <= RECONCILE_WINDOW_DAYS:
                    matched.append(o)
            for dd in range(-RECONCILE_WINDOW_DAYS, RECONCILE_WINDOW_DAYS + 1):
                covered.add((mob, cdate + timedelta(days=dd)))

        connected_any = any(o["status"] == "connected" for o in matched)
        connected_calls = [o for o in matched if o["status"] == "connected"]
        ivr_time = connected_calls[0]["time"] if connected_calls else ""
        ivr_dur  = connected_calls[0]["dur"] if connected_calls else ""
        named_tokens = set(_agent_token(o["agent"]) for o in matched if not _is_shared_agent(o["agent"]))
        named_tokens.discard("")
        shared_present = any(_is_shared_agent(o["agent"]) for o in matched)
        next_day = bool(matched) and all(o["d"] != cdate for o in matched)

        if not matched:
            flag, note = "logged_no_call", "outcome recorded but no IVR call to this number"
        elif cls in REACHED_CLAIM_CLASSES and not connected_any:
            flag = "claims_reached_no_connect"
            note = f"logged '{resp}' (spoke to patient) but all {len(matched)} attempt(s) missed"
        elif cls in NOANSWER_CLAIM_CLASSES and connected_any:
            flag = "claims_noanswer_but_connected"
            _when = (f" {ivr_dur}s" if ivr_dur else "") + (f" at {ivr_time}" if ivr_time else "")
            note = f"logged '{resp}' (no answer) but a call connected{_when}"
        elif (caller_l and caller_l not in CALLERS_NOT_ON_IVR
              and named_tokens and caller_l not in named_tokens):
            flag = "agent_mismatch"
            note = f"logged by {caller or '—'}, IVR shows {', '.join(sorted(t.title() for t in named_tokens))}"
        else:
            flag = "consistent"
            if caller_l in CALLERS_NOT_ON_IVR and shared_present:
                note = "Ranjeet via reception mobile — call confirmed"
            elif shared_present and not named_tokens:
                note = "via reception mobile — caller not verifiable"
            elif next_day:
                note = "called next day"
            else:
                note = ""

        counts[flag] += 1
        ivr_agents = ", ".join(sorted(t.title() for t in named_tokens))
        if shared_present:
            ivr_agents = (ivr_agents + " + Reception Mobile").strip(" +") if ivr_agents else "Reception Mobile"
        rows.append({
            "key": key, "name": fu_name.get(mob, ""), "mobile": mob,
            "date": str(c["Call_Date"]), "response": resp, "resolution": cls,
            "called_by": caller, "attempts": len(matched), "connected": connected_any,
            "ivr_agents": ivr_agents, "ivr_time": ivr_time, "ivr_dur": ivr_dur,
            "flag": flag, "note": note,
        })

    # ── inverse: IVR calls to follow-up patients with NO logged outcome ──────────
    seen_cnl = set()
    for _, o in ob.iterrows():
        ph = str(o["Phone10"]); od = o["_d"]
        if ph not in fu_mobiles or not od:
            continue
        if (ph, od) in covered:
            continue
        if (ph, od) in seen_cnl:
            continue
        seen_cnl.add((ph, od))
        same = [x for x in by_phone.get(ph, []) if x["d"] == od]
        conn = any(x["status"] == "connected" for x in same)
        agents = sorted(set((_agent_token(x["agent"]) or "reception mobile") for x in same))
        counts["call_not_logged"] += 1
        rows.append({
            "key": "", "name": fu_name.get(ph, ""), "mobile": ph,
            "date": date_to_str(od), "response": "", "resolution": "",
            "called_by": "", "attempts": len(same), "connected": conn,
            "ivr_agents": ", ".join(a.title() for a in agents),
            "flag": "call_not_logged",
            "note": "IVR called this follow-up patient — no outcome logged",
        })

    agent = _agent_cross_check(call_log, ob, fu_mobiles)

    order = {"logged_no_call": 0, "claims_reached_no_connect": 1,
             "claims_noanswer_but_connected": 2, "agent_mismatch": 3,
             "call_not_logged": 4, "consistent": 9}
    rows.sort(key=lambda r: (order.get(r["flag"], 5), str(r["date"])))

    return {"rows": rows, "counts": counts, "agent": agent,
            "meta": {"calls": int(len(call_log)), "outbound": int(len(ob))}}

def find_called_in_and_due(fu_ledger=None, outbound=None) -> dict:
    """Feature B — patients who are due/overdue for follow-up AND have called the
    clinic and been MISSED (incoming, status=missed). High intent: they're trying
    to reach us. Returns {'rows', 'count'}; one row per patient, newest call-in first."""
    if fu_ledger is None: fu_ledger = load_followups()
    if outbound is None:  outbound = load_outbound_log()

    # latest missed-incoming date per phone
    inmissed = {}
    if len(outbound) and "Direction" in outbound.columns:
        for _, o in outbound.iterrows():
            if str(o.get("Direction", "")).strip().lower() not in ("incoming", "in"):
                continue
            if str(o.get("Status", "")).strip().lower() != "missed":
                continue
            ph = str(o.get("Phone10", "")).strip()
            dt = parse_date(o.get("Date"))
            if len(ph) == 10 and dt and (ph not in inmissed or dt > inmissed[ph]):
                inmissed[ph] = dt
    if not inmissed:
        return {"rows": [], "count": 0, "meta": {"no_inbound": True}}

    rows = {}
    for _, r in fu_ledger.iterrows():
        if str(r.get("Followup_Status", "")) not in CALLIN_DUE_STATUSES:
            continue
        m = str(r.get("FU_Mobile_Clean", "")).strip()
        if m not in inmissed:
            continue
        cin = inmissed[m]
        prev = rows.get(m)
        # keep the most-overdue row for a patient; attach the call-in date
        if prev is None or _overdue_num(r) > prev["_ov"]:
            rows[m] = {
                "name": str(r.get("Standard_Name") or r.get("FU_Name_Raw") or ""),
                "mobile": m,
                "due": str(r.get("Due_Date", "")),
                "status": str(r.get("Followup_Status", "")),
                "overdue": str(r.get("Days_Overdue", "")),
                "called_in": date_to_str(cin),
                "_cin": cin, "_ov": _overdue_num(r),
            }
    out = sorted(rows.values(), key=lambda x: (x["_cin"], x["_ov"]), reverse=True)
    for r in out:
        r.pop("_cin", None); r.pop("_ov", None)
    return {"rows": out, "count": len(out)}

def _overdue_num(r) -> int:
    try:
        return int(float(str(r.get("Days_Overdue", "")).strip() or 0))
    except (ValueError, TypeError):
        return 0

# ── Auto-fetch the Call_Feed straight from Google (kills the download/upload step) ──
def feed_url_configured() -> str:
    """Return the published Call_Feed CSV URL from data/feed_url.txt, or '' if unset."""
    try:
        if FEED_URL_FILE.exists():
            return FEED_URL_FILE.read_text(encoding="utf-8").strip()
    except Exception:
        pass
    return ""

def fetch_feed_auto() -> dict:
    """Pull the published Call_Feed CSV and ingest it. No download/upload needed —
    needs data/feed_url.txt with a 'publish to web' CSV link for the Call_Feed tab."""
    url = feed_url_configured()
    if not url:
        return {"ok": False, "error": "no feed link set — create data/feed_url.txt (see notes)."}
    import urllib.request, tempfile
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "followup-tracker"})
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = resp.read()
    except Exception as e:
        return {"ok": False, "error": f"could not reach Google ({e}). Check the link / internet."}
    if b"<html" in data[:600].lower() or b"<!doctype html" in data[:600].lower():
        return {"ok": False, "error": "the link returned a web page, not CSV — re-check the "
                                      "'publish to web → CSV' link for the Call_Feed tab."}
    tmp = Path(tempfile.gettempdir()) / "call_feed_fetch.csv"
    tmp.write_bytes(data)
    rep = ingest_outbound_log(str(tmp), source_name="auto-fetch (Google)")
    if rep.get("ok"):
        rep["fetched"] = True
    return rep

# ── Reinstated callbacks — deterrent for "claims no-answer but connected" ─────
# When a follow-up is logged NOT PICK / SWITCH OFF but the IVR shows the call
# actually CONNECTED, the patient likely had a brief/incomplete conversation and
# the callback was quietly dropped. We REINSTATE the callback on the next sheet,
# named to the original caller. It does NOT clear by re-logging NOT PICK — only
# when the patient is genuinely reached (a later RESOLVE), returns (a visit), or
# the doctor clears it after listening to the recording.
# Deterrent by default; judgement before consequence.
REINSTATE_FILE = DATA_DIR / "reinstatements.csv"
REINSTATE_COLS = ["Reinstate_ID", "Followup_ID", "Patient_UID", "Patient_Name",
                  "Mobile", "Orig_Caller", "Orig_Call_Date", "Connect_Info",
                  "Created_On", "Status", "Cleared_On", "Cleared_By", "Cleared_Reason"]

def load_reinstatements() -> pd.DataFrame:
    return _load_csv_with_schema(REINSTATE_FILE, REINSTATE_COLS)

def save_reinstatements(df: pd.DataFrame):
    df[REINSTATE_COLS].to_csv(REINSTATE_FILE, index=False)

def _next_reinstate_id(df) -> int:
    have = df["Reinstate_ID"].astype(str) if len(df) else []
    nums = [int(x[1:]) for x in have if x.startswith("R") and x[1:].isdigit()]
    return (max(nums) + 1) if nums else 1

def sync_reinstatements(call_log=None, fu_ledger=None, visits=None,
                        outbound=None, today=None) -> dict:
    """Open new reinstatements from 'claims no-answer but connected' findings, and
    auto-clear any where the patient has since been genuinely reached (a later
    RESOLVE) or has returned. Idempotent — a finding is opened once and never
    recreated, and re-logging NOT PICK does not clear it. Returns a summary."""
    if today is None:    today = date.today()
    if call_log is None: call_log = load_call_log()
    if fu_ledger is None: fu_ledger = load_followups()
    if visits is None:   visits = load_visits()
    if outbound is None: outbound = load_outbound_log()

    led = load_reinstatements()
    rows = led.to_dict("records") if len(led) else []
    seen_events = {(str(r["Followup_ID"]), str(r["Orig_Call_Date"])) for r in rows}
    open_fids = {str(r["Followup_ID"]) for r in rows if str(r.get("Status")) == "Open"}

    # 1) open new ones from reconciliation findings
    rec = reconcile_calls(call_log, fu_ledger, visits, outbound)
    nid = _next_reinstate_id(led)
    new_n = 0
    for fr in rec.get("rows", []):
        if fr.get("flag") != "claims_noanswer_but_connected":
            continue
        fid = str(fr.get("key", "")); dt = str(fr.get("date", ""))
        # never recreate the exact event; and never duplicate a patient already
        # on an OPEN reinstatement (re-logging NOT PICK can't add a second row,
        # nor clear the first). A fresh event AFTER a clear opens a new one.
        if (fid, dt) in seen_events or fid in open_fids:
            continue
        ci = []
        if fr.get("ivr_dur"):  ci.append(f"connected {fr['ivr_dur']}s")
        if fr.get("ivr_time"): ci.append(f"at {fr['ivr_time']}")
        connect_info = (" ".join(ci) + f" on {dt}").strip()
        rr = {
            "Reinstate_ID": f"R{nid:06d}", "Followup_ID": fid,
            "Patient_UID": "", "Patient_Name": fr.get("name", ""),
            "Mobile": fr.get("mobile", ""), "Orig_Caller": fr.get("called_by", ""),
            "Orig_Call_Date": dt, "Connect_Info": connect_info,
            "Created_On": date_to_str(today), "Status": "Open",
            "Cleared_On": "", "Cleared_By": "", "Cleared_Reason": "",
        }
        rows.append(rr); seen_events.add((fid, dt)); open_fids.add(fid)
        nid += 1; new_n += 1

    # 2) auto-clear opens now genuinely reached / returned
    fu_status = {str(r.get("Followup_ID", "")): str(r.get("Followup_Status", ""))
                 for _, r in fu_ledger.iterrows()}
    resolved_after = {}
    if len(call_log):
        for _, c in call_log.iterrows():
            cls = (str(c.get("Resolution_Class", "")).strip()
                   or resolution_class_for(c.get("Response", "")))
            if cls == "RESOLVE":
                k = str(c.get("Obligation_Key", "")); d = parse_date(c.get("Call_Date"))
                if d and (k not in resolved_after or d > resolved_after[k]):
                    resolved_after[k] = d
    cleared_n = 0
    for r in rows:
        if str(r.get("Status")) != "Open":
            continue
        fid = str(r["Followup_ID"]); ocd = parse_date(r["Orig_Call_Date"])
        if fu_status.get(fid, "").startswith("Returned"):
            r.update(Status="Cleared", Cleared_On=date_to_str(today),
                     Cleared_By="auto", Cleared_Reason="patient returned"); cleared_n += 1
            continue
        ra = resolved_after.get(fid)
        if ra and ocd and ra > ocd:
            r.update(Status="Cleared", Cleared_On=date_to_str(today),
                     Cleared_By="auto", Cleared_Reason="reached on a later call"); cleared_n += 1

    out = pd.DataFrame(rows, columns=REINSTATE_COLS) if rows else pd.DataFrame(columns=REINSTATE_COLS)
    save_reinstatements(out)
    return {"new": new_n, "cleared": cleared_n,
            "open": int((out["Status"] == "Open").sum()) if len(out) else 0}

def open_reinstatement_records() -> list:
    """Callable rows for the Call Sheet's reinstated-callback section."""
    led = load_reinstatements()
    recs = []
    if not len(led):
        return recs
    for _, r in led[led["Status"] == "Open"].iterrows():
        orig = str(r.get("Orig_Caller", "")) or "—"
        ci = str(r.get("Connect_Info", ""))
        if "rejected by" in ci.lower():
            info = f"{ci} (originally logged by {orig})"
        else:
            info = (f"Incomplete call — {ci}; originally logged NOT PICK by {orig}")
        recs.append({
            "pr": "A", "name": r.get("Patient_Name", ""),
            "mobile": str(r.get("Mobile", "")),
            "info": info,
            "date": str(r.get("Orig_Call_Date", "")), "od": "",
            "status": "Callback reinstated", "vip": False,
            "key": str(r.get("Followup_ID", "")),
        })
    return recs

def open_reinstatements_view() -> list:
    """Open reinstatements for the doctor's web review (with IDs to clear)."""
    led = load_reinstatements()
    if not len(led):
        return []
    out = []
    for _, r in led[led["Status"] == "Open"].iterrows():
        out.append({k: str(r.get(k, "")) for k in
                    ("Reinstate_ID", "Patient_Name", "Mobile", "Orig_Caller",
                     "Connect_Info", "Orig_Call_Date")})
    return out

def clear_reinstatement(reinstate_id: str, by: str = "doctor", reason: str = "reviewed") -> bool:
    """Mark a reinstatement cleared. Used by the doctor's review button — NOT
    available to the flagged caller."""
    led = load_reinstatements()
    if not len(led):
        return False
    m = led["Reinstate_ID"].astype(str) == str(reinstate_id)
    if not m.any():
        return False
    led.loc[m, "Status"] = "Cleared"
    led.loc[m, "Cleared_On"] = date_to_str(date.today())
    led.loc[m, "Cleared_By"] = by
    led.loc[m, "Cleared_Reason"] = reason
    save_reinstatements(led)
    return True

# ── Confirmation & escalation layer (decentralised to Shavez + assistants) ────
# Each "closing" response is a factual claim with a natural witness. Closes that
# can be silently wrong are routed to a named owner for confirmation; the
# auto-settled ones (visit-backed YES / DIKHA CHUKE drop off the call list on
# their own) never reach a person. Presence confirms; absence flags for review —
# it never auto-convicts. Only the final summary reaches the doctor.
#   Tt COMPLETE  → owner Shavez    (clinical close — confirm with patient/records)
#   MED HERE     → owner Pharmacy  (corroborate vs the Marg pharmacy upload)
#   NO (decline) → owner Shavez    (retention digest — scan for patterns)
#   MED OUTSIDE  → self-owned, declared outside — no confirmation row
CONFIRM_FILE = DATA_DIR / "confirmations.csv"
CONFIRM_COLS = ["Confirm_ID", "Followup_ID", "Clinic_Id", "Patient_UID",
                "Patient_Name", "Mobile", "Response", "Conf_Type", "Owner",
                "Caller", "Call_Date", "Remarks", "Created_On", "Status",
                "Resolved_On", "Resolved_By", "Resolved_Note"]
CONFIRM_RESPONSES = {   # response → (Conf_Type, Owner, label)
    "Tt COMPLETE": ("Tt COMPLETE", "Shavez",
                    "Confirm treatment-complete with the patient / records"),
    "MED HERE":    ("MED HERE", "Pharmacy",
                    "Corroborate 'bought here' against the pharmacy upload"),
    "NO":          ("DECLINE", "Shavez", "Review decline reason (retention)"),
}

def load_confirmations() -> pd.DataFrame:
    return _load_csv_with_schema(CONFIRM_FILE, CONFIRM_COLS)

def save_confirmations(df: pd.DataFrame):
    df[CONFIRM_COLS].to_csv(CONFIRM_FILE, index=False)

def _next_confirm_id(df) -> int:
    have = df["Confirm_ID"].astype(str) if len(df) else []
    nums = [int(x[1:]) for x in have if x.startswith("C") and x[1:].isdigit()]
    return (max(nums) + 1) if nums else 1

def _add_reinstatement(followup_id, name, mobile, caller, reason, today=None):
    """Directly open a reinstatement callback (used when a confirmation is
    rejected). Idempotent on an already-open one for the same follow-up."""
    if today is None:
        today = date.today()
    led = load_reinstatements()
    rows = led.to_dict("records") if len(led) else []
    if any(str(r["Followup_ID"]) == str(followup_id) and str(r.get("Status")) == "Open"
           for r in rows):
        return
    nid = _next_reinstate_id(led)
    rows.append({
        "Reinstate_ID": f"R{nid:06d}", "Followup_ID": str(followup_id),
        "Patient_UID": "", "Patient_Name": name, "Mobile": mobile,
        "Orig_Caller": caller, "Orig_Call_Date": date_to_str(today),
        "Connect_Info": reason, "Created_On": date_to_str(today),
        "Status": "Open", "Cleared_On": "", "Cleared_By": "", "Cleared_Reason": "",
    })
    save_reinstatements(pd.DataFrame(rows, columns=REINSTATE_COLS))

def sync_confirmations(call_log=None, fu_ledger=None, today=None) -> dict:
    """Open pending confirmations from confirmable closing responses (Tt COMPLETE,
    MED HERE, NO). Idempotent — one open confirmation per (follow-up, type); a new
    event after one is resolved opens a fresh one."""
    if today is None:     today = date.today()
    if call_log is None:  call_log = load_call_log()
    if fu_ledger is None: fu_ledger = load_followups()

    fu_info = {}
    for _, r in fu_ledger.iterrows():
        fu_info[str(r.get("Followup_ID", ""))] = (
            str(r.get("Clinic_Specific_Id_Resolved", "")),
            str(r.get("Standard_Name") or r.get("FU_Name_Raw") or ""),
            str(r.get("FU_Mobile_Clean", "")))

    led = load_confirmations()
    rows = led.to_dict("records") if len(led) else []
    seen_events = {(str(r["Followup_ID"]), str(r["Call_Date"]), str(r["Conf_Type"]))
                   for r in rows}
    open_ft = {(str(r["Followup_ID"]), str(r["Conf_Type"]))
               for r in rows if str(r.get("Status")) == "Open"}

    nid = _next_confirm_id(led)
    new_n = 0
    if len(call_log):
        cl = call_log[call_log["Obligation_Key"].astype(str).str.startswith("F")].copy()
        cl["_d"] = cl["Call_Date"].map(lambda s: parse_date(s) or date(1900, 1, 1))
        cl = cl.sort_values("_d")
        for _, c in cl.iterrows():
            resp = str(c.get("Response", "")).strip()
            if resp not in CONFIRM_RESPONSES:
                continue
            ctype, owner, _lbl = CONFIRM_RESPONSES[resp]
            fid = str(c.get("Obligation_Key", "")); cdate = str(c.get("Call_Date", ""))
            if (fid, cdate, ctype) in seen_events or (fid, ctype) in open_ft:
                continue
            csid, nm, mob = fu_info.get(fid, ("", "", ""))
            rows.append({
                "Confirm_ID": f"C{nid:06d}", "Followup_ID": fid, "Clinic_Id": csid,
                "Patient_UID": str(c.get("Patient_UID", "")), "Patient_Name": nm,
                "Mobile": mob, "Response": resp, "Conf_Type": ctype, "Owner": owner,
                "Caller": str(c.get("Called_By", "")), "Call_Date": cdate,
                "Remarks": str(c.get("Remarks", "")), "Created_On": date_to_str(today),
                "Status": "Open", "Resolved_On": "", "Resolved_By": "", "Resolved_Note": "",
            })
            seen_events.add((fid, cdate, ctype)); open_ft.add((fid, ctype))
            nid += 1; new_n += 1

    out = pd.DataFrame(rows, columns=CONFIRM_COLS) if rows else pd.DataFrame(columns=CONFIRM_COLS)
    save_confirmations(out)
    return {"new": new_n, "open": int((out["Status"] == "Open").sum()) if len(out) else 0}

def open_confirmations_view() -> list:
    """Open confirmations for the assistants' owner-tagged worklist."""
    led = load_confirmations()
    if not len(led):
        return []
    out = []
    for _, r in led[led["Status"] == "Open"].iterrows():
        out.append({k: str(r.get(k, "")) for k in
                    ("Confirm_ID", "Clinic_Id", "Patient_Name", "Mobile", "Response",
                     "Conf_Type", "Owner", "Caller", "Call_Date", "Remarks")})
    return out

def _resolve_confirmation(confirm_id, status, by, note):
    led = load_confirmations()
    if not len(led):
        return None
    m = led["Confirm_ID"].astype(str) == str(confirm_id)
    if not m.any():
        return None
    row = led[m].iloc[0].to_dict()
    led.loc[m, "Status"] = status
    led.loc[m, "Resolved_On"] = date_to_str(date.today())
    led.loc[m, "Resolved_By"] = by
    led.loc[m, "Resolved_Note"] = note
    save_confirmations(led)
    return row

def confirm_confirmation(confirm_id, by="assistant", note="confirmed") -> bool:
    return _resolve_confirmation(confirm_id, "Confirmed", by, note) is not None

def review_confirmation(confirm_id, by="Shavez", note="reviewed") -> bool:
    return _resolve_confirmation(confirm_id, "Reviewed", by, note) is not None

def reject_confirmation(confirm_id, by="Shavez", note="not confirmed — call back") -> bool:
    """Reject a close → reinstate the callback so the patient is called again."""
    row = _resolve_confirmation(confirm_id, "Rejected", by, note)
    if row is None:
        return False
    reason = f"{row.get('Conf_Type','close')} rejected by {by} — call back"
    _add_reinstatement(row.get("Followup_ID", ""), row.get("Patient_Name", ""),
                       row.get("Mobile", ""), row.get("Caller", ""), reason)
    return True

def apply_call_overlay(combined: pd.DataFrame, today: date):
    """Split the follow-up call list using captured call outcomes. Returns
    (show_df, watch_df), and tags show_df rows with a '_call_note':

      RESOLVE / DECLINE                 → hidden (handled; stop calling)
      RESCHEDULE, still inside hold      → hidden until Last_Call_Date+RESCHEDULE_DAYS
      RESCHEDULE, hold expired           → shown, note 'rescheduled — call now'
      RETRY (not yet escalated)          → shown, note 'N× tried — call #N+1'
      ESCALATE                           → moved to watch_df
      blank / no call yet                → shown unchanged

    Presentation-layer only: the status engine and the audit ledger are untouched."""
    empty = (combined.iloc[0:0].copy() if combined is not None else None)
    if combined is None or not len(combined):
        if combined is not None and "_call_note" not in combined.columns:
            combined = combined.copy(); combined["_call_note"] = ""
        return combined, empty
    d = combined.copy()
    if "Call_Resolution" not in d.columns:
        d["_call_note"] = ""
        return d, d.iloc[0:0].copy()
    res = list(d["Call_Resolution"].astype(str))
    lastresp = list(d.get("Last_Response", pd.Series([""] * len(d))).astype(str).str.strip())
    lastcall = [parse_date(s) for s in d.get("Last_Call_Date", pd.Series([""] * len(d)))]
    attempts = pd.to_numeric(d.get("Call_Attempts", pd.Series([""] * len(d))),
                             errors="coerce").fillna(0).astype(int).tolist()
    keep, notes, watch_pos = [], [], []
    for i, (rcls, lc, att) in enumerate(zip(res, lastcall, attempts)):
        lresp = lastresp[i] if i < len(lastresp) else ""
        if rcls == "RESOLVE":
            # A row that is still on the action list here has NOT returned (the
            # engine excludes returned patients). Cross-verify the claim:
            #   'YES/coming' is a PROMISE — if no real return within the grace
            #   window, re-surface it for a re-confirm call. Other resolutions
            #   (treatment complete / medicine / already seen) expect no return,
            #   so they stay closed.
            if lresp in EXPECTS_RETURN_RESPONSES:
                if lc and today < lc + timedelta(days=PROMISE_GRACE_DAYS):
                    keep.append(False); notes.append(""); continue
                _dl = lc.strftime('%d-%b') if lc else ""
                keep.append(True)
                notes.append(f"said coming {_dl} — not back yet, re-confirm")
                continue
            keep.append(False); notes.append(""); continue
        if rcls == "DECLINE":
            keep.append(False); notes.append(""); continue
        if rcls == "ESCALATE":
            keep.append(False); notes.append(""); watch_pos.append(i); continue
        if rcls == "RESCHEDULE":
            if lc and today < lc + timedelta(days=RESCHEDULE_DAYS):
                keep.append(False); notes.append(""); continue
            keep.append(True); notes.append("said call later — call now"); continue
        if rcls == "RETRY":
            keep.append(True)
            notes.append(f"no answer — called {att}\u00d7, call again"); continue
        keep.append(True); notes.append("")
    d["_call_note"] = notes
    show = d[pd.Series(keep, index=d.index)].copy()
    watch = d.iloc[watch_pos].copy() if watch_pos else d.iloc[0:0].copy()
    if len(watch):
        watch["_call_note"] = [
            f"{int(a)}\u00d7 no contact — verify / WABA / drop"
            for a in pd.to_numeric(watch.get("Call_Attempts", 0),
                                   errors="coerce").fillna(0).astype(int)]
    return show, watch

# ── Date helpers ──────────────────────────────────────────────────────────────
def parse_date(val) -> date | None:
    """Parse date from multiple formats. Returns date object or None."""
    if pd.isna(val) or not str(val).strip():
        return None
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    # Try datetime string like "01-06-2026 01:25 PM"
    try:
        return datetime.strptime(s[:10], "%d-%m-%Y").date()
    except ValueError:
        pass
    return None

def date_to_str(d) -> str:
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%Y-%m-%d")
    return str(d)

def next_call_day(visit_day: date) -> date:
    """The day patients are actually called: the day AFTER the visit, except
    Sunday is emergency-only at this clinic (no routine follow-ups scheduled,
    no routine calls), so a SATURDAY visit's call-backs roll to MONDAY.
    This is why a Saturday-night upload should already produce Monday's sheet."""
    nxt = visit_day + timedelta(days=1)
    if nxt.weekday() == 6:           # 6 == Sunday
        nxt = nxt + timedelta(days=1)
    return nxt

# ── Step 1: Ingest consultation report ───────────────────────────────────────
def parse_consultation_report(filepath: str) -> tuple[pd.DataFrame, str]:
    """
    Parse Docterz consultation report CSV.
    Returns (dataframe, log_date_string).
    Skips the clinic-name header row (row 0).
    """
    try:
        raw = pd.read_csv(filepath, dtype=str)
    except pd.errors.EmptyDataError:
        raw = pd.DataFrame()

    if len(raw) == 0:
        # Return empty DataFrame with correct columns
        empty = pd.DataFrame(columns=[
            "Patient_UID","Clinic_Specific_Id","Patient_Name",
            "Mobile_Raw","Mobile_Clean","Visit_Date","Had_Procedure","Source_File"
        ])
        fname = Path(filepath).stem
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fname)
        log_date = date_match.group(1) if date_match else date_to_str(date.today())
        return empty, log_date

    # Row 0 is "Dr. Manoj Agarwal Clinic" — drop it
    if str(raw.iloc[0].get("Sr No", "")).strip().lower() in (
        "dr. manoj agarwal clinic", "dr manoj agarwal clinic", ""
    ):
        raw = raw.iloc[1:].reset_index(drop=True)

    # Report date = the date the visits ACTUALLY happened (from the report's
    # CONTENT), not the filename. Docterz names the export by the DOWNLOAD day, so
    # yesterday's report pulled this morning is named one day ahead of its visits.
    # Keying the report date off the filename made the procedure call-back filter
    # (Visit_Date == report_date) match nothing every day. We take the latest real
    # visit date in the file; the filename is only a last resort.
    _vdates = ([parse_date(x) for x in raw["Consultation Date"]]
               if "Consultation Date" in raw.columns else [])
    _vdates = [d for d in _vdates if d]
    if _vdates:
        log_date = date_to_str(max(_vdates))
    else:
        fname = Path(filepath).stem  # consultation_report_2026-06-01
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fname)
        log_date = date_match.group(1) if date_match else date_to_str(date.today())

    df = pd.DataFrame()
    df["Patient_UID"]        = raw["Patient UID"].str.strip()
    df["Clinic_Specific_Id"] = raw["Clinic Specific Id"].apply(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ("", "nan") else ""
    )
    df["Patient_Name"]       = raw["Patient Name"].str.strip()
    df["Mobile_Raw"]         = raw["Mobile"].astype(str)
    df["Mobile_Clean"]       = df["Mobile_Raw"].apply(clean_mobile)
    df["Visit_Date"]         = raw["Consultation Date"].apply(
        lambda x: date_to_str(parse_date(x))
    )
    df["Source_File"]        = Path(filepath).name
    # Procedure flag: a patient had a procedure when "Procedure Amount" > 0.
    # NOTE: this misses procedures billed at 0 (e.g. Ayushman cashless). If such
    # cases need to be caught, add the alternate signal here.
    def _amt(x):
        try:
            return float(str(x).replace(",", "").strip() or 0)
        except (ValueError, TypeError):
            return 0.0
    # A procedure happened if there is ANY procedure billing signal — amount OR
    # discount. This catches Ayushman/cashless procedures billed then fully
    # discounted to net ₹0. (Procedures entered as a literal 0 with no discount
    # have no signal in the export and are handled via the manual ₹0-procedure
    # marker, which sets this flag on the visit row at run time.)
    if "Procedure Amount" in raw.columns:
        proc_amt = raw["Procedure Amount"].apply(_amt).values
        proc_disc = (raw["Procedure Discount"].apply(_amt).values
                     if "Procedure Discount" in raw.columns else [0] * len(raw))
        df["Had_Procedure"] = ["Y" if (a > 0 or d > 0) else "" for a, d in zip(proc_amt, proc_disc)]
    else:
        df["Had_Procedure"] = ""
    # Drop rows with no Patient_UID
    df = df[df["Patient_UID"].notna() & (df["Patient_UID"] != "")].reset_index(drop=True)
    return df, log_date

# ── Step 2: Update Patient Master from consultation report ────────────────────
def update_master_from_consultation(
    master: pd.DataFrame,
    consult: pd.DataFrame,
    processed_on: str
) -> tuple[pd.DataFrame, int]:
    """
    Update master with new/updated patient data.
    Returns (updated_master, new_patients_count).
    """
    master = master.copy()
    new_count = 0

    for _, row in consult.iterrows():
        uid  = row["Patient_UID"]
        name = row["Patient_Name"]
        mob_raw   = row["Mobile_Raw"]
        mob_clean = row["Mobile_Clean"]
        visit_dt  = row["Visit_Date"]
        csid      = row["Clinic_Specific_Id"]

        existing = master[master["Patient_UID"] == uid]

        if len(existing) == 0:
            # New patient
            new_row = {
                "Patient_UID":           uid,
                "Clinic_Specific_Id":    csid,
                "Patient_Name":          name,
                "Mobile_Raw":            mob_raw,
                "Mobile_Clean":          mob_clean,
                "Mobile_Status":         mobile_status(mob_clean),
                "First_Seen_Date":       visit_dt,
                "Last_Seen_Date":        visit_dt,
                "Mobile_Duplicate_Count": "0",
                "Identity_Status":       "",
                "Added_From":            f"Daily Seen Report ({row['Source_File']})",
                "Last_Updated":          processed_on
            }
            master = pd.concat([master, pd.DataFrame([new_row])], ignore_index=True)
            new_count += 1
        else:
            idx = existing.index[0]
            # Update Last_Seen_Date if newer
            existing_last = parse_date(master.at[idx, "Last_Seen_Date"])
            this_date     = parse_date(visit_dt)
            if this_date and (not existing_last or this_date > existing_last):
                master.at[idx, "Last_Seen_Date"] = visit_dt
            # Update First_Seen_Date if older
            existing_first = parse_date(master.at[idx, "First_Seen_Date"])
            if this_date and (not existing_first or this_date < existing_first):
                master.at[idx, "First_Seen_Date"] = visit_dt
            # Update mobile if master has none and this one is valid
            if not master.at[idx, "Mobile_Clean"] and mob_clean:
                master.at[idx, "Mobile_Raw"]   = mob_raw
                master.at[idx, "Mobile_Clean"] = mob_clean
                master.at[idx, "Mobile_Status"] = mobile_status(mob_clean)
            # Update Clinic_Specific_Id if blank
            if not master.at[idx, "Clinic_Specific_Id"] and csid:
                master.at[idx, "Clinic_Specific_Id"] = csid
            master.at[idx, "Last_Updated"] = processed_on

    master = recalculate_duplicate_flags(master)
    return master, new_count

# ── Step 3: Append visits to Visit Ledger ────────────────────────────────────
def append_visits(
    visits: pd.DataFrame,
    consult: pd.DataFrame,
    processed_on: str
) -> tuple[pd.DataFrame, int]:
    """Append new visits; skip duplicates (same UID + same date)."""
    new_rows = []
    existing_keys = set(
        zip(visits["Patient_UID"].tolist(), visits["Visit_Date"].tolist())
    )
    # Determine next Visit_ID
    if len(visits) > 0:
        max_id = visits["Visit_ID"].apply(
            lambda x: int(re.sub(r'\D', '', str(x)) or 0)
        ).max()
    else:
        max_id = 0

    added = 0
    proc_refreshed = 0
    for _, row in consult.iterrows():
        key = (row["Patient_UID"], row["Visit_Date"])
        if key in existing_keys:
            # Same-day re-upload: the visit already exists, so we do NOT duplicate
            # it. But if the re-uploaded report now shows a procedure the stored row
            # is missing (a procedure billed AFTER the first upload), flip
            # Had_Procedure to "Y" so the patient reaches the PROCEDURE CALL-BACKS
            # list. Upgrade only — never downgrade a "Y" back to blank.
            if str(row.get("Had_Procedure", "")).strip().upper() in ("Y", "YES", "TRUE", "1"):
                _m = ((visits["Patient_UID"].astype(str) == str(row["Patient_UID"]))
                      & (visits["Visit_Date"].astype(str) == str(row["Visit_Date"]))
                      & (~visits["Had_Procedure"].astype(str).str.strip().str.upper()
                           .isin(["Y", "YES", "TRUE", "1"])))
                if _m.any():
                    visits.loc[_m, "Had_Procedure"] = "Y"
                    proc_refreshed += int(_m.sum())
            continue
        max_id += 1
        new_rows.append({
            "Visit_ID":          f"V{max_id:06d}",
            "Visit_Date":        row["Visit_Date"],
            "Patient_UID":       row["Patient_UID"],
            "Clinic_Specific_Id": row["Clinic_Specific_Id"],
            "Patient_Name":      row["Patient_Name"],
            "Mobile_Raw":        row["Mobile_Raw"],
            "Mobile_Clean":      row["Mobile_Clean"],
            "Had_Procedure":     row.get("Had_Procedure", ""),
            "Source_File":       row["Source_File"],
            "Processed_On":      processed_on
        })
        existing_keys.add(key)
        added += 1

    if new_rows:
        visits = pd.concat([visits, pd.DataFrame(new_rows)], ignore_index=True)
    if proc_refreshed:
        INGEST_NOTICES.append(
            f"Re-upload: refreshed the procedure flag on {proc_refreshed} patient(s) "
            f"already in today's visit ledger (a procedure was added after the first "
            f"upload). They now appear in the PROCEDURE CALL-BACKS list."
        )
    return visits, added

# ── Step 4: Resolve follow-up identities ─────────────────────────────────────
def resolve_identity(mobile_clean: str, name_raw: str, master: pd.DataFrame):
    """
    Returns (patient_uid, csid, std_name, confidence, issue)
    """
    if not mobile_clean:
        return "", "", "", "Invalid Mobile", "Mobile missing or invalid"

    matches = master[master["Mobile_Clean"] == mobile_clean]

    if len(matches) == 0:
        return "", "", "", "Unresolved", "Mobile not found in Patient Master"

    if len(matches) == 1:
        row = matches.iloc[0]
        return (
            row["Patient_UID"],
            row["Clinic_Specific_Id"],
            row["Patient_Name"],
            "High",
            ""
        )

    # Shared mobile (family/attendant) — try name matching
    name_norm = str(name_raw).strip().lower()
    matches = matches.copy()
    matches["_name_match"] = matches["Patient_Name"].apply(
        lambda n: name_match_score(name_norm, str(n).lower())
    )
    best = matches.nlargest(1, "_name_match").iloc[0]

    if best["_name_match"] >= 0.7:
        return (
            best["Patient_UID"],
            best["Clinic_Specific_Id"],
            best["Patient_Name"],
            "Medium",
            f"Shared mobile; name matched to {best['Patient_Name']}"
        )

    # Name match failed — cannot safely assign
    uids = ", ".join(matches["Patient_UID"].tolist())
    return "", "", "", "Ambiguous", f"Shared family mobile — patients: {uids}. Verify name before contacting."

def name_match_score(a: str, b: str) -> float:
    """Simple token overlap score between two name strings."""
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    if not tokens_a or not tokens_b:
        return 0.0
    overlap = tokens_a & tokens_b
    return len(overlap) / max(len(tokens_a), len(tokens_b))

# ── Step 5: Ingest follow-up log ─────────────────────────────────────────────
def parse_followup_log(filepath: str, anchor_date=None) -> tuple[pd.DataFrame, str]:
    """
    Parse Docterz follow-up log CSV.
    The file contains no export date — log_date is derived as Due_Date minus 1 day,
    since Docterz follow-up logs are always generated one day in advance.
    If rows have mixed due dates, the earliest due date minus 1 is used as log_date.
    """
    try:
        raw = pd.read_csv(filepath, dtype=str)
    except pd.errors.EmptyDataError:
        raw = pd.DataFrame()

    if len(raw) == 0:
        empty = pd.DataFrame(columns=[
            "Appointment_ID","FU_Name_Raw","FU_Mobile_Raw",
            "FU_Mobile_Clean","Due_Date","Followup_Log_Date"
        ])
        return empty, date_to_str(date.today())

    df = pd.DataFrame()
    df["Appointment_ID"]  = raw["Appointment ID"].str.strip()
    df["FU_Name_Raw"]     = raw["Patient Name"].str.strip()
    df["FU_Mobile_Raw"]   = raw["Mobile No"].astype(str)
    df["FU_Mobile_Clean"] = df["FU_Mobile_Raw"].apply(clean_mobile)
    df["Due_Date"]        = raw["Followup Date"].apply(
        lambda x: date_to_str(parse_date(x))
    )

    # Derive log_date: earliest Due_Date in this file minus 1 day
    due_dates = [parse_date(d) for d in df["Due_Date"] if parse_date(d)]
    if due_dates:
        earliest_due = min(due_dates)
        latest_due   = max(due_dates)
        log_date = date_to_str(earliest_due - timedelta(days=1))

        # GUARD: a normal daily file is all one due date (span 0). A Docterz
        # default export spans ~30 days. If the span is too wide, keep only the
        # earliest day so an accidental month upload cannot flood the ledger.
        span_days = (latest_due - earliest_due).days
        if span_days > MAX_FOLLOWUP_SPAN_DAYS:
            # Wide file = Docterz's default ~one-month export. Pick a SINGLE day so
            # an accidental month upload can't flood the ledger. Prefer the
            # consultation-anchored next-call day (anchor + 1); fall back to the
            # earliest day only when no consultation report is in the run to anchor
            # against. Narrow, deliberate exports (span <= MAX) are never sliced.
            before = len(df)
            rng = f"{date_to_str(earliest_due)} to {date_to_str(latest_due)}"
            if anchor_date is not None:
                # Target the clinic's NEXT CALL DAY. Sunday is emergency-only, so a
                # Saturday run rolls to Monday (next_call_day handles the skip). Load
                # every row due AFTER the anchor up to and including that call day —
                # this picks the next working day and folds in any rare Sunday-due
                # rows so they are called on Monday rather than lost.
                call_day = next_call_day(anchor_date)
                call_day_str = date_to_str(call_day)
                def _in_window(s, _a=anchor_date, _c=call_day):
                    d = parse_date(s)
                    return (d is not None) and (_a < d <= _c)
                df = df[df["Due_Date"].apply(_in_window)].copy()
                log_date = date_to_str(anchor_date)
                _skip = "; Sunday skipped" if (call_day - anchor_date).days > 1 else ""
                if len(df):
                    INGEST_NOTICES.append(
                        f"Follow-up file covered {span_days + 1} days ({rng}) — Docterz's "
                        f"default one-month export. Anchored to consultation date "
                        f"{date_to_str(anchor_date)} \u2192 loaded the next call day "
                        f"({call_day_str}{_skip}): {len(df)} of {before} rows. Other days "
                        f"load on their own eves."
                    )
                else:
                    INGEST_NOTICES.append(
                        f"Follow-up file covered {span_days + 1} days ({rng}) but had NO rows "
                        f"due by the next call day ({call_day_str}{_skip}), anchored to "
                        f"consultation {date_to_str(anchor_date)}; loaded 0 rows. If you "
                        f"expected calls for {call_day_str}, check the Docterz export range."
                    )
            else:
                earliest_str = date_to_str(earliest_due)
                df = df[df["Due_Date"] == earliest_str].copy()
                INGEST_NOTICES.append(
                    f"Follow-up file covered {span_days + 1} days ({rng}) — this looks like "
                    f"Docterz's default one-month export. No consultation report to anchor "
                    f"to, so loaded ONLY the earliest day ({earliest_str}): {len(df)} of "
                    f"{before} rows. The other days load on their own dates."
                )
    else:
        log_date = date_to_str(date.today())

    df["Followup_Log_Date"] = log_date
    return df, log_date

# ── Step 6: Merge follow-up log into Follow-Up Ledger ────────────────────────
def append_followups(
    fu_ledger: pd.DataFrame,
    fu_log: pd.DataFrame,
    master: pd.DataFrame,
    processed_on: str
) -> tuple[pd.DataFrame, int]:
    """Add new follow-up obligations; skip exact duplicates."""
    # Existing key: Appointment_ID (unique per appointment in Docterz)
    existing_appt_ids = set(fu_ledger["Appointment_ID"].tolist())

    if len(fu_ledger) > 0:
        max_id = fu_ledger["Followup_ID"].apply(
            lambda x: int(re.sub(r'\D', '', str(x)) or 0)
        ).max()
    else:
        max_id = 0

    new_rows = []
    added = 0

    for _, row in fu_log.iterrows():
        appt_id = row["Appointment_ID"]
        if appt_id in existing_appt_ids:
            continue

        mob_clean = row["FU_Mobile_Clean"]
        uid, csid, std_name, confidence, issue = resolve_identity(
            mob_clean, row["FU_Name_Raw"], master
        )

        max_id += 1
        new_rows.append({
            "Followup_ID":               f"F{max_id:06d}",
            "Followup_Log_Date":         row["Followup_Log_Date"],
            "Due_Date":                  row["Due_Date"],
            "FU_Name_Raw":               row["FU_Name_Raw"],
            "FU_Mobile_Raw":             row["FU_Mobile_Raw"],
            "FU_Mobile_Clean":           mob_clean,
            "Appointment_ID":            appt_id,
            "Patient_UID_Resolved":      uid,
            "Clinic_Specific_Id_Resolved": csid,
            "Standard_Name":             std_name,
            "Identity_Confidence":       confidence,
            "Identity_Issue":            issue,
            "Matched_Return_Visit_Date": "",
            "Return_Delay_Days":         "",
            "Followup_Status":           "",
            "Days_Overdue":              "",
            "Suggested_Action":          "",
            "Last_Status_Update":        processed_on
        })
        existing_appt_ids.add(appt_id)
        added += 1

    if new_rows:
        fu_ledger = pd.concat([fu_ledger, pd.DataFrame(new_rows)], ignore_index=True)
    return fu_ledger, added

# ── Step 7: Recalculate all open follow-up statuses ──────────────────────────
def recalculate_statuses(
    fu_ledger: pd.DataFrame,
    visits: pd.DataFrame,
    today: date,
    processed_on: str
) -> pd.DataFrame:
    """
    For every follow-up row, determine current status using the Visit Ledger.

    One-visit-one-follow-up rule (P0-01 fix):
      A single return visit can close at most ONE follow-up obligation.
      Visits already claimed by terminal (Returned*) rows — via Matched_Visit_ID,
      or for legacy rows via UID+date — are excluded before matching. Pending
      follow-ups for the same patient are matched in due-date order, each
      consuming the earliest still-unclaimed visit on/after its log date.

    60-day expiry rule (P0-02 fix):
      Any follow-up still unreturned more than EXPIRY_DAYS past its due date
      becomes "Expired Unresolved". It stays in the ledger and can still flip
      to "Returned Late" if the patient eventually comes back, but it is
      hidden from Staff Action Today. Identity-problem rows expire the same
      way so stale exceptions don't pollute the action sheet either.
    """
    fu_ledger = fu_ledger.copy()

    # Build lookup: uid -> sorted list of (visit_date, visit_id)
    visit_map: dict[str, list] = {}
    visit_id_by_key: dict[tuple, str] = {}   # (uid, date_str) -> visit_id (legacy retro-claim)
    for _, vrow in visits.iterrows():
        uid = vrow["Patient_UID"]
        vd  = parse_date(vrow["Visit_Date"])
        vid = str(vrow.get("Visit_ID", "")).strip()
        if uid and vd:
            visit_map.setdefault(uid, []).append((vd, vid))
            visit_id_by_key[(uid, date_to_str(vd))] = vid
    for uid in visit_map:
        visit_map[uid].sort(key=lambda t: (t[0], t[1]))

    # Terminal statuses — never re-evaluated
    TERMINAL = {"Returned Early", "Returned On Time", "Returned Late"}

    # ── Pass 1: collect visits already consumed by terminal rows ─────────────
    consumed: set[str] = set()
    for idx, row in fu_ledger.iterrows():
        if row.get("Followup_Status") not in TERMINAL:
            continue
        vid = str(row.get("Matched_Visit_ID", "")).strip()
        if not vid:
            # Legacy terminal row (matched before this column existed):
            # retro-claim its visit by UID + matched date, and backfill the ID.
            uid = str(row.get("Patient_UID_Resolved", "")).strip()
            mdate = str(row.get("Matched_Return_Visit_Date", "")).strip()
            vid = visit_id_by_key.get((uid, mdate), "")
            if vid:
                fu_ledger.at[idx, "Matched_Visit_ID"] = vid
        if vid:
            consumed.add(vid)

    # ── Pass 2: identity-problem rows (status fixed, but they can expire) ────
    pending_idx = []   # rows eligible for visit matching, grouped later by UID
    for idx, row in fu_ledger.iterrows():
        if row.get("Followup_Status") in TERMINAL:
            continue

        due_date = parse_date(row.get("Due_Date"))
        conf = row.get("Identity_Confidence", "")

        if conf in ("Invalid Mobile", "Unresolved", "Ambiguous"):
            if due_date and (today - due_date).days > EXPIRY_DAYS:
                fu_ledger.at[idx, "Followup_Status"]   = "Expired Unresolved"
                fu_ledger.at[idx, "Days_Overdue"]      = str((today - due_date).days)
                fu_ledger.at[idx, "Suggested_Action"]  = "Archived (>%d days)" % EXPIRY_DAYS
            elif conf == "Invalid Mobile":
                fu_ledger.at[idx, "Followup_Status"]   = "Invalid Mobile / No Contact"
                fu_ledger.at[idx, "Suggested_Action"]  = "Correct mobile number in Docterz"
            elif conf == "Unresolved":
                fu_ledger.at[idx, "Followup_Status"]   = "Identity Unresolved"
                fu_ledger.at[idx, "Suggested_Action"]  = "Check registration / mobile in Docterz"
            else:
                fu_ledger.at[idx, "Followup_Status"]   = "Ambiguous Mobile"
                fu_ledger.at[idx, "Suggested_Action"]  = "Verify patient identity before contacting"
            fu_ledger.at[idx, "Last_Status_Update"] = processed_on
            continue

        if not due_date:
            fu_ledger.at[idx, "Followup_Status"]    = "Invalid Due Date"
            fu_ledger.at[idx, "Last_Status_Update"] = processed_on
            continue

        pending_idx.append(idx)

    # ── Pass 3: match pending follow-ups to visits, one visit per follow-up ──
    # Group by UID; within a patient, earlier due dates claim visits first.
    by_uid: dict[str, list] = {}
    for idx in pending_idx:
        uid = str(fu_ledger.at[idx, "Patient_UID_Resolved"]).strip()
        by_uid.setdefault(uid, []).append(idx)

    for uid, idxs in by_uid.items():
        idxs.sort(key=lambda i: (parse_date(fu_ledger.at[i, "Due_Date"]) or date.max,
                                 str(fu_ledger.at[i, "Followup_ID"])))
        for idx in idxs:
            row      = fu_ledger.loc[idx]
            due_date = parse_date(row.get("Due_Date"))
            log_date = parse_date(row.get("Followup_Log_Date"))

            matched_visit = None
            matched_vid   = ""
            if uid and uid in visit_map:
                for vd, vid in visit_map[uid]:
                    if vid in consumed:
                        continue
                    if log_date:
                        # COUNT_LOG_DATE_VISIT_AS_RETURN = False (the safer
                        # default): a visit ON the log date is normally the
                        # prescribing visit itself ("come tomorrow" written
                        # today) — counting it as a return silently hides
                        # dropouts as "Returned Early". The cost: a genuine
                        # one-day-early return shows as due/missed until staff
                        # call and learn the patient already came. Set the
                        # constant True to restore the old behaviour.
                        ok = (vd >= log_date) if COUNT_LOG_DATE_VISIT_AS_RETURN \
                             else (vd > log_date)
                    else:
                        ok = vd >= due_date - timedelta(days=30)
                    if ok:
                        matched_visit, matched_vid = vd, vid
                        break

            if matched_visit:
                consumed.add(matched_vid)
                delay = (matched_visit - due_date).days
                fu_ledger.at[idx, "Matched_Return_Visit_Date"] = date_to_str(matched_visit)
                fu_ledger.at[idx, "Matched_Visit_ID"]          = matched_vid
                fu_ledger.at[idx, "Return_Delay_Days"]         = str(delay)
                fu_ledger.at[idx, "Days_Overdue"]              = ""
                if delay < 0:
                    status = "Returned Early"
                elif delay <= 2:
                    status = "Returned On Time"
                else:
                    status = "Returned Late"
                fu_ledger.at[idx, "Followup_Status"]  = status
                fu_ledger.at[idx, "Suggested_Action"] = "—"
            else:
                days_overdue = (today - due_date).days
                fu_ledger.at[idx, "Days_Overdue"]              = str(days_overdue)
                fu_ledger.at[idx, "Matched_Return_Visit_Date"] = ""
                fu_ledger.at[idx, "Matched_Visit_ID"]          = ""
                fu_ledger.at[idx, "Return_Delay_Days"]         = ""

                if days_overdue < 0:
                    status = "Not Due"
                    action = "—"
                elif days_overdue == 0:
                    status = "Due Today"
                    action = "WhatsApp reminder"
                elif days_overdue <= 3:
                    status = "Grace Period"
                    action = "WhatsApp reminder"
                elif days_overdue <= 10:
                    status = "Actionable Missed Follow-Up"
                    action = "Call / WhatsApp"
                elif days_overdue <= EXPIRY_DAYS:
                    status = "Probable Dropout"
                    action = "Call and mark outcome"
                else:
                    status = "Expired Unresolved"
                    action = "Archived (>%d days)" % EXPIRY_DAYS
                fu_ledger.at[idx, "Followup_Status"]  = status
                fu_ledger.at[idx, "Suggested_Action"] = action

            fu_ledger.at[idx, "Last_Status_Update"] = processed_on

    return fu_ledger

# ── Step 8: Build output Excel ────────────────────────────────────────────────
def build_staff_call_workbook(
    df_reminder: pd.DataFrame,
    df_action: pd.DataFrame,
    df_vac_out: pd.DataFrame,
    visits: pd.DataFrame,
    today: date,
    out_path: str,
    consult_date: date | None = None,
    day_revenue_df: pd.DataFrame | None = None,
    full_fu_ledger: pd.DataFrame | None = None,
):
    """Separate, hand-to-staff CALL SHEET — one printable tab, three sections:

      1. REMINDER CALLS    — follow-ups due TOMORROW (call to remind: 'kal aaiye').
      2. FOLLOW-UP CALLS    — due today + overdue + identity problems (the chase).
      3. PROCEDURE CALL-BACKS — patients who had a procedure YESTERDAY, by Shavez.

    Plus a Vacation Notice List tab (reference, not calls).

    Sections share one column layout so it prints as one clean table. RESPONSE and
    CALLED BY are dropdowns; CALL TIME + REMARKS are free. A hidden KEY column
    carries Followup_ID / Visit_ID for the future call-log read-back (Phase 2).
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    AR = "Arial"
    NAVY = "1F3864"; LIGHTBLUE = "D6E4F0"; GREEN = "548235"; BROWN = "806000"
    PINK = "FCE4EC"
    AMBER = "FFE0B2"   # reinstated callbacks, folded inline so they stand out
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    PRI_FILL = {"A": "FFC7CE", "B": "FFEB9C", "C": "C6EFCE", "D": "E0E0E0"}

    def fmt_d(s):
        d = parse_date(s)
        return d.strftime("%d-%b") if d else ""

    def clean_diag(d):
        d = "" if d is None else str(d).strip()
        return "" if d in ("No Diagnosis Recorded", "nan", "None") else d

    cols = ["S.N", "PR", "PATIENT NAME", "MOBILE", "DIAGNOSIS / INFO", "DATE", "OD",
            "STATUS", "RESPONSE", "CALLED BY", "CALL TIME", "REMARKS", "KEY"]
    widths = [5, 4, 22, 13, 28, 9, 5, 20, 15, 13, 11, 20, 12]
    RESP_COL, CALL_COL, KEY_COL = 9, 10, 13

    wb = Workbook()
    ws = wb.active; ws.title = "Call Sheet"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # title
    last = chr(64 + len(cols))
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = "DR. MANOJ AGARWAL CLINIC  —  STAFF CALL SHEET"
    c.font = Font(AR, bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]
    c.value = (f"Date: {today.strftime('%d-%b-%Y')}    Pick RESPONSE and CALLED BY "
               "from the dropdowns; add notes in REMARKS.")
    c.font = Font(AR, italic=True, size=10)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    dv_resp = DataValidation(type="list", formula1='"' + ",".join(CALL_RESPONSES) + '"', allow_blank=True)
    dv_call = DataValidation(type="list", formula1='"' + ",".join(CALL_STAFF) + '"', allow_blank=True)
    ws.add_data_validation(dv_resp); ws.add_data_validation(dv_call)

    def section_header(r, label, fill):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        c = ws.cell(r, 1, label)
        c.font = Font(AR, bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1
        for j, h in enumerate(cols, 1):
            c = ws.cell(r, j, h)
            c.font = Font(AR, bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=LIGHTBLUE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        return r + 1

    def write_rows(r, records, default_caller=""):
        sn = 1
        for rec in records:
            vals = [sn, rec["pr"], rec["name"], rec["mobile"], rec["info"],
                    rec["date"], rec["od"], rec["status"], "", default_caller, "", "",
                    rec["key"]]
            for j, v in enumerate(vals, 1):
                c = ws.cell(r, j, v)
                c.font = Font(AR, size=10); c.border = border
                c.alignment = Alignment(
                    horizontal="center" if j in (1, 2, 6, 7) else "left", vertical="center")
            if rec.get("vip"):
                for j in range(1, len(cols) + 1):
                    ws.cell(r, j).fill = PatternFill("solid", fgColor=PINK)
            elif rec.get("flag") == "reinstated":
                for j in range(1, len(cols) + 1):
                    ws.cell(r, j).fill = PatternFill("solid", fgColor=AMBER)
            elif rec["pr"] in PRI_FILL:
                cc = ws.cell(r, 2)
                cc.fill = PatternFill("solid", fgColor=PRI_FILL[rec["pr"]])
                cc.font = Font(AR, bold=True, size=10)
            dv_resp.add(ws.cell(r, RESP_COL)); dv_call.add(ws.cell(r, CALL_COL))
            r += 1; sn += 1
        if sn == 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
            ws.cell(r, 1, "— none —").font = Font(AR, italic=True, size=10)
            r += 1
        return r + 1

    def fu_records(df, status_label=None, notes_override=None):
        out = []
        if df is None or not len(df):
            return out
        _PRI = {"A": 0, "B": 1, "C": 2, "D": 3}
        d = df.copy()
        d["_dd"] = d["Due_Date"].map(lambda s: parse_date(s) or date(1900, 1, 1))
        d["_vip"] = d.get("VIP_Flag", "").map(lambda v: 0 if v == "VIP" else 1)
        d["_pri"] = d.get("Priority", "").map(lambda p: _PRI.get(str(p), 4))
        d = d.sort_values(["_dd", "_vip", "_pri"], ascending=[False, True, True])
        for _, row in d.iterrows():
            _fid = str(row.get("Followup_ID", ""))
            _base_status = status_label or {"Not Due": "Upcoming"}.get(
                str(row.get("Followup_Status", "")), str(row.get("Followup_Status", "")))
            _note = str(row.get("_call_note", "") or "").strip()
            _flag = ""
            if notes_override and _fid in notes_override:
                _note = notes_override[_fid]      # clearer reinstated text
                _flag = "reinstated"
            out.append({
                "pr": str(row.get("Priority", "") or ""),
                "name": row.get("FU_Name_Raw", ""),
                "mobile": str(row.get("FU_Mobile_Clean", "")),
                "info": clean_diag(row.get("Diagnosis", "")),
                "date": fmt_d(row.get("Due_Date")),
                "od": row.get("Days_Overdue", ""),
                "status": (_base_status + "  ·  " + _note) if _note else _base_status,
                "vip": row.get("VIP_Flag", "") == "VIP",
                "flag": _flag,
                "key": row.get("Followup_ID", ""),
            })
        return out

    # ── Section 1: Follow-up calls (due today + overdue + upcoming) ──────────
    # One follow-up list: due today + overdue (chase) + any upcoming (so the
    # sheet is complete whether run in the morning or the evening before).
    parts = [d for d in (df_action, df_reminder) if d is not None and len(d)]
    combined = pd.concat(parts, ignore_index=True, sort=False) if parts else df_action

    # Phase 2 overlay: hide resolved/declined, hold rescheduled, annotate retries,
    # and peel off escalated (unreachable) follow-ups into their own section below.
    combined, watch_df = apply_call_overlay(combined, today)

    # ── Reinstated callbacks, folded INLINE ──────────────────────────────────
    # A callback that was logged NOT PICK but connected (or a close a supervisor
    # rejected) is NOT orphaned in a section at the bottom. It sits in the main
    # list at the patient's own due-date, highlighted amber with a plain note, so
    # staff see it alongside that day's other calls.
    reinstate_note = {}
    try:
        _ri = load_reinstatements()
        if len(_ri):
            import re as _re
            for _, rr in _ri[_ri["Status"] == "Open"].iterrows():
                _fid = str(rr.get("Followup_ID", ""))
                _ci = str(rr.get("Connect_Info", ""))
                _orig = str(rr.get("Orig_Caller", "")) or "—"
                if "rejected by" in _ci.lower():
                    reinstate_note[_fid] = f"\u26a0 not confirmed — {_ci}"
                else:
                    _m = _re.search(r"connected\s+\d+s", _ci)
                    _frag = _m.group(0) if _m else "call connected"
                    reinstate_note[_fid] = (f"\u26a0 {_frag} last time but marked NOT PICK "
                                            f"({_orig}) — call back & complete")
            # pull in any reinstated follow-up that the overlay had suppressed
            _present = set(combined["Followup_ID"].astype(str)) if len(combined) else set()
            _missing = [f for f in reinstate_note if f not in _present]
            if _missing:
                _full = load_followups()
                _add = _full[_full["Followup_ID"].astype(str).isin(_missing)].copy()
                if len(_add):
                    if "_call_note" not in _add.columns:
                        _add["_call_note"] = ""
                    combined = pd.concat([combined, _add], ignore_index=True, sort=False)
    except Exception:
        reinstate_note = {}   # a reinstatement hiccup must never break the sheet


    # ── De-duplicate the follow-up list: one row per patient (S102, Item 1) ──────
    # A patient can carry several OPEN follow-ups from different visit cycles that
    # all land on the same day's sheet. Staff then see the same person 2–3 times.
    # RULE (owner-confirmed, S102):
    #   • Group by mobile + name (+ diagnosis, so two genuinely different clinical
    #     problems for one patient stay as separate rows).
    #   • Keep ONLY the most recent follow-up cycle = latest Due_Date. Older cycles
    #     are removed from the sheet entirely (NO note — a note would confuse staff).
    #   • EXCEPTION: a reinstated ("call back & complete", amber) row always wins its
    #     group, even if older — that flag means we owe the patient a callback and it
    #     must never be dropped.
    #   • Blank / invalid mobile → group by name only (they are un-callable anyway).
    # Only the FOLLOW-UP list is touched here; Procedure and Watch sections are not.
    # Wrapped so a de-dupe hiccup can NEVER break the sheet — on any error we fall
    # back to the full, un-deduped list.
    try:
        if combined is not None and len(combined):
            _c = combined.copy().reset_index(drop=True)

            def _norm_name(v):
                return " ".join(str(v or "").strip().lower().split())

            def _grp_key(row):
                mob = str(row.get("FU_Mobile_Clean", "") or "").strip()
                nm = _norm_name(row.get("FU_Name_Raw", ""))
                dg = str(row.get("Diagnosis", "") or "").strip().lower()
                if mob and mob.lower() not in ("", "nan", "none"):
                    return ("M", mob, nm, dg)     # real mobile → mobile+name+diagnosis
                return ("N", nm)                  # no mobile → name only

            def _due_sort(v):
                d = parse_date(v)
                return d or date(1900, 1, 1)

            _reinstated_ids = set(reinstate_note.keys()) if reinstate_note else set()

            _keep_idx = []
            _seen = {}
            for _i, _row in _c.iterrows():
                _k = _grp_key(_row)
                _fid = str(_row.get("Followup_ID", ""))
                _is_reins = _fid in _reinstated_ids
                _due = _due_sort(_row.get("Due_Date"))
                if _k not in _seen:
                    _seen[_k] = (_i, _due, _is_reins)
                    continue
                _pi, _pdue, _pr = _seen[_k]
                # reinstated always beats non-reinstated; else newer Due_Date wins
                if _is_reins and not _pr:
                    _seen[_k] = (_i, _due, _is_reins)
                elif _is_reins == _pr and _due > _pdue:
                    _seen[_k] = (_i, _due, _is_reins)
                # otherwise keep the incumbent
            _keep_idx = [v[0] for v in _seen.values()]
            combined = _c.loc[_keep_idx].reset_index(drop=True)
    except Exception:
        pass  # a de-dupe hiccup must never break the sheet — keep full list


    # ── Item 3 (S103): daily cap + drip + 3-strike Hard-to-Reach split ───────────
    # Owner policy (Sessions 109-114):
    #   • DAILY CAP = 120 total follow-up callbacks on the staff sheet.
    #   • Fill order: winnable buckets first (Due Today > Grace Period >
    #     Actionable Missed), in the engine's existing priority order. Whatever
    #     room is left under 120 is back-filled with the OLDEST Probable Dropouts
    #     (drip). When winnable alone >= 120, dropouts get zero room and the
    #     winnable OVERFLOW simply rolls to tomorrow (not shown today; it reappears
    #     next run because its ledger status is unchanged).
    #   • 3 STRIKES: any row with Call_Attempts >= 3 and still no contact
    #     (Call_Resolution not RESOLVE/DECLINE) is pulled OFF the daily list into a
    #     Hard-to-Reach queue (its own tab), carrying name, Clinic ID, mobile,
    #     diagnosis, last-visit date and attempts — so the doctor decides per
    #     patient: keep calling or archive. Not auto-archived.
    #   • The AUDIT workbook is untouched and stays full; only THIS staff sheet is
    #     capped. Wrapped in try/except → any error falls back to the full list; it
    #     can never blank the sheet.
    hard_to_reach_records = []   # filled by the 3-strike split below
    DAILY_CALL_CAP = 120
    STRIKE_LIMIT   = 3
    try:
        if combined is not None and len(combined):
            _w = combined.copy().reset_index(drop=True)

            # last-visit date per patient (read-only from the visit ledger)
            _lastvisit = {}
            try:
                if visits is not None and len(visits):
                    for _, _vr in visits.iterrows():
                        _uid = str(_vr.get("Patient_UID", "") or "").strip()
                        _vd  = parse_date(_vr.get("Visit_Date"))
                        if _uid and _vd:
                            _prev = _lastvisit.get(_uid)
                            if _prev is None or _vd > _prev:
                                _lastvisit[_uid] = _vd
            except Exception:
                _lastvisit = {}

            def _attempts(row):
                try:
                    return int(float(str(row.get("Call_Attempts", "") or 0).strip() or 0))
                except Exception:
                    return 0

            def _no_contact(row):
                # a strike counts only if the row is NOT already resolved/declined
                return str(row.get("Call_Resolution", "") or "").strip().upper() \
                       not in ("RESOLVE", "DECLINE")

            # reinstated rows are protected — never pulled into Hard-to-Reach
            _reins_ids = set(reinstate_note.keys()) if reinstate_note else set()

            # ── Step A: 3-strike split ───────────────────────────────────────────
            _keep_mask = []
            for _i, _row in _w.iterrows():
                _fid = str(_row.get("Followup_ID", ""))
                if (_fid not in _reins_ids
                        and _attempts(_row) >= STRIKE_LIMIT
                        and _no_contact(_row)):
                    _uid = str(_row.get("Patient_UID_Resolved", "") or "").strip()
                    _lv  = _lastvisit.get(_uid)
                    hard_to_reach_records.append({
                        "name":     _row.get("FU_Name_Raw", ""),
                        "clinic":   _row.get("Clinic_Specific_Id_Resolved", ""),
                        "mobile":   str(_row.get("FU_Mobile_Clean", "") or ""),
                        "diag":     clean_diag(_row.get("Diagnosis", "")),
                        "lastvisit": _lv.strftime("%d-%b-%Y") if _lv else "",
                        "attempts": _attempts(_row),
                    })
                    _keep_mask.append(False)
                else:
                    _keep_mask.append(True)
            _w = _w[pd.Series(_keep_mask, index=_w.index)].reset_index(drop=True)

            # ── Step B: 120 cap + drip + roll-to-tomorrow ────────────────────────
            if len(_w) > DAILY_CALL_CAP:
                _WINNABLE = {"Due Today", "Grace Period", "Actionable Missed Follow-Up"}
                _st = _w["Followup_Status"].astype(str)
                _win = _w[_st.isin(_WINNABLE)].copy()
                _drop = _w[~_st.isin(_WINNABLE)].copy()   # Probable Dropout (+ any other)

                # _w already arrived in the engine's priority order (freshest-first
                # + post-op float); the de-dupe preserved that order. So the first
                # N winnable rows ARE the top-priority winnable rows.
                if len(_win) >= DAILY_CALL_CAP:
                    # winnable alone fills the cap → keep top 120, rest roll tomorrow;
                    # dropouts get no room today.
                    combined = _win.head(DAILY_CALL_CAP).reset_index(drop=True)
                else:
                    # room left after winnable → drip in OLDEST dropouts (most overdue)
                    _room = DAILY_CALL_CAP - len(_win)
                    def _od(v):
                        try:
                            return int(float(str(v or 0).strip() or 0))
                        except Exception:
                            return 0
                    if len(_drop):
                        _drop = _drop.assign(_od=_drop["Days_Overdue"].map(_od)) \
                                     .sort_values("_od", ascending=False) \
                                     .drop(columns=["_od"])
                        _drip = _drop.head(_room)
                    else:
                        _drip = _drop.iloc[0:0]
                    combined = pd.concat([_win, _drip], ignore_index=True, sort=False)
            else:
                # already within cap → keep all (minus the 3-strike rows already removed)
                combined = _w
    except Exception:
        pass  # any cap/split hiccup must never break the sheet — keep the de-duped list


    r = 4
    r = section_header(r, "1.  FOLLOW-UP CALLS  —  due today, overdue & upcoming", NAVY)
    r = write_rows(r, fu_records(combined, notes_override=reinstate_note))

    # ── Procedure call-backs ─────────────────────────────────────────────────
    # FIX: call-backs are for procedures done in TODAY'S consultation report
    # (the report just uploaded), to be phoned on the NEXT call day by Shavez —
    # NOT "today − 1". On a Saturday upload the call day rolls to Monday
    # (Sunday is emergency-only), so the sheet is ready for Monday with no
    # Sunday work. `consult_date` is the report's own date; we fall back to the
    # latest visit date in the ledger only if it wasn't supplied.
    if consult_date is None:
        vd_all = [parse_date(s) for s in visits.get("Visit_Date", []) if parse_date(s)]
        consult_date = max(vd_all) if vd_all else today
    call_day = next_call_day(consult_date)

    proc_recs = []
    _rendered_proc_uids = set()
    p = visits.copy()
    if "Had_Procedure" in p.columns and len(p):
        p = p[(p["Visit_Date"].astype(str) == date_to_str(consult_date))
              & (p["Had_Procedure"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"]))]
        for _, row in p.iterrows():
            proc_recs.append({
                "pr": "A", "name": row.get("Patient_Name", ""),
                "mobile": str(row.get("Mobile_Clean", "")),
                "info": "Procedure welfare call", "date": fmt_d(row.get("Visit_Date")),
                "od": "", "status": "Procedure call-back", "vip": False,
                "key": row.get("Visit_ID", ""),
            })
            _rendered_proc_uids.add(str(row.get("Patient_UID", "")).strip())
    _sun = "  (Sunday skipped — emergency only)" if (consult_date.weekday() == 5) else ""
    r = section_header(
        r,
        f"2.  PROCEDURE CALL-BACKS  —  done {consult_date.strftime('%d-%b')}, "
        f"call on {call_day.strftime('%a %d-%b')} by {PROCEDURE_CALLER}{_sun}",
        BROWN)
    # ── Safety net + staleness (added 18-Jun-2026) ───────────────────────────
    # Shout any gap at the TOP of the section so a missed procedure call-back can
    # never pass silently again. (a) staleness: report older than yesterday;
    # (b) reconcile: every procedure on the Day Revenue tab must appear here.
    _warn = []
    try:
        _lag = (today - consult_date).days
        if _lag > 1:
            _warn.append("STALE REPORT: visits are from " + consult_date.strftime("%d-%b")
                         + " (" + str(_lag) + " days ago) - these call-backs may already be overdue.")
        if day_revenue_df is not None and len(day_revenue_df):
            _dd = day_revenue_df.copy()
            if "Procedure_Disc" not in _dd.columns: _dd["Procedure_Disc"] = 0
            if "Manual_Procedure" not in _dd.columns: _dd["Manual_Procedure"] = False
            _pm = ((pd.to_numeric(_dd.get("Procedure", 0), errors="coerce").fillna(0) > 0)
                   | (pd.to_numeric(_dd.get("Procedure_Disc", 0), errors="coerce").fillna(0) > 0)
                   | (_dd["Manual_Procedure"].astype(bool)))
            for _, _br in _dd[_pm].iterrows():
                _uid = str(_br.get("Patient_UID", "")).strip()
                if _uid and _uid not in _rendered_proc_uids:
                    _warn.append("PROCEDURE WITHOUT CALL-BACK: "
                                 + str(_br.get("Patient_Name", "") or _uid)
                                 + " (Clinic ID " + str(_br.get("Clinic_Specific_Id", "")) + ") "
                                 + "was billed a procedure but is not in this list - verify & call.")
    except Exception:
        pass  # a recon hiccup must never break the sheet
    for _wl in _warn:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        _c = ws.cell(r, 1, "\u26a0  " + _wl)
        _c.font = Font(AR, bold=True, size=10, color="B00000")
        _c.fill = PatternFill("solid", fgColor="FFF2CC")
        _c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    r = write_rows(r, proc_recs, default_caller=PROCEDURE_CALLER)

    # ── Section 3: Unreachable / Watch (escalated after repeated no-contact) ──
    if watch_df is not None and len(watch_df):
        r = section_header(
            r,
            f"3.  UNREACHABLE / WATCH  —  {len(watch_df)} follow-up(s) after "
            f"{ESCALATE_AFTER}+ failed calls (verify number, WABA, or drop)",
            "C00000")
        r = write_rows(r, fu_records(watch_df))

    ws.column_dimensions[chr(64 + KEY_COL)].hidden = True
    ws.freeze_panes = "A4"

    # ── Vacation Notice List tab ──────────────────────────────────────────────
    wsv = wb.create_sheet("Vacation Notice List")
    wsv.append(list(df_vac_out.columns))
    for cell in wsv[1]:
        cell.font = Font(AR, bold=True, color="1F4E79")
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for _, row in df_vac_out.iterrows():
        wsv.append([str(v) if pd.notna(v) else "" for v in row.values])
    for col in wsv.columns:
        ln = max([len(str(c.value)) for c in col if c.value] + [10])
        wsv.column_dimensions[col[0].column_letter].width = min(ln + 2, 40)
    wsv.freeze_panes = "A2"

    # ── Settled Follow-Ups tab ────────────────────────────────────────────────
    # Disposition of patients who were on the call list but have dropped off
    # because they were settled — reached, declined, returned, or holding. This
    # answers "what happened to yesterday's calls?" so nothing vanishes silently.
    if full_fu_ledger is not None and len(full_fu_ledger):
        _build_settled_sheet(wb, full_fu_ledger, today, AR, border)

    # ── Hard-to-Reach tab (Item 3, S103) ──────────────────────────────────────
    # Patients pulled OFF the daily call list after 3 no-contact attempts. Shown
    # so staff see they are parked (not lost) and the doctor can decide per
    # patient: keep calling, or archive. Carries the four LOCAL fields now
    # (last-visit date, mobile, Clinic ID, diagnosis); recording/transcript links
    # are a planned fast follow-up (owner choice 'b', S111).
    try:
        wsh = wb.create_sheet("Hard-to-Reach")
        _htr_title = ("HARD-TO-REACH  —  %d patient(s) off the call list after %d+ "
                      "no-contact attempts. DOCTOR TO DECIDE: keep calling or archive."
                      % (len(hard_to_reach_records), STRIKE_LIMIT))
        _htr_cols = ["Patient", "Clinic ID", "Mobile", "Diagnosis",
                     "Last Visit", "Attempts"]
        wsh.merge_cells(start_row=1, start_column=1, end_row=1,
                        end_column=len(_htr_cols))
        _tc = wsh.cell(1, 1, _htr_title)
        _tc.font = Font(AR, bold=True, size=11, color="FFFFFF")
        _tc.fill = PatternFill("solid", fgColor="C00000")
        _tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for _ci, _cn in enumerate(_htr_cols, 1):
            _hc = wsh.cell(2, _ci, _cn)
            _hc.font = Font(AR, bold=True, color="1F4E79")
            _hc.fill = PatternFill("solid", fgColor="F2F2F2")
            _hc.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True)
            _hc.border = border
        _rr = 3
        for _rec in hard_to_reach_records:
            _vals = [_rec.get("name", ""), _rec.get("clinic", ""),
                     _rec.get("mobile", ""), _rec.get("diag", ""),
                     _rec.get("lastvisit", ""), _rec.get("attempts", "")]
            for _ci, _v in enumerate(_vals, 1):
                _dc = wsh.cell(_rr, _ci, "" if _v is None else str(_v))
                _dc.font = Font(AR, size=10)
                _dc.alignment = Alignment(horizontal="left", vertical="center")
                _dc.border = border
            _rr += 1
        _htr_widths = [22, 12, 14, 30, 13, 9]
        for _ci, _w in enumerate(_htr_widths, 1):
            wsh.column_dimensions[chr(64 + _ci)].width = _w
        wsh.freeze_panes = "A3"
    except Exception:
        pass  # a Hard-to-Reach tab hiccup must never break the sheet

    # ── Day Revenue tab ───────────────────────────────────────────────────────
    if day_revenue_df is not None and len(day_revenue_df):
        _build_day_revenue_sheet(wb, day_revenue_df, consult_date, AR, border)

    wb.save(out_path)


def _build_settled_sheet(wb, fu_ledger: pd.DataFrame, today: date, AR: str, border):
    """Read-only disposition of recently-due follow-ups now OFF the call list:
    reached / declined / returned / holding. Window: last SETTLED_WINDOW_DAYS by
    due date. Built from the recomputed ledger's summary + status columns."""
    from openpyxl.styles import PatternFill, Font, Alignment
    WINDOW = SETTLED_WINDOW_DAYS
    NAVY = "1F3864"; LIGHTBLUE = "D6E4F0"; GREEN = "C6EFCE"; GREY = "E0E0E0"; AMBER = "FFE0B2"
    LABELS = {
        "YES":          ("Reached — patient coming", "green"),
        "Tt COMPLETE":  ("Treatment complete (confirming)", "green"),
        "MED HERE":     ("Medicine — bought here", "green"),
        "MED OUTSIDE":  ("Medicine — bought outside", "green"),
        "MEDICIN LI H": ("Medicine — self-managing", "green"),
        "DIKHA CHUKE":  ("Already seen elsewhere", "green"),
        "NO":           ("Declined", "grey"),
        "LATER":        ("Said call later — holding", "amber"),
    }
    rows = []
    for _, r in fu_ledger.iterrows():
        dd = parse_date(r.get("Due_Date"))
        if not dd:
            continue
        age = (today - dd).days
        if age < 0 or age > WINDOW:
            continue
        st = str(r.get("Followup_Status", "") or "")
        cres = str(r.get("Call_Resolution", "") or "")
        resp = str(r.get("Last_Response", "") or "").strip()
        by = str(r.get("Last_Called_By", "") or "")
        when = str(r.get("Last_Call_Date", "") or "")
        if st.startswith("Returned"):
            rv = parse_date(r.get("Matched_Return_Visit_Date", ""))
            outcome = "Returned on " + rv.strftime("%d-%b") if rv else "Returned to clinic"
            tone, by, when = "green", "", (r.get("Matched_Return_Visit_Date", "") or when)
        elif cres in ("RESOLVE", "DECLINE", "RESCHEDULE"):
            outcome, tone = LABELS.get(resp, (resp or cres, "grey"))
        else:
            continue   # still open / on the call list — not settled
        rows.append({"due": dd, "name": str(r.get("FU_Name_Raw") or r.get("Standard_Name") or ""),
                     "mobile": str(r.get("FU_Mobile_Clean", "")),
                     "cid": str(r.get("Clinic_Specific_Id_Resolved", "")),
                     "outcome": outcome, "by": by, "when": when, "tone": tone})
    rows.sort(key=lambda x: x["due"], reverse=True)

    n_ret = sum(1 for x in rows if x["outcome"].startswith("Returned"))
    n_dec = sum(1 for x in rows if x["tone"] == "grey")
    n_hold = sum(1 for x in rows if x["tone"] == "amber")
    n_reach = len(rows) - n_ret - n_dec - n_hold

    ws = wb.create_sheet("Settled Follow-Ups")
    headers = ["Due", "Patient", "Mobile", "Clinic ID", "Outcome", "Handled By", "When"]
    ws.merge_cells("A1:G1")
    c = ws.cell(1, 1, f"SETTLED FOLLOW-UPS (last {WINDOW} days)  —  {len(rows)} off the call list: "
                      f"{n_ret} returned · {n_reach} reached · {n_dec} declined · {n_hold} holding")
    c.font = Font(AR, bold=True, size=11, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 22
    for j, h in enumerate(headers, 1):
        cc = ws.cell(2, j, h); cc.font = Font(AR, bold=True, size=10)
        cc.fill = PatternFill("solid", fgColor=LIGHTBLUE)
        cc.alignment = Alignment(horizontal="center", vertical="center"); cc.border = border
    tone_fill = {"green": GREEN, "grey": GREY, "amber": AMBER}
    rr = 3
    for x in rows:
        wv = parse_date(x["when"])
        vals = [x["due"].strftime("%d-%b"), x["name"], x["mobile"], x["cid"],
                x["outcome"], x["by"], wv.strftime("%d-%b") if wv else x["when"]]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(rr, j, v); cc.font = Font(AR, size=10); cc.border = border
            cc.alignment = Alignment(horizontal="center" if j in (1, 4, 7) else "left", vertical="center")
        ws.cell(rr, 5).fill = PatternFill("solid", fgColor=tone_fill.get(x["tone"], GREY))
        rr += 1
    if not rows:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
        ws.cell(3, 1, f"— nothing settled in the last {WINDOW} days —").font = Font(AR, italic=True, size=10)
    for i, w in enumerate([10, 22, 13, 11, 32, 12, 10], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"


def _build_day_revenue_sheet(wb, dr: pd.DataFrame, report_date: date, AR: str, border):
    """A hand-to-reception revenue summary for the day's consultation report:
    consultation / X-ray / procedure subtotals, paid-consultation count with
    each name flagged ▲/▼ vs the ₹600 default, free-revisit vs free-case lists,
    a cash-vs-online split, and the day's grand total. Built from the rich
    per-visit revenue frame produced by revenue.parse_day_revenue()."""
    from openpyxl.styles import PatternFill, Font, Alignment

    NAVY = "1F3864"; HEADFILL = "D6E4F0"; GREEN = "548235"
    money = '#,##0'

    ws = wb.create_sheet("Day Revenue")
    widths = [5, 26, 11, 12, 12, 12, 12, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    rd = report_date.strftime("%d-%b-%Y") if report_date else ""
    ncol = len(widths)
    last = chr(64 + ncol)

    def title(text, fill=NAVY, color="FFFFFF", size=13):
        nonlocal row
        ws.merge_cells(f"A{row}:{last}{row}")
        c = ws.cell(row, 1, text)
        c.font = Font(AR, bold=True, size=size, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

    def head(cells):
        nonlocal row
        for j, h in enumerate(cells, 1):
            c = ws.cell(row, j, h)
            c.font = Font(AR, bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=HEADFILL)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        row += 1

    def line(cells, bold=False, fills=None, num_cols=()):
        nonlocal row
        for j, v in enumerate(cells, 1):
            c = ws.cell(row, j, v)
            c.font = Font(AR, bold=bold, size=10)
            c.border = border
            if j in num_cols and isinstance(v, (int, float)):
                c.number_format = money
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
            if fills and j in fills:
                c.fill = PatternFill("solid", fgColor=fills[j])
        row += 1

    row = 1
    title(f"DR. MANOJ AGARWAL CLINIC  —  DAY REVENUE  ·  {rd}")

    # numeric views
    cons = dr[dr["Consultation"] > 0].copy()
    xray = dr[dr["Xray"] > 0].copy()
    # Procedures: include billed (gross>0), billed-then-discounted (disc>0, net may
    # be ₹0), AND manually-marked free/₹0 procedures. "Procedure" column is gross;
    # net = gross − discount; the procedure SUBTOTAL uses net (revenue actually due).
    if "Manual_Procedure" not in dr.columns:
        dr["Manual_Procedure"] = False
    proc = dr[(dr["Procedure"] > 0) | (dr["Procedure_Disc"] > 0) | (dr["Manual_Procedure"])].copy()
    proc["_net"] = proc["Procedure"] - proc["Procedure_Disc"]
    free_revisit = dr[dr["Case_Type"] == "Free Revisit"]
    free_case    = dr[dr["Case_Type"] == "Free / Concession Case"]

    cons_total = float(cons["Consultation"].sum())
    xray_total = float(xray["Xray"].sum())
    proc_total = float(proc["_net"].sum()) if len(proc) else 0.0
    grand = cons_total + xray_total + proc_total
    cash_total   = float(dr["Cash"].sum())
    online_total = float(dr["Online"].sum())
    pending_total = float(dr["Pending"].sum())
    morn = int((dr["Shift"].str.lower() == "morning").sum())
    eve  = int((dr["Shift"].str.lower() == "evening").sum())

    # ── Summary block ─────────────────────────────────────────────────────────
    row += 0
    title("SUMMARY", fill="1F4E79", size=11)
    head(["", "Head", "Count", "Amount (₹)", "", "", "", ""])
    line(["", "Consultations (paid)", int(len(cons)), cons_total, "", "", "", ""], num_cols=(4,))
    line(["", "X-ray (lab column)",  int(len(xray)), xray_total, "", "", "", ""], num_cols=(4,))
    line(["", "Procedures",          int(len(proc)), proc_total, "", "", "", ""], num_cols=(4,))
    line(["", "GRAND TOTAL (collected billing)", int(len(dr)), grand, "", "", "", ""],
         bold=True, num_cols=(4,), fills={2: "FFF2CC", 4: "FFF2CC"})
    row += 1
    line(["", "Cash", "", cash_total, "", "", "", ""], num_cols=(4,), fills={2: "E2EFDA"})
    line(["", "Online / UPI", "", online_total, "", "", "", ""], num_cols=(4,), fills={2: "DDEBF7"})
    if pending_total:
        line(["", "Pending (unpaid at counter)", "", pending_total, "", "", "", ""],
             num_cols=(4,), fills={2: "FCE4EC"})
    line(["", f"Shift split — Morning: {morn}   Evening: {eve}", "", "", "", "", "", ""])
    row += 1

    # ── Paid consultations, each name ▲/▼ vs ₹600 ────────────────────────────
    title("PAID CONSULTATIONS  (default ₹600 — ▲ above / ▼ below flagged)", fill="1F4E79", size=11)
    head(["S.N", "Patient", "Clinic ID", "Consult ₹", "vs ₹600", "Mode", "Shift", "Notes"])
    sn = 1
    for _, r in cons.sort_values("Consultation", ascending=False).iterrows():
        amt = float(r["Consultation"])
        flag = "▲ +%d" % (amt - DEFAULT_CONSULT_FEE) if amt > DEFAULT_CONSULT_FEE else \
               ("▼ −%d" % (DEFAULT_CONSULT_FEE - amt) if amt < DEFAULT_CONSULT_FEE else "= 600")
        fills = {5: ("FCE4EC" if amt < DEFAULT_CONSULT_FEE else
                     ("E2EFDA" if amt > DEFAULT_CONSULT_FEE else None))}
        fills = {k: v for k, v in fills.items() if v}
        line([sn, r["Patient_Name"], r["Clinic_Specific_Id"], amt, flag,
              r["Mode"], r["Shift"], ""], num_cols=(4,), fills=fills)
        sn += 1
    line(["", "Subtotal", int(len(cons)), cons_total, "", "", "", ""],
         bold=True, num_cols=(4,), fills={2: "FFF2CC", 4: "FFF2CC"})
    row += 1

    # ── X-ray ─────────────────────────────────────────────────────────────────
    if len(xray):
        title("X-RAY  (billed under Docterz 'Laboratory' column)", fill="1F4E79", size=11)
        head(["S.N", "Patient", "Clinic ID", "X-ray ₹", "Mode", "Shift", "", ""])
        sn = 1
        for _, r in xray.sort_values("Xray", ascending=False).iterrows():
            line([sn, r["Patient_Name"], r["Clinic_Specific_Id"], float(r["Xray"]),
                  r["Mode"], r["Shift"], "", ""], num_cols=(4,))
            sn += 1
        line(["", "Subtotal", int(len(xray)), xray_total, "", "", "", ""],
             bold=True, num_cols=(4,), fills={2: "FFF2CC", 4: "FFF2CC"})
        row += 1

    # ── Procedures ──────────────────────────────────────────────────────────
    if len(proc):
        title("PROCEDURES  (incl. ₹0 / free — e.g. Ayushman cashless)", fill="1F4E79", size=11)
        head(["S.N", "Patient", "Clinic ID", "Procedure ₹ (net)", "Note", "Shift", "", ""])
        sn = 1
        for _, r in proc.sort_values("_net", ascending=False).iterrows():
            net = float(r["_net"])
            disc = float(r["Procedure_Disc"])
            if bool(r.get("Manual_Procedure")) and float(r["Procedure"]) == 0 and disc == 0:
                note = "free / ₹0 (marked)"
            elif disc > 0:
                note = f"{r['Mode']} (billed {float(r['Procedure']):.0f}, disc {disc:.0f})"
            else:
                note = r["Mode"]
            fills = {4: "FCE4EC"} if net == 0 else {}
            line([sn, r["Patient_Name"], r["Clinic_Specific_Id"], net,
                  note, r["Shift"], "", ""], num_cols=(4,), fills=fills)
            sn += 1
        line(["", "Subtotal (net)", int(len(proc)), proc_total, "", "", "", ""],
             bold=True, num_cols=(4,), fills={2: "FFF2CC", 4: "FFF2CC"})
        row += 1

    # ── Free revisits (5-day window) and free / concession cases ─────────────
    title("FREE REVISITS  (within 5-day window — fee already paid, expected — NOT lost revenue)",
          fill="548235", size=11)
    if len(free_revisit):
        head(["S.N", "Patient", "Clinic ID", "Shift", "", "", "", ""])
        sn = 1
        for _, r in free_revisit.iterrows():
            line([sn, r["Patient_Name"], r["Clinic_Specific_Id"], r["Shift"], "", "", "", ""])
            sn += 1
    else:
        line(["", "— none —", "", "", "", "", "", ""])
    row += 1

    title("FREE / CONCESSION CASES  (genuinely complimentary — CC 0 / Ayushman / staff / VIP)",
          fill="806000", size=11)
    if len(free_case):
        head(["S.N", "Patient", "Clinic ID", "Shift", "", "", "", ""])
        sn = 1
        for _, r in free_case.iterrows():
            line([sn, r["Patient_Name"], r["Clinic_Specific_Id"], r["Shift"], "", "", "", ""])
            sn += 1
    else:
        line(["", "— none —", "", "", "", "", "", ""])

    ws.freeze_panes = "A2"


def build_output_excel(
    fu_ledger: pd.DataFrame,
    visits: pd.DataFrame,
    master: pd.DataFrame,
    new_patients: int,
    new_visits: int,
    new_followups: int,
    today: date,
    output_path: str,
    staff_output_path: str | None = None,
    consult_date: date | None = None,
    day_revenue_df: pd.DataFrame | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── colour palette ────────────────────────────────────────────────────────
    RED    = "FFC7CE"   # Probable Dropout
    ORANGE = "FFEB9C"   # Actionable Missed
    YELLOW = "FFFF99"   # Grace Period / Due Today
    GREEN  = "C6EFCE"   # Returned
    BLUE   = "DDEBF7"   # Not Due
    GREY   = "F2F2F2"   # Header
    WHITE  = "FFFFFF"
    HEADER_FONT_COLOR = "1F4E79"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLORS = {
        "Probable Dropout":              RED,
        "Actionable Missed Follow-Up":   ORANGE,
        "Grace Period":                  YELLOW,
        "Due Today":                     YELLOW,
        "Not Due":                       BLUE,
        "Returned Early":                GREEN,
        "Returned On Time":              GREEN,
        "Returned Late":                 "E2EFDA",
        "Identity Unresolved":           "FCE4D6",
        "Ambiguous Mobile":              "FCE4D6",
        "Invalid Mobile / No Contact":   "FCE4D6",
        "Expired Unresolved":            "D9D9D9",
    }

    def style_header(ws, row=1, fill_color=GREY):
        for cell in ws[row]:
            cell.font      = Font(bold=True, color=HEADER_FONT_COLOR)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border

    def autofit(ws, min_w=10, max_w=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

    def write_df(ws, df, status_col=None):
        """Write dataframe to worksheet with optional status-based row colouring."""
        ws.append(list(df.columns))
        style_header(ws)
        for _, row in df.iterrows():
            ws.append([str(v) if pd.notna(v) else "" for v in row.values])
            if status_col and status_col in df.columns:
                status = str(row.get(status_col, ""))
                color  = STATUS_COLORS.get(status, WHITE)
                for cell in ws[ws.max_row]:
                    cell.fill   = PatternFill("solid", fgColor=color)
                    cell.border = border
        autofit(ws)
        ws.freeze_panes = "A2"

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    status_counts = fu_ledger["Followup_Status"].value_counts().to_dict()

    def sc(k): return status_counts.get(k, 0)

    open_fu = sc("Not Due") + sc("Due Today") + sc("Grace Period") + \
              sc("Actionable Missed Follow-Up") + sc("Probable Dropout")

    metrics = [
        ("", "FOLLOW-UP TRACKING DASHBOARD", ""),
        ("", f"Report Date: {date_to_str(today)}  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", ""),
        ("", "", ""),
        ("METRIC", "COUNT", ""),
        ("Total follow-up obligations (all time)", len(fu_ledger), ""),
        ("Open follow-ups (not yet returned)", open_fu, ""),
        ("", "", ""),
        ("── STATUS BREAKDOWN ──", "", ""),
        ("Not Due (future dated)", sc("Not Due"), ""),
        ("Due Today", sc("Due Today"), ""),
        ("Grace Period (1–3 days overdue)", sc("Grace Period"), ""),
        ("Actionable Missed Follow-Up (4–10 days)", sc("Actionable Missed Follow-Up"), ""),
        ("Probable Dropout (>10 days)", sc("Probable Dropout"), ""),
        (f"Expired Unresolved (>{EXPIRY_DAYS} days, archived)", sc("Expired Unresolved"), ""),
        ("", "", ""),
        ("── RETURNED ──", "", ""),
        ("Returned Early", sc("Returned Early"), ""),
        ("Returned On Time", sc("Returned On Time"), ""),
        ("Returned Late", sc("Returned Late"), ""),
        ("", "", ""),
        ("── IDENTITY ISSUES ──", "", ""),
        ("Identity Unresolved", sc("Identity Unresolved"), ""),
        ("Ambiguous Mobile", sc("Ambiguous Mobile"), ""),
        ("Invalid Mobile / No Contact", sc("Invalid Mobile / No Contact"), ""),
        ("", "", ""),
        ("── TODAY'S PROCESSING ──", "", ""),
        ("New patients added to Master", new_patients, ""),
        ("New visits appended to Ledger", new_visits, ""),
        ("New follow-up obligations added", new_followups, ""),
        ("", "", ""),
        ("── MASTER STATS ──", "", ""),
        ("Total patients in Master", len(master), ""),
        ("Patients with shared mobile (family)", len(master[master["Identity_Status"] == "Shared Mobile"]), ""),
        ("Patients with no/invalid mobile", len(master[master["Identity_Status"] == "No/Invalid Mobile"]), ""),
    ]

    for m in metrics:
        ws_dash.append(list(m))

    # Style dashboard
    for row in ws_dash.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.row == 1 or (cell.value and str(cell.value).startswith("──")):
                cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
                cell.fill = PatternFill("solid", fgColor="DDEBF7")

    ws_dash.column_dimensions["A"].width = 45
    ws_dash.column_dimensions["B"].width = 12
    ws_dash.freeze_panes = "A5"

    # ── Sheet 2: Staff Action Today ───────────────────────────────────────────
    ACTION_STATUSES = {
        "Probable Dropout", "Actionable Missed Follow-Up", "Grace Period",
        "Due Today", "Ambiguous Mobile", "Identity Unresolved",
        "Invalid Mobile / No Contact"
    }
    STATUS_ORDER = [
        "Probable Dropout", "Actionable Missed Follow-Up", "Grace Period",
        "Due Today", "Ambiguous Mobile", "Identity Unresolved",
        "Invalid Mobile / No Contact"
    ]
    order_map = {s: i for i, s in enumerate(STATUS_ORDER)}

    df_action = fu_ledger[fu_ledger["Followup_Status"].isin(ACTION_STATUSES)].copy()
    df_action["_sort"] = df_action["Followup_Status"].map(lambda x: order_map.get(x, 99))
    df_action = df_action.sort_values(["_sort", "Due_Date"]).drop(columns=["_sort"])

    # ── enrich with diagnosis layer (Diagnosis / Priority / VIP) ──────────────
    # P0-03: keyed by Patient_UID. Mobile fallback ONLY for High-confidence
    # (unique mobile) rows — never shared/ambiguous, where a relative's
    # diagnosis could be attached to the wrong patient.
    diag_uid, diag_mob = load_diagnosis_lookup()
    _PRI_RANK = {"A": 0, "B": 1, "C": 2, "D": 3, "": 4}
    _EMPTY_DIAG = {"diag": "", "priority": "", "vip": False, "taxonomy": ""}

    def _diag_for(row):
        uid = str(row.get("Patient_UID_Resolved", "")).strip()
        if uid and uid in diag_uid:
            return diag_uid[uid]
        if str(row.get("Identity_Confidence", "")) == "High":
            rec = diag_mob.get(str(row.get("FU_Mobile_Clean", "")).strip())
            if rec:
                return rec
        return _EMPTY_DIAG

    _diag_rows = df_action.apply(_diag_for, axis=1) if len(df_action) else []
    df_action["Diagnosis"] = [r["diag"] for r in _diag_rows] if len(df_action) else ""
    df_action["Priority"]  = [r["priority"] for r in _diag_rows] if len(df_action) else ""
    df_action["VIP_Flag"]  = ["VIP" if r["vip"] else "" for r in _diag_rows] if len(df_action) else ""
    # Taxonomy → message-category tag (provision for later WABA; no sending)
    df_action["Message_Category"] = [
        message_category_for(r["diag"], r.get("taxonomy", "")) for r in _diag_rows
    ] if len(df_action) else ""

    # ── vacation notice flag ──────────────────────────────────────────────────
    vacations = load_vacations()
    df_action["Vacation_Notice"] = df_action["Due_Date"].map(
        lambda d: vacation_notice_for(parse_date(d), vacations)) if len(df_action) else ""

    # ── upcoming list: follow-ups not yet due, confidently identified, so staff ──
    # can call them on the follow-up morning. Include ALL "Not Due" (not just
    # due-tomorrow) so upcoming patients never fall in a gap regardless of which
    # day the tracker is run. The month-export guard keeps Not Due near-term.
    df_reminder = fu_ledger[
        (fu_ledger["Followup_Status"] == "Not Due")
        & (fu_ledger["Identity_Confidence"].isin(["High", "Medium"]))
    ].copy()
    if len(df_reminder):
        _rem_rows = df_reminder.apply(_diag_for, axis=1)
        df_reminder["Diagnosis"] = [r["diag"] for r in _rem_rows]
        df_reminder["Priority"]  = [r["priority"] for r in _rem_rows]
        df_reminder["VIP_Flag"]  = ["VIP" if r["vip"] else "" for r in _rem_rows]
    else:
        for c in ("Diagnosis", "Priority", "VIP_Flag"):
            df_reminder[c] = ""

    diag_lookup = diag_uid or diag_mob
    if diag_lookup:
        # within each follow-up status band, surface VIP first, then Priority A→D
        df_action["_vip"] = (df_action["VIP_Flag"] == "VIP").map(lambda v: 0 if v else 1)
        df_action["_pri"] = df_action["Priority"].map(lambda p: _PRI_RANK.get(p, 4))
        df_action["_sort"] = df_action["Followup_Status"].map(lambda x: order_map.get(x, 99))
        df_action = (df_action.sort_values(["_sort", "_vip", "_pri", "Due_Date"])
                              .drop(columns=["_sort", "_vip", "_pri"]))

    action_cols = [
        "VIP_Flag", "FU_Name_Raw", "FU_Mobile_Clean", "Clinic_Specific_Id_Resolved",
        "Patient_UID_Resolved", "Diagnosis", "Priority", "Due_Date", "Days_Overdue",
        "Followup_Status", "Suggested_Action", "Vacation_Notice", "Message_Category",
        "Identity_Confidence", "Identity_Issue"
    ]
    action_display = {
        "VIP_Flag":                    "VIP",
        "FU_Name_Raw":                 "Patient Name",
        "FU_Mobile_Clean":             "Mobile",
        "Clinic_Specific_Id_Resolved": "Clinic ID",
        "Patient_UID_Resolved":        "Patient UID",
        "Diagnosis":                   "Diagnosis",
        "Priority":                    "Priority",
        "Due_Date":                    "Due Date",
        "Days_Overdue":                "Days Overdue",
        "Followup_Status":             "Status",
        "Suggested_Action":            "Suggested Action",
        "Vacation_Notice":             "Doctor Unavailable",
        "Message_Category":            "Msg Category",
        "Identity_Confidence":         "ID Confidence",
        "Identity_Issue":              "Identity Issue"
    }
    df_action_out = df_action[action_cols].rename(columns=action_display)
    ws_action = wb.create_sheet("Staff Action Today")
    write_df(ws_action, df_action_out, status_col="Status")

    # highlight VIP cells (reception + caller awareness) and colour priority
    _PRI_FILL = {"A": "FFC7CE", "B": "FFEB9C", "C": "C6EFCE", "D": "E0E0E0"}
    hdr = [c.value for c in ws_action[1]]
    vip_idx = hdr.index("VIP") + 1 if "VIP" in hdr else None
    pri_idx = hdr.index("Priority") + 1 if "Priority" in hdr else None
    for row in ws_action.iter_rows(min_row=2):
        if vip_idx and row[vip_idx - 1].value == "VIP":
            row[vip_idx - 1].font = Font(bold=True, color="C00000")
        if pri_idx and row[pri_idx - 1].value in _PRI_FILL:
            row[pri_idx - 1].fill = PatternFill("solid", fgColor=_PRI_FILL[row[pri_idx - 1].value])
            row[pri_idx - 1].alignment = Alignment(horizontal="center")

    # ── Sheet 2b: Vacation Notice List ────────────────────────────────────────
    # Patients whose follow-up falls inside a declared vacation window and who
    # have a contactable, confidently-identified mobile (High/Medium only).
    # Staff use this to inform patients the doctor is unavailable; later the
    # same list feeds the WABA unavailability template.
    VAC_OPEN = {"Not Due", "Due Today", "Grace Period", "Actionable Missed Follow-Up"}
    df_vac = fu_ledger[
        fu_ledger["Followup_Status"].isin(VAC_OPEN)
        & fu_ledger["Identity_Confidence"].isin(["High", "Medium"])
    ].copy()
    if len(df_vac):
        df_vac["Vacation_Notice"] = df_vac["Due_Date"].map(
            lambda d: vacation_notice_for(parse_date(d), vacations))
        df_vac = df_vac[df_vac["Vacation_Notice"] != ""]
    if len(df_vac):
        df_vac = df_vac.sort_values("Due_Date")
        df_vac_out = df_vac[[
            "FU_Name_Raw", "FU_Mobile_Clean", "Clinic_Specific_Id_Resolved",
            "Patient_UID_Resolved", "Due_Date", "Followup_Status",
            "Vacation_Notice", "Identity_Confidence"
        ]].rename(columns={
            "FU_Name_Raw": "Patient Name", "FU_Mobile_Clean": "Mobile",
            "Clinic_Specific_Id_Resolved": "Clinic ID",
            "Patient_UID_Resolved": "Patient UID", "Due_Date": "Due Date",
            "Followup_Status": "Status", "Vacation_Notice": "Doctor Unavailable",
            "Identity_Confidence": "ID Confidence",
        })
    else:
        df_vac_out = pd.DataFrame(columns=[
            "Patient Name", "Mobile", "Clinic ID", "Patient UID", "Due Date",
            "Status", "Doctor Unavailable", "ID Confidence"])
    ws_vac = wb.create_sheet("Vacation Notice List")
    write_df(ws_vac, df_vac_out, status_col="Status")

    # ── Sheet 3: Probable Dropouts ────────────────────────────────────────────
    df_dropout = fu_ledger[fu_ledger["Followup_Status"] == "Probable Dropout"].copy()
    df_dropout = df_dropout.sort_values("Days_Overdue",
        key=lambda x: pd.to_numeric(x, errors='coerce'), ascending=False)
    dropout_cols = [
        "FU_Name_Raw", "FU_Mobile_Clean", "Clinic_Specific_Id_Resolved",
        "Patient_UID_Resolved", "Due_Date", "Days_Overdue", "Suggested_Action"
    ]
    dropout_display = {
        "FU_Name_Raw":                 "Patient Name",
        "FU_Mobile_Clean":             "Mobile",
        "Clinic_Specific_Id_Resolved": "Clinic ID",
        "Patient_UID_Resolved":        "Patient UID",
        "Due_Date":                    "Due Date",
        "Days_Overdue":                "Days Overdue",
        "Suggested_Action":            "Action"
    }
    df_dropout_out = df_dropout[[c for c in dropout_cols if c in df_dropout.columns]].rename(columns=dropout_display)
    ws_dropout = wb.create_sheet("Probable Dropouts")
    write_df(ws_dropout, df_dropout_out, status_col=None)
    for row in ws_dropout.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=RED)
            cell.border = border

    # ── Sheet 4: Returned Late ────────────────────────────────────────────────
    df_late = fu_ledger[fu_ledger["Followup_Status"] == "Returned Late"].copy()
    late_cols = [
        "FU_Name_Raw", "FU_Mobile_Clean", "Clinic_Specific_Id_Resolved",
        "Due_Date", "Matched_Return_Visit_Date", "Return_Delay_Days"
    ]
    late_display = {
        "FU_Name_Raw":                 "Patient Name",
        "FU_Mobile_Clean":             "Mobile",
        "Clinic_Specific_Id_Resolved": "Clinic ID",
        "Due_Date":                    "Due Date",
        "Matched_Return_Visit_Date":   "Returned On",
        "Return_Delay_Days":           "Delay (days)"
    }
    df_late_out = df_late[[c for c in late_cols if c in df_late.columns]].rename(columns=late_display)
    ws_late = wb.create_sheet("Returned Late")
    write_df(ws_late, df_late_out, status_col=None)

    # ── Sheet 5: Identity Problems ────────────────────────────────────────────
    id_problem_statuses = {"Identity Unresolved", "Ambiguous Mobile", "Invalid Mobile / No Contact"}
    df_id = fu_ledger[fu_ledger["Followup_Status"].isin(id_problem_statuses)].copy()
    id_cols = [
        "FU_Name_Raw", "FU_Mobile_Raw", "FU_Mobile_Clean",
        "Identity_Confidence", "Identity_Issue", "Suggested_Action"
    ]
    id_display = {
        "FU_Name_Raw":         "Follow-Up Name",
        "FU_Mobile_Raw":       "Mobile (Raw)",
        "FU_Mobile_Clean":     "Mobile (Cleaned)",
        "Identity_Confidence": "Problem Type",
        "Identity_Issue":      "Detail",
        "Suggested_Action":    "Suggested Fix"
    }
    df_id_out = df_id[[c for c in id_cols if c in df_id.columns]].rename(columns=id_display)
    ws_id = wb.create_sheet("Identity Problems")
    write_df(ws_id, df_id_out, status_col=None)

    # Note: Shared Mobile (family numbers) are intentional — not listed as problems

    # ── Sheet 6: Follow-Up Ledger (full) ─────────────────────────────────────
    ws_fu = wb.create_sheet("Follow-Up Ledger")
    write_df(ws_fu, fu_ledger, status_col="Followup_Status")

    # ── Sheet 7: Visit Ledger (full) ──────────────────────────────────────────
    ws_vis = wb.create_sheet("Visit Ledger")
    write_df(ws_vis, visits)

    # ── Sheet 8: Patient Master ───────────────────────────────────────────────
    ws_master = wb.create_sheet("Patient Master")
    write_df(ws_master, master)

    wb.save(output_path)

    # ── Staff-only workbook: the hand-to-staff CALL SHEET (P1-08 + call dropdowns).
    # Separate file, no ledgers/master. Follow-up calls + procedure call-backs +
    # vacation notices, with RESPONSE / CALLED BY dropdowns staff fill in.
    if staff_output_path:
        build_staff_call_workbook(
            df_reminder=df_reminder,
            df_action=df_action,
            df_vac_out=df_vac_out,
            visits=visits,
            today=today,
            out_path=staff_output_path,
            consult_date=consult_date,
            day_revenue_df=day_revenue_df,
            full_fu_ledger=fu_ledger,
        )

# ── Main run function ─────────────────────────────────────────────────────────
def run_daily(
    consultation_csv: str,
    followup_csv: str,
    today: date | None = None,
    callsheet_path: str | None = None
) -> tuple[str, str]:
    """
    Full daily processing pipeline.
    Returns (full_audit_path, staff_action_path).

    `callsheet_path` (optional): a FILLED Staff Call Sheet (.xlsx). If given, its
    RESPONSE / CALLED BY entries are captured into call_log.csv and the ledger's
    call-summary columns are refreshed. Capture-only — does not change the sheet
    that is generated this run.
    """
    INGEST_NOTICES.clear()   # fresh notices for this run

    # Safety net: snapshot existing ledgers before we touch them. Pending
    # follow-ups live only in data/*.csv, so this protects them across runs
    # and version updates.
    backup_data("daily")

    # Load persistent state
    master   = load_master()
    visits   = load_visits()
    fu_ledger= load_followups()

    # Parse both inputs up front so the reference date can come from the FILES.
    consult, consult_date_str = parse_consultation_report(consultation_csv)
    # Anchor the follow-up slice to the consultation report's CONTENT date so the
    # run always targets tomorrow's calls (anchor + 1), even from Docterz's default
    # one-month export. Falls back to the legacy earliest-day clamp if no anchor.
    fu_log, _ = parse_followup_log(followup_csv, anchor_date=parse_date(consult_date_str))

    # ── Reference date ("today") anchored to the FILES, not the PC clock ──────
    # The follow-up date in each log row is the call day (best evidence); the
    # consultation filename carries the visit day. Take the LATEST of these so the
    # run is correct whether it's deployed on the morning of the follow-up day or
    # the evening before — and so it does not depend on the clinic PC's clock.
    # An explicitly passed `today` (e.g. for testing) always wins.
    if today is None:
        cands = []
        cd = parse_date(consult_date_str)
        if cd:
            cands.append(cd)
        for s in list(fu_log.get("Due_Date", [])):
            dd = parse_date(s)
            if dd:
                cands.append(dd)
        today = max(cands) if cands else date.today()
    processed_on = date_to_str(today)

    # Step 1+2+3: Consultation report
    master, new_patients = update_master_from_consultation(master, consult, processed_on)
    visits, new_visits = append_visits(visits, consult, processed_on)

    # Step 4+5+6: Follow-up log
    fu_ledger, new_followups = append_followups(fu_ledger, fu_log, master, processed_on)

    # Step 7: Recalculate all statuses
    fu_ledger = recalculate_statuses(fu_ledger, visits, today, processed_on)

    # ── Phase 2: capture staff call outcomes (read-back) ─────────────────────
    # If a filled call sheet was uploaded, fold its RESPONSE / CALLED BY into
    # call_log.csv (idempotent), then refresh the ledger's call-summary columns
    # from the full log. Capture-only: this records outcomes but does NOT change
    # the sheet generated this run (that overlay is the next build step). Wrapped
    # so a call-sheet hiccup can never break the core follow-up run.
    if callsheet_path:
        try:
            _ci = ingest_call_sheet(callsheet_path, fu_ledger=fu_ledger,
                                    visits=visits, source_name=Path(callsheet_path).name)
            if _ci.get("error"):
                INGEST_NOTICES.append("Call sheet not captured: " + _ci["error"])
            else:
                INGEST_NOTICES.append(
                    f"Call outcomes captured for {_ci['date']}: {_ci['new']} new"
                    + (f", {_ci['duplicate']} already recorded" if _ci['duplicate'] else "")
                    + f" (from {_ci['read']} filled rows).")
        except Exception as _e:
            INGEST_NOTICES.append(f"Call sheet ingest skipped (non-fatal): {_e}")
    try:
        fu_ledger = recompute_call_summary(fu_ledger, load_call_log())
    except Exception as _e:
        print(f"  [call-summary] skipped (non-fatal): {_e}")

    # ── Auto-pull the Call_Feed from Google (if a feed link is configured) so the
    # Call Audit + Called-In pages are current after the normal run — no download,
    # no upload. Non-fatal: a network hiccup never breaks the follow-up run.
    if feed_url_configured():
        try:
            _ff = fetch_feed_auto()
            if _ff.get("ok"):
                INGEST_NOTICES.append(
                    f"Call feed refreshed from Google: {_ff.get('added', 0)} new call(s).")
            else:
                INGEST_NOTICES.append("Call feed not refreshed: " + _ff.get("error", "unknown"))
        except Exception as _e:
            INGEST_NOTICES.append(f"Call feed refresh skipped (non-fatal): {_e}")

    # ── Reinstated callbacks: open new "logged NOT PICK but connected" findings
    # and auto-clear any since genuinely reached/returned. Non-fatal. Surfaces in
    # Section 4 of the call sheet, named to the original caller.
    try:
        _rs = sync_reinstatements(load_call_log(), fu_ledger, visits,
                                  load_outbound_log(), today)
        if _rs.get("new") or _rs.get("cleared"):
            INGEST_NOTICES.append(
                f"Reinstated callbacks: {_rs['new']} new, {_rs['cleared']} cleared, "
                f"{_rs['open']} open.")
    except Exception as _e:
        INGEST_NOTICES.append(f"Reinstatement sync skipped (non-fatal): {_e}")

    # ── Confirmation worklist: open pending confirmations from confirmable closes
    # (Tt COMPLETE → Shavez, MED HERE → Pharmacy, NO → Shavez digest). Routed to
    # the assistants' /confirm worklist; never auto-convicts. Non-fatal.
    try:
        _cs = sync_confirmations(load_call_log(), fu_ledger, today)
        if _cs.get("new"):
            INGEST_NOTICES.append(
                f"Confirmations: {_cs['new']} new to confirm, {_cs['open']} open.")
    except Exception as _e:
        INGEST_NOTICES.append(f"Confirmation sync skipped (non-fatal): {_e}")

    # ── Revenue (Phase 1 auto-capture): pull Consultation / X-ray / Procedure
    # revenue from the SAME consultation report, fold it into the persistent
    # per-patient revenue ledger (de-duplicated), and prepare the day frame for
    # the "Day Revenue" tab. Done in a try/except so a revenue hiccup can never
    # break the follow-up run, which is the system's core job.
    day_rev_df = None
    consult_date = parse_date(consult_date_str)
    try:
        import revenue
        day_rev_df, _rd = revenue.parse_day_revenue(consultation_csv)
        if consult_date is None:
            consult_date = parse_date(_rd)

        # Apply manual ₹0/free-procedure markers for this report's date:
        #   • flip Had_Procedure = "Y" on the matching visit-ledger rows so they
        #     appear in the staff PROCEDURE CALL-BACKS section, and
        #   • mark the matching day-revenue rows so the Day Revenue tab lists them.
        man = revenue.manual_proc_for_date(_rd)
        man_uids = set(man["Patient_UID"].astype(str).str.strip()) if len(man) else set()
        if man_uids:
            vd_str = _rd
            mask = (visits["Visit_Date"].astype(str) == vd_str) & \
                   (visits["Patient_UID"].astype(str).str.strip().isin(man_uids))
            visits.loc[mask, "Had_Procedure"] = "Y"
            if day_rev_df is not None and len(day_rev_df):
                day_rev_df["Manual_Procedure"] = day_rev_df["Patient_UID"].astype(str).str.strip().isin(man_uids)
        if day_rev_df is not None and "Manual_Procedure" not in day_rev_df.columns:
            day_rev_df["Manual_Procedure"] = False

        added_rev = revenue.update_revenue_ledger_from_day(
            day_rev_df, _rd, Path(consultation_csv).name, processed_on)
        print(f"  [revenue] {added_rev} new revenue rows captured from consultation report"
              f"{'; ' + str(len(man_uids)) + ' manual ₹0 procedure(s) applied' if man_uids else ''}")
        conc_added = revenue.append_concession_log(
            day_rev_df, Path(consultation_csv).name, processed_on)
        if conc_added:
            print(f"  [revenue] {conc_added} new concession case(s) logged to standing list")
    except Exception as _e:
        print(f"  [revenue] skipped (non-fatal): {_e}")

    # Save persistent state
    save_master(master)
    save_visits(visits)
    save_followups(fu_ledger)

    # Step 8: Generate output Excel(s)
    out_filename   = f"Followup_Audit_{processed_on}.xlsx"
    staff_filename = f"Staff_Action_Today_{processed_on}.xlsx"
    out_path   = str(OUTPUTS_DIR / out_filename)
    staff_path = str(OUTPUTS_DIR / staff_filename)
    build_output_excel(
        fu_ledger=fu_ledger,
        visits=visits,
        master=master,
        new_patients=new_patients,
        new_visits=new_visits,
        new_followups=new_followups,
        today=today,
        output_path=out_path,
        staff_output_path=staff_path,
        consult_date=consult_date,
        day_revenue_df=day_rev_df,
    )

    return out_path, staff_path


def parse_patient_list(filepath: str) -> pd.DataFrame:
    """
    Parse the Docterz full patient list export.
    This file has different column names from the daily consultation report:
      - 'Name' (not 'Patient Name')
      - 'Mobile' (already clean string, not float)
      - 'Registered On' (not 'Consultation Date'), format DD/MM/YYYY
      - No clinic header junk row
    Deduplicates rows with identical Patient UID (exact row duplicates in export).
    """
    df = pd.read_csv(filepath, dtype=str)

    # Deduplicate exact Patient UID duplicates — keep first occurrence
    before = len(df)
    df = df.drop_duplicates(subset=["Patient UID"], keep="first").reset_index(drop=True)
    dupes_removed = before - len(df)

    out = pd.DataFrame()
    out["Patient_UID"]        = df["Patient UID"].str.strip()
    out["Clinic_Specific_Id"] = df["Clinic Specific Id"].apply(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ("", "nan") else ""
    )
    out["Patient_Name"]       = df["Name"].str.strip()
    out["Mobile_Raw"]         = df["Mobile"].astype(str)
    out["Mobile_Clean"]       = out["Mobile_Raw"].apply(clean_mobile)
    out["Registered_On"]      = df["Registered On"].apply(
        lambda x: date_to_str(parse_date(x))
    )
    out["_dupes_removed"]     = dupes_removed  # carry through for reporting

    # Drop rows with no Patient_UID
    out = out[out["Patient_UID"].notna() & (out["Patient_UID"] != "")].reset_index(drop=True)
    return out


def run_initial_master(master_csv: str) -> str:
    """
    One-time setup: ingest the full Docterz patient list CSV
    (DOCTERZ_PATIENT_LIST_UPTO_31_MAY_2026.csv or similar).

    This file uses the patient-list export format (Name, Mobile as string,
    Registered On) — not the daily consultation report format.

    Merges with any patients already in the master from daily runs,
    so it is safe to run even after daily processing has begun.
    Returns a summary string.
    """
    patient_list = parse_patient_list(master_csv)
    dupes_removed = int(patient_list["_dupes_removed"].iloc[0]) if len(patient_list) > 0 else 0
    patient_list = patient_list.drop(columns=["_dupes_removed"])

    backup_data("init_master")   # snapshot existing ledgers before merging the roster
    processed_on = date_to_str(date.today())
    master = load_master()
    added = 0
    updated = 0

    for _, row in patient_list.iterrows():
        uid       = row["Patient_UID"]
        csid      = row["Clinic_Specific_Id"]
        name      = row["Patient_Name"]
        mob_raw   = row["Mobile_Raw"]
        mob_clean = row["Mobile_Clean"]
        reg_date  = row["Registered_On"]

        existing = master[master["Patient_UID"] == uid]

        if len(existing) == 0:
            new_row = {
                "Patient_UID":            uid,
                "Clinic_Specific_Id":     csid,
                "Patient_Name":           name,
                "Mobile_Raw":             mob_raw,
                "Mobile_Clean":           mob_clean,
                "Mobile_Status":          mobile_status(mob_clean),
                "First_Seen_Date":        reg_date,
                "Last_Seen_Date":         reg_date,
                "Mobile_Duplicate_Count": "0",
                "Identity_Status":        "",
                "Added_From":             "Patient List Import",
                "Last_Updated":           processed_on
            }
            master = pd.concat([master, pd.DataFrame([new_row])], ignore_index=True)
            added += 1
        else:
            # Patient already in master (from daily runs) — fill in any gaps
            idx = existing.index[0]
            if not master.at[idx, "Mobile_Clean"] and mob_clean:
                master.at[idx, "Mobile_Raw"]    = mob_raw
                master.at[idx, "Mobile_Clean"]  = mob_clean
                master.at[idx, "Mobile_Status"] = mobile_status(mob_clean)
            if not master.at[idx, "Clinic_Specific_Id"] and csid:
                master.at[idx, "Clinic_Specific_Id"] = csid
            # Use registration date as First_Seen_Date if earlier
            existing_first = parse_date(master.at[idx, "First_Seen_Date"])
            this_reg       = parse_date(reg_date)
            if this_reg and (not existing_first or this_reg < existing_first):
                master.at[idx, "First_Seen_Date"] = reg_date
            master.at[idx, "Last_Updated"] = processed_on
            updated += 1

    master = recalculate_duplicate_flags(master)
    save_master(master)

    # Build summary
    shared   = len(master[master["Identity_Status"] == "Shared Mobile"])
    no_mob   = len(master[master["Identity_Status"] == "No/Invalid Mobile"])
    unique   = len(master[master["Identity_Status"] == "Unique Mobile"])

    summary = (
        f"Patient Master initialised from patient list export.\n"
        f"  Total patients: {len(master)}\n"
        f"  New patients added: {added}\n"
        f"  Existing patients updated: {updated}\n"
        f"  Duplicate UIDs removed from source: {dupes_removed}\n"
        f"  Unique mobile: {unique}\n"
        f"  Shared mobile (family): {shared}\n"
        f"  No/Invalid mobile (incl. 1111111111): {no_mob}"
    )
    return summary
