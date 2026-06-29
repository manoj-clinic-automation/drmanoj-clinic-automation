"""End-to-end test of the 11-Jun fixes. Run inside an empty sandbox folder."""
import sys, shutil
from pathlib import Path
from datetime import date
import pandas as pd

SANDBOX = Path("/home/claude/fufix/sandbox")
if SANDBOX.exists():
    shutil.rmtree(SANDBOX)
SANDBOX.mkdir()
shutil.copy("/home/claude/fufix/processor.py", SANDBOX / "processor.py")
sys.path.insert(0, str(SANDBOX))
import processor as P  # noqa: E402

FAIL = []
def check(name, cond):
    print(("PASS  " if cond else "FAIL  ") + name)
    if not cond:
        FAIL.append(name)

# ── seed master via initial patient list ─────────────────────────────────────
plist = SANDBOX / "patient_list.csv"
pd.DataFrame({
    "Name":               ["Ram Kumar", "Sita Devi", "Mohan Lal", "Geeta Sharma", "Raju Sharma"],
    "Mobile":             ["9000000001", "9000000002", "9000000003", "9000000044", "9000000044"],
    "Patient UID":        ["UID001", "UID002", "UID003", "UID004", "UID005"],
    "Clinic Specific Id": ["101", "102", "103", "104", "105"],
    "Registered On":      ["01/01/2026"] * 5,
}).to_csv(plist, index=False)
P.run_initial_master(str(plist))

# ── diagnosis file: UID-keyed for Ram (RA), legacy mobile-only for Mohan,
#    and a shared-mobile pair (Geeta/Raju) that must NOT enrich by mobile ─────
pd.DataFrame({
    "Patient_UID":            ["UID001", "",           "",           ""],
    "Mobile_Clean":           ["9000000001", "9000000003", "9000000044", "9000000044"],
    "Standardized_Diagnosis": ["RA",        "OA Knee",    "Fracture",   "AS"],
    "Diagnosis_Priority":     ["A",         "B",          "A",          "A"],
    "Is_VIP":                 ["True",      "False",      "False",      "False"],
}).to_csv(P.DIAGNOSIS_FILE, index=False)

# ── vacation calendar: 20–22 Aug 2026, Morning ───────────────────────────────
pd.DataFrame([{"Start_Date": "20-08-2026", "End_Date": "22-08-2026",
               "Slot": "Morning", "Note": "Conference"}]).to_csv(P.VACATION_FILE, index=False)

# ── Day 1 (01-06-2026): follow-up log gives Ram TWO obligations (due 02-06 and
#    05-06 via a 4-day-span file... span guard allows <=3 days, so use 02 & 04),
#    Sita one due 65+ days in the past can't come from a log (logs are future),
#    so instead her due is 02-06 and we test expiry by running 'today' far ahead.
cons1 = SANDBOX / "consultation_report_2026-06-01.csv"
pd.DataFrame({
    "Sr No":              ["Dr. Manoj Agarwal Clinic", "1", "2", "3", "4", "5"],
    "Patient UID":        ["", "UID001", "UID002", "UID003", "UID004", "UID005"],
    "Clinic Specific Id": ["", "101", "102", "103", "104", "105"],
    "Patient Name":       ["", "Ram Kumar", "Sita Devi", "Mohan Lal", "Geeta Sharma", "Raju Sharma"],
    "Mobile":             ["", "9000000001.0", "9000000002.0", "9000000003.0", "9000000044.0", "9000000044.0"],
    "Consultation Date":  ["", "01-06-2026 10:00 AM"] + ["01-06-2026 11:00 AM"] * 4,
}).to_csv(cons1, index=False)

fu1 = SANDBOX / "followup_log_day1.csv"
pd.DataFrame({
    "Appointment ID": ["A1", "A2", "A3", "A4"],
    "Mobile No":      ["9000000001", "9000000001", "9000000002", "9000000003"],
    "Patient Name":   ["Ram Kumar", "Ram Kumar", "Sita Devi", "Mohan Lal"],
    "Followup Date":  ["02-06-2026", "04-06-2026", "02-06-2026", "02-06-2026"],
}).to_csv(fu1, index=False)

audit1, staff1 = P.run_daily(str(cons1), str(fu1), today=date(2026, 6, 1))

# Second same-day upload: Geeta/Raju follow-ups due 20-08 (inside the vacation
# window; single-day span so the guard passes). Shared family mobile, distinct
# names -> each resolves to own UID with Medium confidence.
cons1b = SANDBOX / "consultation_report_empty.csv"
pd.DataFrame({
    "Sr No": ["Dr. Manoj Agarwal Clinic"], "Patient UID": [""],
    "Clinic Specific Id": [""], "Patient Name": [""],
    "Mobile": [""], "Consultation Date": [""],
}).to_csv(cons1b, index=False)
fu1b = SANDBOX / "followup_log_day1b.csv"
pd.DataFrame({
    "Appointment ID": ["A5", "A6"],
    "Mobile No":      ["9000000044", "9000000044"],
    "Patient Name":   ["Geeta Sharma", "Raju Sharma"],
    "Followup Date":  ["20-08-2026", "20-08-2026"],
}).to_csv(fu1b, index=False)
P.run_daily(str(cons1b), str(fu1b), today=date(2026, 6, 1))

# Strict-after rule: the 01-06 prescribing visits must NOT have closed the
# 02-06 follow-ups created the same day.
_fu = P.load_followups()
check("Prescribing-day visit does not close its own follow-up",
      not any(_fu["Followup_Status"].str.startswith("Returned")))

# ── Day 2 (06-06-2026): Ram returns ONCE. P0-01: only ONE of his two pending
#    follow-ups may close. ─────────────────────────────────────────────────────
cons2 = SANDBOX / "consultation_report_2026-06-06.csv"
pd.DataFrame({
    "Sr No":              ["Dr. Manoj Agarwal Clinic", "1"],
    "Patient UID":        ["", "UID001"],
    "Clinic Specific Id": ["", "101"],
    "Patient Name":       ["", "Ram Kumar"],
    "Mobile":             ["", "9000000001.0"],
    "Consultation Date":  ["", "06-06-2026 10:30 AM"],
}).to_csv(cons2, index=False)
fu2 = SANDBOX / "followup_log_day2.csv"
pd.DataFrame({
    "Appointment ID": ["A7"],
    "Mobile No":      ["9000000003"],
    "Patient Name":   ["Mohan Lal"],
    "Followup Date":  ["07-06-2026"],
}).to_csv(fu2, index=False)

audit2, staff2 = P.run_daily(str(cons2), str(fu2), today=date(2026, 6, 6))

fu = P.load_followups()
ram = fu[fu["FU_Mobile_Clean"] == "9000000001"].sort_values("Due_Date")
ram_statuses = ram["Followup_Status"].tolist()
check("P0-01: Ram's earlier follow-up (02-06) closed by the single 06-06 visit",
      ram_statuses[0] == "Returned Late")
check("P0-01: Ram's later follow-up (04-06) NOT closed by the same visit",
      not ram_statuses[1].startswith("Returned"))
check("P0-01: Matched_Visit_ID stored on the closed row",
      bool(str(ram.iloc[0]["Matched_Visit_ID"]).strip()))
check("P0-01: open row carries no Matched_Visit_ID",
      not str(ram.iloc[1]["Matched_Visit_ID"]).strip())

# ── Day 3 (15-08-2026, ~74 days after Sita's 02-06 due): expiry check ────────
cons3 = SANDBOX / "consultation_report_2026-08-15.csv"
pd.DataFrame({
    "Sr No": ["Dr. Manoj Agarwal Clinic"], "Patient UID": [""],
    "Clinic Specific Id": [""], "Patient Name": [""],
    "Mobile": [""], "Consultation Date": [""],
}).to_csv(cons3, index=False)
fu3 = SANDBOX / "followup_log_day3.csv"
pd.DataFrame({"Appointment ID": [], "Mobile No": [], "Patient Name": [],
              "Followup Date": []}).to_csv(fu3, index=False)
audit3, staff3 = P.run_daily(str(cons3), str(fu3), today=date(2026, 8, 15))

fu = P.load_followups()
sita = fu[fu["FU_Mobile_Clean"] == "9000000002"].iloc[0]
check("P0-02: 74-days-overdue follow-up becomes Expired Unresolved",
      sita["Followup_Status"] == "Expired Unresolved")
ram_open = fu[(fu["FU_Mobile_Clean"] == "9000000001")
              & (~fu["Followup_Status"].str.startswith("Returned"))].iloc[0]
check("P0-02: Ram's unmatched 04-06 row also expired by 15-08",
      ram_open["Followup_Status"] == "Expired Unresolved")

# Staff Action Today must not contain Expired rows
from openpyxl import load_workbook
wb = load_workbook(audit3)
ws = wb["Staff Action Today"]
statuses_in_action = {ws.cell(r, [c.value for c in ws[1]].index("Status") + 1).value
                      for r in range(2, ws.max_row + 1)} if ws.max_row > 1 else set()
check("P0-02: 'Expired Unresolved' absent from Staff Action Today",
      "Expired Unresolved" not in statuses_in_action)
check("Dashboard exists with Expired line",
      any("Expired Unresolved" in str(ws2.cell(r, 1).value)
          for ws2 in [wb["Dashboard"]] for r in range(1, ws2.max_row + 1)))

# Vacation notice: Geeta & Raju due 20-08 (inside 20–22 Aug window)
ws_vac = wb["Vacation Notice List"]
vac_names = {ws_vac.cell(r, 1).value for r in range(2, ws_vac.max_row + 1)}
check("Vacation: Geeta in Vacation Notice List", "Geeta Sharma" in vac_names)
check("Vacation: Raju in Vacation Notice List", "Raju Sharma" in vac_names)

# Action sheet vacation + msg category columns
hdr = [c.value for c in ws[1]]
check("Action sheet has 'Doctor Unavailable' column", "Doctor Unavailable" in hdr)
check("Action sheet has 'Msg Category' column", "Msg Category" in hdr)

# P0-03 diagnosis checks on the action sheet (Geeta/Raju rows are open Not Due
# until 20-08 — by 15-08 they're Not Due, not in ACTION sheet... so check via
# the enrichment by re-running lookup logic directly)
diag_uid, diag_mob = P.load_diagnosis_lookup()
check("P0-03: UID lookup carries Ram's RA", diag_uid.get("UID001", {}).get("diag") == "RA")
check("P0-03: legacy unique mobile (Mohan) still available via mobile fallback",
      diag_mob.get("9000000003", {}).get("diag") == "OA Knee")
check("P0-03: shared mobile 9000000044 dropped from mobile fallback entirely",
      "9000000044" not in diag_mob)
check("Taxonomy provision: RA maps to inflammatory_arthritis_bilingual",
      P.message_category_for("RA") == "inflammatory_arthritis_bilingual")
check("Taxonomy provision: unmapped diagnosis maps to empty",
      P.message_category_for("OA Knee") == "")

# Staff workbook: exactly the two staff sheets, no ledgers/master
wbs = load_workbook(staff3)
check("Staff workbook has only Staff Action Today + Vacation Notice List",
      set(wbs.sheetnames) == {"Staff Action Today", "Vacation Notice List"})

# ── Day 4 (10-09-2026): Sita returns VERY late → expired row reconciles ──────
cons4 = SANDBOX / "consultation_report_2026-09-10.csv"
pd.DataFrame({
    "Sr No":              ["Dr. Manoj Agarwal Clinic", "1"],
    "Patient UID":        ["", "UID002"],
    "Clinic Specific Id": ["", "102"],
    "Patient Name":       ["", "Sita Devi"],
    "Mobile":             ["", "9000000002.0"],
    "Consultation Date":  ["", "10-09-2026 10:00 AM"],
}).to_csv(cons4, index=False)
audit4, staff4 = P.run_daily(str(cons4), str(fu3), today=date(2026, 9, 10))
fu = P.load_followups()
sita = fu[fu["FU_Mobile_Clean"] == "9000000002"].iloc[0]
check("P0-02: expired row reconciles to Returned Late on a very late return",
      sita["Followup_Status"] == "Returned Late")

print()
print("ALL PASS" if not FAIL else f"{len(FAIL)} FAILURE(S): {FAIL}")
sys.exit(1 if FAIL else 0)
