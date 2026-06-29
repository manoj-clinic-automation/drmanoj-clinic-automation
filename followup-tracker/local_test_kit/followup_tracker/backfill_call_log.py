"""
Backfill the call log from a folder of saved, FILLED call sheets.
Dr. Manoj Agarwal Clinic — Follow-Up Tracker, Phase 2.

Use this once, when you have several days of filled Staff Call Sheets saved up
(e.g. the habit-phase weeks) and want them all captured at once. It ingests every
.xlsx in the given folder, in the order of the date printed on each sheet, into
data/call_log.csv. It is idempotent — running it twice changes nothing — so it is
always safe to re-run.

USAGE
    python backfill_call_log.py "C:\\path\\to\\filled_call_sheets"

If no folder is given it uses ./filled_call_sheets next to this script.
"""
import sys
from pathlib import Path
from datetime import date

import processor as P


def _sheet_date(path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb["Call Sheet"] if "Call Sheet" in wb.sheetnames else wb[wb.sheetnames[0]]
        return P._parse_callsheet_date(ws, date(1900, 1, 1))
    except Exception:
        return date(1900, 1, 1)


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "filled_call_sheets"
    if not folder.exists():
        print(f"Folder not found: {folder}")
        return
    files = sorted([f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")],
                   key=_sheet_date)
    if not files:
        print(f"No .xlsx call sheets found in {folder}")
        return

    fu = P.load_followups()
    visits = P.load_visits()
    print(f"Backfilling {len(files)} sheet(s) from {folder}\n")
    total_new = total_dup = 0
    for f in files:
        res = P.ingest_call_sheet(str(f), fu_ledger=fu, visits=visits, source_name=f.name)
        if res.get("error"):
            print(f"  SKIP {f.name}: {res['error']}")
            continue
        total_new += res["new"]; total_dup += res["duplicate"]
        print(f"  {res['date']}  {f.name}: {res['new']} new, {res['duplicate']} already recorded")

    # refresh the ledger summary columns from the full log
    try:
        fu = P.recompute_call_summary(P.load_followups(), P.load_call_log())
        P.save_followups(fu)
    except Exception as e:
        print(f"  [summary] skipped: {e}")

    print(f"\nDone. {total_new} new call(s) captured, {total_dup} already present.")
    print("call_log.csv updated; follow-up ledger summary refreshed.")


if __name__ == "__main__":
    main()
