"""
Phase 2 — Call Read-Back & Overlay test suite.
Dr. Manoj Agarwal Clinic, Follow-Up Tracker.

Self-contained: it builds its own call sheet, so it needs no sample file. Run it
from the tracker folder (next to processor.py) after deploying, to confirm the
read-back loop is sound on this machine:

    python test_call_readback.py

It uses a throwaway temp folder and never touches your real data\\ ledgers.
"""
import tempfile, shutil
from pathlib import Path
from datetime import date
import pandas as pd
import openpyxl
import processor as P

PASS = 0

def ok(msg):
    global PASS; PASS += 1
    print("  PASS —", msg)

# ── isolate all state in a temp dir ───────────────────────────────────────────
tmp = Path(tempfile.mkdtemp())
P.CALL_LOG_FILE = tmp / "call_log.csv"
P.FOLLOWUP_FILE = tmp / "followup_ledger.csv"
P.VISITS_FILE   = tmp / "visit_ledger.csv"

def make_row(fid, name, status="Due Today", od="0", due="22-06-2026",
             res="", lastcall="", attempts="", pri="C", uid=None):
    r = {c: "" for c in P.FU_COLS}
    r.update({"Followup_ID": fid, "FU_Name_Raw": name, "FU_Mobile_Clean": "9000000000",
              "Followup_Status": status, "Days_Overdue": od, "Due_Date": due,
              "Patient_UID_Resolved": uid or ("U" + fid), "Identity_Confidence": "High",
              "Call_Resolution": res, "Last_Call_Date": lastcall, "Call_Attempts": attempts})
    r["Diagnosis"] = "Knee Osteoarthritis"; r["Priority"] = pri; r["VIP_Flag"] = ""
    return r

def build_sheet(df_action, today, path, visits=None):
    P.build_staff_call_workbook(
        df_reminder=df_action.iloc[0:0].copy(), df_action=df_action,
        df_vac_out=pd.DataFrame(columns=["A"]),
        visits=visits if visits is not None else pd.DataFrame(columns=P.VISIT_COLS),
        today=today, out_path=str(path), consult_date=today, day_revenue_df=None)

def fill_sheet(path, date_label, fills):
    """Set the date banner and fill RESPONSE/CALLED BY by KEY."""
    wb = openpyxl.load_workbook(path); ws = wb["Call Sheet"]
    ws.cell(2, 1).value = f"Date: {date_label}    Pick RESPONSE and CALLED BY"
    keyc = respc = byc = None
    for r in range(1, 12):
        labs = {str(ws.cell(r, c).value).strip().upper(): c
                for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "KEY" in labs and "RESPONSE" in labs:
            keyc, respc, byc = labs["KEY"], labs["RESPONSE"], labs["CALLED BY"]; break
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, keyc).value
        if k and str(k).strip() in fills:
            resp, by = fills[str(k).strip()]
            ws.cell(r, respc).value = resp; ws.cell(r, byc).value = by
    wb.save(path)

def sheet_rows(path):
    """Return {KEY: (section, status_text)} for every keyed data row."""
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb["Call Sheet"]
    sec = None; seen = {}
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "")
        if a.strip().startswith(("1.", "2.", "3.")):
            sec = a.strip()[:2]
        k = ws.cell(r, 13).value
        if k and str(k).strip() != "KEY":
            seen[str(k).strip()] = (sec, str(ws.cell(r, 8).value or ""))
    return seen

today = date(2026, 6, 24)

print("TEST 1 — capture + resolution classes")
base = pd.DataFrame([make_row("F000001", "A"), make_row("F000002", "B"),
                     make_row("F000003", "C"), make_row("F000004", "D")])
P.save_followups(base); P.save_visits(pd.DataFrame(columns=P.VISIT_COLS))
s22 = tmp / "s22.xlsx"; build_sheet(base, date(2026, 6, 22), s22)
fill_sheet(s22, "22-Jun-2026", {"F000001": ("YES", "Shavez"),
           "F000002": ("NOT PICK", "Alisha"), "F000003": ("LATER", "Shivani"),
           "F000004": ("NO", "Ranjeet")})
r = P.ingest_call_sheet(str(s22), fu_ledger=base, visits=pd.DataFrame(columns=P.VISIT_COLS))
assert r["new"] == 4 and r["date"] == "2026-06-22", r
cls = dict(zip(P.load_call_log()["Obligation_Key"], P.load_call_log()["Resolution_Class"]))
assert cls == {"F000001": "RESOLVE", "F000002": "RETRY",
               "F000003": "RESCHEDULE", "F000004": "DECLINE"}, cls
ok("4 outcomes captured with correct resolution classes")

print("TEST 2 — idempotent re-upload")
r2 = P.ingest_call_sheet(str(s22), fu_ledger=base, visits=pd.DataFrame(columns=P.VISIT_COLS))
assert r2["new"] == 0 and r2["duplicate"] == 4 and len(P.load_call_log()) == 4, r2
ok("re-upload adds nothing")

print("TEST 3 — multi-attempt numbering + escalation")
for d, by in (("23-Jun-2026", "Shavez"), ("24-Jun-2026", "Ranjeet")):
    sx = tmp / f"s{d[:2]}.xlsx"; build_sheet(base, date(2026, 6, int(d[:2])), sx)
    fill_sheet(sx, d, {"F000002": ("NOT PICK", by)})
    P.ingest_call_sheet(str(sx), fu_ledger=base, visits=pd.DataFrame(columns=P.VISIT_COLS))
log = P.load_call_log()
att = list(log[log["Obligation_Key"] == "F000002"].sort_values("Attempt_No")["Attempt_No"])
assert att == ["1", "2", "3"], att
fu2 = P.recompute_call_summary(base, log)
row = fu2[fu2["Followup_ID"] == "F000002"].iloc[0]
assert row["Call_Attempts"] == "3" and row["Call_Resolution"] == "ESCALATE", dict(row)
ok("attempts 1/2/3 numbered; escalation fires at 3")

print("TEST 4 — sheet overlay hides / holds / annotates / escalates")
ov = pd.DataFrame([
    make_row("F000001", "ResolveYes", "Actionable Missed Follow-Up", "4", "20-06-2026", res="RESOLVE",   lastcall="2026-06-22", attempts="1"),
    make_row("F000002", "DeclineNo",  "Actionable Missed Follow-Up", "4", "20-06-2026", res="DECLINE",   lastcall="2026-06-22", attempts="1"),
    make_row("F000003", "Retry",      "Grace Period",                "2", "22-06-2026", res="RETRY",     lastcall="2026-06-23", attempts="2"),
    make_row("F000004", "Hold",       "Grace Period",                "2", "22-06-2026", res="RESCHEDULE",lastcall="2026-06-23", attempts="1"),
    make_row("F000005", "ReschedDue", "Actionable Missed Follow-Up", "5", "19-06-2026", res="RESCHEDULE",lastcall="2026-06-21", attempts="1"),
    make_row("F000006", "Escalate",   "Probable Dropout",            "12","12-06-2026", res="ESCALATE",  lastcall="2026-06-23", attempts="3"),
    make_row("F000007", "Fresh",      "Due Today",                   "0", "24-06-2026"),
])
sov = tmp / "ov.xlsx"; build_sheet(ov, today, sov)
seen = sheet_rows(sov)
assert "F000001" not in seen and "F000002" not in seen, "resolve/decline must be hidden"
assert "F000004" not in seen, "reschedule in hold must be hidden"
assert seen["F000003"][0] == "1." and "call again" in seen["F000003"][1]
assert "call later" in seen["F000005"][1] or "call now" in seen["F000005"][1]
assert seen["F000006"][0] == "3.", "escalate must be in watch section"
assert seen["F000007"][0] == "1."
ok("overlay: resolved/declined/held hidden; retry annotated; escalate in watch")

print("TEST 5 — monitor stats")
st = P.call_monitor_stats("2026-06-22")
assert st["total"] == 4 and st["reached"] == 3 and st["not_reached"] == 1 and st["contact_rate"] == 75, st
assert st["all_dates"][0] == "2026-06-24"
ok("monitor: 4 calls, 75% contact rate, day list correct")

print("TEST 6 — unfulfilled 'YES/coming' re-surfaces after grace; genuine resolve stays hidden")
prom = pd.DataFrame([
    make_row("F000010", "PromiseDue",  "Grace Period", "3", "21-06-2026", res="RESOLVE", lastcall="2026-06-20", attempts="1"),
    make_row("F000011", "PromiseHold", "Grace Period", "1", "23-06-2026", res="RESOLVE", lastcall="2026-06-23", attempts="1"),
    make_row("F000012", "TreatDone",   "Grace Period", "3", "21-06-2026", res="RESOLVE", lastcall="2026-06-20", attempts="1"),
])
prom.loc[prom["Followup_ID"] == "F000010", "Last_Response"] = "YES"          # promise, grace elapsed (20+3=23<=24)
prom.loc[prom["Followup_ID"] == "F000011", "Last_Response"] = "YES"          # promise, within grace (23+3=26>24)
prom.loc[prom["Followup_ID"] == "F000012", "Last_Response"] = "Tt COMPLETE"  # genuine resolve
sp = tmp / "prom.xlsx"; build_sheet(prom, today, sp)
seen = sheet_rows(sp)
assert "F000010" in seen and "re-confirm" in seen["F000010"][1], seen.get("F000010")
assert "F000011" not in seen, "promise still within grace must stay hidden"
assert "F000012" not in seen, "treatment-complete must stay hidden"
ok("YES promise re-surfaces only after grace; Tt COMPLETE stays resolved")

print("TEST 7 — reconciliation: responses cross-checked against actual returns")
recon = pd.DataFrame([
    make_row("F000020", "CalledReturned", "Returned On Time", "0", "20-06-2026", res="RESOLVE"),
    make_row("F000021", "PromiseNoReturn", "Grace Period",    "3", "21-06-2026", res="RESOLVE"),
    make_row("F000022", "SelfResolved",    "Actionable Missed Follow-Up", "4", "20-06-2026", res="RESOLVE"),
    make_row("F000023", "Declined",        "Actionable Missed Follow-Up", "4", "20-06-2026", res="DECLINE"),
    make_row("F000024", "Organic",         "Returned Early",  "0", "20-06-2026"),
    make_row("F000025", "Unreachable",     "Probable Dropout","12","12-06-2026", res="ESCALATE"),
])
recon.loc[recon["Followup_ID"] == "F000020", "Last_Response"] = "YES"
recon.loc[recon["Followup_ID"] == "F000021", "Last_Response"] = "YES"
recon.loc[recon["Followup_ID"] == "F000022", "Last_Response"] = "Tt COMPLETE"
recon.loc[recon["Followup_ID"] == "F000023", "Last_Response"] = "NO"
recon.loc[recon["Followup_ID"] == "F000024", "Last_Response"] = ""
recon.loc[recon["Followup_ID"] == "F000025", "Last_Response"] = "NOT PICK"
P.save_followups(recon)
rc = P.call_reconciliation_stats()
assert rc["called"] == 5 and rc["called_returned"] == 1 and rc["call_return_rate"] == 20, rc
assert rc["unfulfilled_promise"] == 1 and rc["self_resolved_no_return"] == 1, rc
assert rc["declined"] == 1 and rc["unreached"] == 1 and rc["organic_return"] == 1, rc
ok("reconciliation: 20% call→return rate; promise/self-resolved/declined/unreached/organic all correct")

print("\nALL %d PHASE-2 CHECKS PASSED" % PASS)
shutil.rmtree(tmp, ignore_errors=True)
