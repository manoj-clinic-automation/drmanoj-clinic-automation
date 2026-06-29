"""
build_insight.py
Dr. Manoj Agarwal Clinic — Diagnosis Insight Workbook

Reads data/patient_diagnosis.csv (built by seed_diagnosis.py) and produces a
formatted Excel report so the doctor can SEE the cleaned full patient list and
taxonomy-wise trends.

Sheets:
  1. Summary            — totals, priority mix, category mix (Excel formulas)
  2. Full Patient List  — cleaned, one row per patient
  3. Taxonomy Breakdown — patients per standardized diagnosis & category
  4. Age x Category     — cross-tab (cohort targeting)
  5. VIP & Concessions  — VIP list + CC/PD/BID concession rows
  6. Review (Unclassified) — strings the taxonomy could not place (for tuning)
"""

from pathlib import Path
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent / "data"
DIAG_FILE = DATA_DIR / "patient_diagnosis.csv"

FONT = "Arial"
NAVY  = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
PRI_FILL = {"A": "FFC7CE", "B": "FFEB9C", "C": "C6EFCE", "D": "E0E0E0"}
HDR = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BOLD = Font(name=FONT, bold=True, size=11)
REG = Font(name=FONT, size=10)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _hdr(ws, row, headers, start=1, fill=BLUE):
    for j, h in enumerate(headers, start):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR; c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN


def _title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def build(out_path):
    df = pd.read_csv(DIAG_FILE, dtype=str).fillna("")
    total = len(df)
    df["Age_num"] = pd.to_numeric(df["Age"], errors="coerce")
    n_age   = int((df["Age"].astype(str).str.strip() != "").sum())
    n_nodx  = int((df["Diagnosis_Category"] == "No Diagnosis Recorded").sum())
    n_uncl  = int((df["Diagnosis_Category"] == "Unclassified").sum())
    n_dx    = total - n_nodx - n_uncl     # patients with a classified diagnosis
    n_textdx = n_dx + n_uncl              # patients who have any diagnosis text
    cov = (n_dx / n_textdx * 100) if n_textdx else 0

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    _title(ws, "Dr. Manoj Agarwal Clinic — Patient Diagnosis Insight", 4)
    ws.cell(row=2, column=1, value=f"Generated {dt.date.today().isoformat()}   •   "
            f"Total patients: {total}   •   Age known: {n_age} ({n_age/total*100:.1f}%)   •   "
            f"With diagnosis: {n_dx}   •   No diagnosis yet: {n_nodx}   •   "
            f"Taxonomy coverage of diagnosed: {cov:.1f}%").font = \
        Font(name=FONT, italic=True, size=9, color="595959")

    r = 4
    ws.cell(row=r, column=1, value="PRIORITY MIX (A = most urgent; among patients with a diagnosis)").font = BOLD
    r += 1
    _hdr(ws, r, ["Priority", "Meaning", "Patients", "Share"]); r += 1
    pri_meta = [("A","Trauma / Post-op / Infection / Tumor"),
                ("B","Inflammatory / Surgical candidate (PIVD, RA, OA-bilateral, hip)"),
                ("C","Degenerative / Mechanical / Soft-tissue / chronic"),
                ("D","Unclassified — has text but needs taxonomy review")]
    pri_start = r
    for p, meaning in pri_meta:
        cnt = int((df["Diagnosis_Priority"] == p).sum())
        ws.cell(row=r, column=1, value=p).font = BOLD
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=PRI_FILL[p])
        ws.cell(row=r, column=2, value=meaning).font = REG
        ws.cell(row=r, column=3, value=cnt).font = REG
        ws.cell(row=r, column=4, value=f"=C{r}/$C${pri_start+5}").font = REG
        ws.cell(row=r, column=4).number_format = "0.0%"
        for col in range(1, 5): ws.cell(row=r, column=col).border = THIN
        r += 1
    # No-diagnosis row
    ws.cell(row=r, column=1, value="—").font = BOLD
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=r, column=2, value="No diagnosis recorded yet (in roster, awaiting clinical entry)").font = REG
    ws.cell(row=r, column=3, value=n_nodx).font = REG
    ws.cell(row=r, column=4, value=f"=C{r}/$C${pri_start+5}").font = REG
    ws.cell(row=r, column=4).number_format = "0.0%"
    for col in range(1, 5): ws.cell(row=r, column=col).border = THIN
    r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD
    ws.cell(row=r, column=3, value=f"=SUM(C{pri_start}:C{r-1})").font = BOLD
    ws.cell(row=r, column=4, value=f"=C{r}/C{r}").font = BOLD
    ws.cell(row=r, column=4).number_format = "0.0%"
    for col in range(1, 5): ws.cell(row=r, column=col).border = THIN

    r += 2
    ws.cell(row=r, column=1, value="CATEGORY MIX").font = BOLD; r += 1
    _hdr(ws, r, ["Diagnosis Category", "Patients", "Share"]); r += 1
    cat_start = r
    cat_counts = df["Diagnosis_Category"].value_counts()
    for cat, cnt in cat_counts.items():
        ws.cell(row=r, column=1, value=cat).font = REG
        ws.cell(row=r, column=2, value=int(cnt)).font = REG
        ws.cell(row=r, column=3, value=f"=B{r}/$B${cat_start+len(cat_counts)}").font = REG
        ws.cell(row=r, column=3).number_format = "0.0%"
        for col in range(1, 4): ws.cell(row=r, column=col).border = THIN
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{cat_start}:B{r-1})").font = BOLD
    for col in range(1, 4): ws.cell(row=r, column=col).border = THIN

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A2"

    # ── Sheet 2: Full Patient List ───────────────────────────────────────────
    ws = wb.create_sheet("Full Patient List")
    cols = ["Patient_Name","Mobile_Clean","Patient_UID","Age","Sex",
            "Standardized_Diagnosis","Diagnosis_Category","Diagnosis_Priority",
            "Diagnosis_Status","Comorbidities","Concession_Scheme",
            "Admin_CC","Admin_PD","Admin_BID","Is_VIP"]
    nice = ["Patient Name","Mobile","Clinic UID","Age","Sex","Standardized Diagnosis",
            "Category","Priority","Status","Comorbidities","Concession Scheme",
            "Cons.Charge","Pharm.Disc%","BloodInv.Disc%","VIP"]
    _title(ws, f"Cleaned Patient List — {total} patients", len(cols)); 
    _hdr(ws, 2, nice)
    view = df.sort_values(["Diagnosis_Priority","Diagnosis_Category","Patient_Name"])
    rr = 3
    vip_col = len(cols)            # VIP is last column
    for _, row in view.iterrows():
        for j, c in enumerate(cols, 1):
            val = row[c]
            if c == "Is_VIP": val = "VIP" if str(val) == "True" else ""
            cell = ws.cell(row=rr, column=j, value=val)
            cell.font = REG; cell.border = THIN
        # colour the priority cell
        pcell = ws.cell(row=rr, column=8)
        pcell.fill = PatternFill("solid", fgColor=PRI_FILL.get(row["Diagnosis_Priority"], "FFFFFF"))
        pcell.alignment = Alignment(horizontal="center")
        if str(row["Is_VIP"]) == "True":
            ws.cell(row=rr, column=vip_col).font = Font(name=FONT, bold=True, color="C00000")
        rr += 1
    widths = [24,12,12,5,5,38,24,8,11,22,20,11,11,13,6]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{rr-1}"

    # ── Sheet 3: Taxonomy Breakdown ──────────────────────────────────────────
    ws = wb.create_sheet("Taxonomy Breakdown")
    _title(ws, "Patients per Standardized Diagnosis (diagnosed patients only)", 4)
    _hdr(ws, 2, ["Standardized Diagnosis","Category","Priority","Patients"])
    # strip the "(Suspected)" suffix for grouping the standardized label
    df["Std_base"] = df["Standardized_Diagnosis"].str.replace(r"\s*\(Suspected\)","",regex=True)
    dx = df[df["Diagnosis_Category"] != "No Diagnosis Recorded"]   # only actual diagnoses
    grp = (dx.groupby(["Std_base","Diagnosis_Category","Diagnosis_Priority"])
             .size().reset_index(name="n").sort_values("n", ascending=False))
    rr = 3
    for _, row in grp.iterrows():
        ws.cell(row=rr, column=1, value=row["Std_base"]).font = REG
        ws.cell(row=rr, column=2, value=row["Diagnosis_Category"]).font = REG
        pc = ws.cell(row=rr, column=3, value=row["Diagnosis_Priority"])
        pc.font = REG; pc.alignment = Alignment(horizontal="center")
        pc.fill = PatternFill("solid", fgColor=PRI_FILL.get(row["Diagnosis_Priority"],"FFFFFF"))
        ws.cell(row=rr, column=4, value=int(row["n"])).font = REG
        for col in range(1,5): ws.cell(row=rr, column=col).border = THIN
        rr += 1
    ws.cell(row=rr, column=1, value="TOTAL").font = BOLD
    ws.cell(row=rr, column=4, value=f"=SUM(D3:D{rr-1})").font = BOLD
    for w,j in zip([42,30,9,11],[1,2,3,4]): ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"

    # ── Sheet 4: Age x Category cross-tab ────────────────────────────────────
    ws = wb.create_sheet("Age x Category")
    _title(ws, "Age Band × Diagnosis Category (cohort targeting)", 8)
    bands = [("0-17",0,17),("18-39",18,39),("40-59",40,59),("60-74",60,74),("75+",75,200)]
    cats = list(cat_counts.index)
    _hdr(ws, 2, ["Diagnosis Category"] + [b[0] for b in bands] + ["Total"])
    rr = 3
    for cat in cats:
        ws.cell(row=rr, column=1, value=cat).font = REG
        sub = df[df["Diagnosis_Category"] == cat]
        for k,(lbl,lo,hi) in enumerate(bands, 2):
            n = int(((sub["Age_num"]>=lo)&(sub["Age_num"]<=hi)).sum())
            ws.cell(row=rr, column=k, value=n).font = REG
            ws.cell(row=rr, column=k).border = THIN
        ws.cell(row=rr, column=7+0)  # ensure exists
        last = len(bands)+1
        ws.cell(row=rr, column=last+1, value=f"=SUM(B{rr}:{get_column_letter(last)}{rr})").font = BOLD
        ws.cell(row=rr, column=1).border = THIN
        ws.cell(row=rr, column=last+1).border = THIN
        rr += 1
    ws.column_dimensions["A"].width = 32
    for j in range(2, len(bands)+3): ws.column_dimensions[get_column_letter(j)].width = 10
    ws.freeze_panes = "B3"
    ws.cell(row=rr+1, column=1, value="Note: age is matched by mobile from the Docterz exports (~99.7% of patients).").font = \
        Font(name=FONT, italic=True, size=9, color="595959")

    # ── Sheet 5: VIP & Concessions ───────────────────────────────────────────
    ws = wb.create_sheet("VIP & Concessions")
    _title(ws, "VIP, Concession Codes (CC / PD / BID) & Schemes", 9)
    _hdr(ws, 2, ["Patient Name","Mobile","Standardized Diagnosis","Priority",
                 "VIP","Scheme","Cons.Charge","Pharm.Disc%","BloodInv.Disc%"])
    flag = df[(df["Is_VIP"]=="True") | (df["Admin_CC"]!="") |
              (df["Admin_PD"]!="") | (df["Admin_BID"]!="") | (df["Concession_Scheme"]!="")]
    rr = 3
    for _, row in flag.sort_values(["Is_VIP","Concession_Scheme"], ascending=False).iterrows():
        ws.cell(row=rr, column=1, value=row["Patient_Name"]).font = REG
        ws.cell(row=rr, column=2, value=row["Mobile_Clean"]).font = REG
        ws.cell(row=rr, column=3, value=row["Standardized_Diagnosis"]).font = REG
        ws.cell(row=rr, column=4, value=row["Diagnosis_Priority"]).font = REG
        v = ws.cell(row=rr, column=5, value="VIP" if row["Is_VIP"]=="True" else "")
        if row["Is_VIP"]=="True": v.font = Font(name=FONT, bold=True, color="C00000")
        ws.cell(row=rr, column=6, value=row["Concession_Scheme"]).font = REG
        ws.cell(row=rr, column=7, value=row["Admin_CC"]).font = REG
        ws.cell(row=rr, column=8, value=row["Admin_PD"]).font = REG
        ws.cell(row=rr, column=9, value=row["Admin_BID"]).font = REG
        for col in range(1,10): ws.cell(row=rr, column=col).border = THIN
        rr += 1
    for w,j in zip([24,12,34,8,6,22,11,11,13],range(1,10)): ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    ws.cell(row=rr+1, column=1, value="VIP flag is shown to BOTH reception (on prescription revisit) and follow-up callers (before calling). VIP implies consultation charge 0.").font = \
        Font(name=FONT, italic=True, size=9, color="595959")
    ws.cell(row=rr+2, column=1, value="Concession numbers are literal: CC 5 = Rs.5, CC 500 = Rs.500; PD/BID are percentages. Scheme = Ayushman / Army-ECHS / Staff etc.").font = \
        Font(name=FONT, italic=True, size=9, color="595959")

    # ── Sheet 6: Review (Unclassified) ───────────────────────────────────────
    ws = wb.create_sheet("Review (Unclassified)")
    _title(ws, "Unclassified diagnoses — review to extend taxonomy", 3)
    _hdr(ws, 2, ["Raw Diagnosis Text","Count","Suggested Category (doctor to fill)"])
    unc = (df[df["Diagnosis_Category"]=="Unclassified"]["Diagnosis_Raw"]
           .str.upper().str.strip().value_counts())
    rr = 3
    for txt, cnt in unc.items():
        if not txt: continue
        ws.cell(row=rr, column=1, value=txt).font = REG
        ws.cell(row=rr, column=2, value=int(cnt)).font = REG
        for col in range(1,4): ws.cell(row=rr, column=col).border = THIN
        rr += 1
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 38
    ws.freeze_panes = "A3"

    # ── Sheet 7: Comorbidity / Pathology Targets ─────────────────────────────
    ws = wb.create_sheet("Pathology Targets")
    _title(ws, "Comorbidity Cohort — NK Pathology Package Targeting", 6)
    # summary counts per comorbidity
    from collections import Counter
    cc = Counter()
    for v in df["Comorbidities"]:
        for t in str(v).split("; "):
            if t: cc[t] += 1
    ws.cell(row=2, column=1, value="Comorbidity").font = HDR
    ws.cell(row=2, column=2, value="Patients").font = HDR
    for col in (1, 2):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=2, column=col).border = THIN
    rr = 3
    for name, n in cc.most_common():
        ws.cell(row=rr, column=1, value=name).font = REG
        ws.cell(row=rr, column=2, value=n).font = REG
        for col in (1, 2): ws.cell(row=rr, column=col).border = THIN
        rr += 1
    summary_end = rr - 1

    # full patient list with comorbidities (the actual targeting list)
    rr += 1
    list_hdr = rr
    _hdr(ws, rr, ["Patient Name","Mobile","Comorbidities","Orthopedic Diagnosis","Priority","VIP"])
    rr += 1
    cohort = df[df["Comorbidities"] != ""].sort_values("Comorbidities")
    for _, row in cohort.iterrows():
        ws.cell(row=rr, column=1, value=row["Patient_Name"]).font = REG
        ws.cell(row=rr, column=2, value=row["Mobile_Clean"]).font = REG
        ws.cell(row=rr, column=3, value=row["Comorbidities"]).font = REG
        ws.cell(row=rr, column=4, value=row["Standardized_Diagnosis"]).font = REG
        ws.cell(row=rr, column=5, value=row["Diagnosis_Priority"]).font = REG
        ws.cell(row=rr, column=6, value="VIP" if str(row["Is_VIP"])=="True" else "").font = REG
        for col in range(1, 7): ws.cell(row=rr, column=col).border = THIN
        rr += 1
    for w, j in zip([24,12,30,34,8,6], range(1, 7)):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{list_hdr+1}"
    ws.cell(row=summary_end+2, column=4,
            value=f"{len(cohort)} patients carry a recorded comorbidity — eligible for "
                  f"diabetic / thyroid / cardiac / renal pathology packages.").font = \
        Font(name=FONT, italic=True, size=9, color="595959")

    wb.save(out_path)
    return total, len(grp), len(flag), len(unc)


if __name__ == "__main__":
    out = Path("/mnt/user-data/outputs/MKA_Patient_Diagnosis_Insight.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    t, g, f, u = build(out)
    print(f"Built insight workbook: {out}")
    print(f"  patients={t}  diagnoses={g}  vip/concession={f}  unclassified_terms={u}")
