#!/usr/bin/env python3
"""
plan_followups_from_xlsx.py  —  Dr Manoj Agarwal Clinic
DRY-RUN ONLY WABA follow-up planner.

WHAT IT DOES
  Reads today's Staff Action sheet (the .xlsx the nightly clinic-PC watcher
  pushes to /root/wa/followup-inbox/), groups the follow-up rows by their
  STATUS section, works out which approved WABA template each patient WOULD
  receive, and PRINTS that plan.

WHAT IT DOES NOT DO
  It sends NOTHING. There is no --send flag. It imports waba.py only to reuse
  its phone-normaliser and template-payload builder for an accurate preview.
  This version is physically incapable of messaging a patient.

WHY
  Prove the reading + grouping + template mapping is correct against real data
  BEFORE any send capability or approve-surface is added. One safe step.

USAGE (on the VPS, after a normal nightly push):
    /root/wa/venv/bin/python3 plan_followups_from_xlsx.py
    /root/wa/venv/bin/python3 plan_followups_from_xlsx.py --file /root/wa/followup-inbox/Staff_Action_Today_2026-07-06.xlsx
    /root/wa/venv/bin/python3 plan_followups_from_xlsx.py --show-all      # list every eligible row, not just 8 per section

PRIVACY
  Mobiles are printed MASKED to the last 4 digits (••••2656). Names are shown
  (staff need to recognise the patient); nothing is written to any file.
"""

from __future__ import annotations
import argparse
import glob
import os
import sys
from datetime import date

# waba.py sits next to this script on the VPS (followup-tracker send arm,
# copied into /root/wa alongside this planner). We import ONLY its helpers.
try:
    import waba
except Exception as e:  # pragma: no cover
    print("ERROR: could not import waba.py (must sit next to this script):", e)
    sys.exit(2)

try:
    import openpyxl
except Exception as e:  # pragma: no cover
    print("ERROR: openpyxl not available in this python:", e)
    sys.exit(2)


INBOX_DIR = "/root/wa/followup-inbox"
CALL_SHEET = "Call Sheet"

# Column indexes on the "Call Sheet" tab (0-based), header row = the row whose
# first cell is exactly "S.N". Confirmed from the on-VPS file 06-Jul-2026.
COL = {
    "sn": 0, "pr": 1, "name": 2, "mobile": 3, "diagnosis": 4,
    "date": 5, "od": 6, "status": 7, "key": 12,
}

# STATUS (text BEFORE any " ·" sub-note)  ->  (template, kind, var-count)
# var-count: 1 = name only; 2 = name+date; 3 = name+date+overdue-days
STATUS_TEMPLATE = {
    "Due Today":                   ("drmanoj_followup_due",     "B3", 2),
    "Grace Period":                ("drmanoj_followup_due",     "B3", 2),
    "Actionable Missed Follow-Up": ("drmanoj_followup_missed",  "B4", 2),
    "Probable Dropout":            ("drmanoj_followup_dropout", "B5", 3),
    "Procedure call-back":         ("drmanoj_post_visit",       "B1", 1),
}
# Rows we must NEVER send to (no valid contact / unknown patient).
HARD_EXCLUDE = {"Invalid Mobile / No Contact", "Identity Unresolved"}

# Order sections print in (clinical urgency, dropouts last).
SECTION_ORDER = ["Due Today", "Grace Period", "Actionable Missed Follow-Up",
                 "Probable Dropout", "Procedure call-back"]


def base_status(raw: str) -> str:
    """'Probable Dropout · said call later' -> 'Probable Dropout'."""
    if not raw:
        return ""
    # split on the middot used in the sheet, and on a plain '-' fallback note
    for sep in (" · ", " ·", "·"):
        if sep in raw:
            raw = raw.split(sep, 1)[0]
            break
    return raw.strip()


def mask_mobile(m: str) -> str:
    d = "".join(ch for ch in str(m or "") if ch.isdigit())
    return ("••••" + d[-4:]) if len(d) >= 4 else "(no number)"


def find_latest_file(explicit: str | None) -> str | None:
    if explicit:
        return explicit if os.path.exists(explicit) else None
    # today's dated file first, else the most recent Staff_Action_Today_*.xlsx
    today = date.today().strftime("%Y-%m-%d")
    dated = os.path.join(INBOX_DIR, f"Staff_Action_Today_{today}.xlsx")
    if os.path.exists(dated):
        return dated
    cands = sorted(glob.glob(os.path.join(INBOX_DIR, "Staff_Action_Today_*.xlsx")))
    return cands[-1] if cands else None


def find_header_row(ws) -> int | None:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and str(row[0]).strip() == "S.N":
            return i
        if i > 30:
            break
    return None


def load_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if CALL_SHEET not in wb.sheetnames:
        raise SystemExit(f"'{CALL_SHEET}' tab not found in {os.path.basename(path)}")
    ws = wb[CALL_SHEET]
    all_rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for i, row in enumerate(all_rows):
        if row and str(row[0]).strip() == "S.N":
            hdr = i
            break
    if hdr is None:
        raise SystemExit("Could not find the header row (looking for 'S.N').")
    out = []
    for row in all_rows[hdr + 1:]:
        if not row:
            continue
        sn = row[COL["sn"]] if len(row) > COL["sn"] else None
        status_raw = row[COL["status"]] if len(row) > COL["status"] else None
        # skip blank / banner rows: a real data row has a numeric S.N and a STATUS
        if sn is None or status_raw is None:
            continue
        try:
            int(str(sn).strip())
        except (ValueError, TypeError):
            continue
        out.append({
            "sn": str(sn).strip(),
            "pr": str(row[COL["pr"]] or "").strip(),
            "name": str(row[COL["name"]] or "").strip(),
            "mobile_raw": str(row[COL["mobile"]] or "").strip(),
            "diagnosis": str(row[COL["diagnosis"]] or "").strip(),
            "date": str(row[COL["date"]] or "").strip(),
            "od": str(row[COL["od"]] or "").strip(),
            "status_raw": str(status_raw).strip(),
            "key": str(row[COL["key"]] or "").strip() if len(row) > COL["key"] else "",
        })
    return out


def build_plan(rows: list[dict]):
    """Return (sections, excluded). sections = {status: [entries]}."""
    sections: dict[str, list[dict]] = {}
    excluded: list[tuple[dict, str]] = []
    for r in rows:
        bstat = base_status(r["status_raw"])
        if bstat in HARD_EXCLUDE:
            excluded.append((r, "hard-exclude: " + bstat))
            continue
        if bstat not in STATUS_TEMPLATE:
            excluded.append((r, "unmapped status: " + (bstat or "(blank)")))
            continue
        mobile = waba.normalize_mobile(r["mobile_raw"])
        if not mobile:
            excluded.append((r, "invalid/missing mobile"))
            continue
        if not r["name"]:
            excluded.append((r, "missing name"))
            continue
        tmpl, kind, ncount = STATUS_TEMPLATE[bstat]
        variables = [r["name"]]
        if ncount >= 2:
            variables.append(r["date"])
        if ncount >= 3:
            variables.append(r["od"] or "0")
        sections.setdefault(bstat, []).append({
            "kind": kind, "template": tmpl, "mobile": mobile,
            "name": r["name"], "variables": variables, "key": r["key"],
            "pr": r["pr"], "diagnosis": r["diagnosis"],
        })
    return sections, excluded


def main():
    ap = argparse.ArgumentParser(description="DRY-RUN WABA follow-up planner (sends nothing).")
    ap.add_argument("--file", default=None, help="explicit .xlsx path (default: today's in the inbox)")
    ap.add_argument("--show-all", action="store_true", help="print every eligible row, not just 8 per section")
    args = ap.parse_args()

    path = find_latest_file(args.file)
    if not path:
        print("No Staff_Action_Today_*.xlsx found in", INBOX_DIR)
        sys.exit(1)

    print("=" * 68)
    print(" DRY-RUN WABA PLAN  —  SENDS NOTHING")
    print(" File:", os.path.basename(path))
    print("=" * 68)

    rows = load_rows(path)
    sections, excluded = build_plan(rows)

    total_eligible = sum(len(v) for v in sections.values())
    limit = 10_000 if args.show_all else 8

    ordered = [s for s in SECTION_ORDER if s in sections] + \
              [s for s in sections if s not in SECTION_ORDER]

    for stat in ordered:
        entries = sections[stat]
        tmpl = entries[0]["template"]
        print(f"\n▎ SECTION: {stat}   ({len(entries)} patients)   → template: {tmpl}")
        for e in entries[:limit]:
            varshow = " | ".join(e["variables"])
            print(f"    {mask_mobile(e['mobile'])}  {e['name']:<22}  vars=[{varshow}]  key={e['key']}")
        if len(entries) > limit:
            print(f"    … and {len(entries) - limit} more (use --show-all)")

    print("\n" + "-" * 68)
    print(f" ELIGIBLE (would send): {total_eligible}")
    for stat in ordered:
        print(f"     {stat:<32} {len(sections[stat]):>4}")
    print(f" EXCLUDED (skipped):    {len(excluded)}")
    # summarise exclusion reasons
    reasons: dict[str, int] = {}
    for _, why in excluded:
        key = why.split(":")[0]
        reasons[key] = reasons.get(key, 0) + 1
    for why, n in sorted(reasons.items(), key=lambda x: -x[1]):
        print(f"     {why:<32} {n:>4}")
    print("-" * 68)
    print(" This was a DRY RUN. No WhatsApp message was sent. No file was written.")
    print("=" * 68)


if __name__ == "__main__":
    main()
