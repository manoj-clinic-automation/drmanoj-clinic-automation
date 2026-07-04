#!/usr/bin/env python3
"""
wa_approve.py  —  Dr Manoj Agarwal Clinic
The approve/deselect surface for the follow-up WABA sender. Self-contained.

WHAT IT IS
  A tiny local web page (own port, own service, walled off like clinic-portal).
  It reads today's Staff Action .xlsx, shows the follow-up patients grouped into
  their STATUS sections with checkboxes, and lets Dr Manoj deselect whole
  sections (e.g. "Probable Dropout") or individual patients, then fire the
  approved WABA templates to the ones left ticked.

SAFETY (the whole point)
  OPEN-GATE 1  — a secret key must be in the URL (?k=KEY). No key → 403.
  OPEN-GATE 2  — TEST MODE is the default. In test mode every send goes ONLY to
                 WA_TEST_NUMBER (Dr Manoj's own number). The page cannot reach a
                 patient until he flips the toggle to LIVE.
  LIVE-GATE 1  — the LIVE toggle itself (a deliberate, separate act).
  LIVE-GATE 2  — cap check: if ticked-count > WA_DAILY_CAP, LIVE send is refused
                 unless the explicit override box is also ticked.
  Plus: opt-out list honoured; a send log appended (dedupe + audit); the raw
  .xlsx is never modified; mobiles shown masked to last-4.

CONFIG  (/root/wa/wa_approve.env  — chmod 600; secrets NEVER in this file)
  WA_APPROVE_KEY     (required) the secret that must appear as ?k=... to open
  WA_TEST_NUMBER     (required) Dr Manoj's own 10-digit mobile (test-mode target)
  WA_DAILY_CAP       (optional) default 100
  WA_APPROVE_PORT    (optional) default 8100
  WA_APPROVE_HOST    (optional) default 127.0.0.1
  INBOX_DIR          (optional) default /root/wa/followup-inbox
  OPTOUT_FILE        (optional) default /root/wa/wa_opt_outs.csv
  SEND_LOG           (optional) default /root/wa/wa_send_log.csv
  waba.py's own config (MYOP_* incl. the SECRET token) stays in /root/wa/.env,
  read by waba.py itself — this app never sees the token directly.

RUN
  test:  /root/wa/venv/bin/python3 wa_approve.py
  prod:  gunicorn -w 1 -b 127.0.0.1:8100 wa_approve:app   (via wa-approve.service)
"""

from __future__ import annotations
import csv
import glob
import html
import os
import re
import sys
import uuid
from datetime import date, datetime

from flask import Flask, request, Response, redirect

try:
    import openpyxl
except Exception as e:  # pragma: no cover
    print("ERROR: openpyxl not available:", e); sys.exit(2)

# waba.py must sit next to this file on the VPS (/root/wa/). We reuse it.
try:
    import waba
except Exception as e:  # pragma: no cover
    print("ERROR: could not import waba.py:", e); sys.exit(2)


# ── tiny .env loader (no dependency) ──────────────────────────────────────────
def _load_env(path):
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

_HERE = os.path.dirname(os.path.abspath(__file__))
_load_env(os.path.join(_HERE, "wa_approve.env"))

APPROVE_KEY  = os.environ.get("WA_APPROVE_KEY", "").strip()
TEST_NUMBER  = waba.normalize_mobile(os.environ.get("WA_TEST_NUMBER", "")) or ""
DAILY_CAP    = int(os.environ.get("WA_DAILY_CAP", "100"))
PORT         = int(os.environ.get("WA_APPROVE_PORT", "8100"))
HOST         = os.environ.get("WA_APPROVE_HOST", "127.0.0.1")
INBOX_DIR    = os.environ.get("INBOX_DIR", "/root/wa/followup-inbox")
OPTOUT_FILE  = os.environ.get("OPTOUT_FILE", "/root/wa/wa_opt_outs.csv")
SEND_LOG     = os.environ.get("SEND_LOG", "/root/wa/wa_send_log.csv")

CALL_SHEET = "Call Sheet"
COL = {"sn":0,"pr":1,"name":2,"mobile":3,"diagnosis":4,"date":5,"od":6,"status":7,"key":12}

STATUS_TEMPLATE = {
    "Due Today":                   ("drmanoj_followup_due",     "B3", 2),
    "Grace Period":                ("drmanoj_followup_due",     "B3", 2),
    "Actionable Missed Follow-Up": ("drmanoj_followup_missed",  "B4", 2),
    "Probable Dropout":            ("drmanoj_followup_dropout", "B5", 3),
    "Procedure call-back":         ("drmanoj_post_visit",       "B1", 1),
}
HARD_EXCLUDE = {"Invalid Mobile / No Contact", "Identity Unresolved"}
SECTION_ORDER = ["Due Today", "Grace Period", "Actionable Missed Follow-Up",
                 "Probable Dropout", "Procedure call-back"]

LOG_COLS = ["Timestamp","Mode","Kind","Key","Name","Mobile","Template",
            "Var1","Var2","Var3","Myop_Ref_Id","Message_ID","Conversation_ID",
            "Status","Error"]

app = Flask(__name__)


# ── helpers ───────────────────────────────────────────────────────────────────
def base_status(raw):
    if not raw:
        return ""
    for sep in (" · ", " ·", "·"):
        if sep in raw:
            return raw.split(sep, 1)[0].strip()
    return raw.strip()

def mask(m):
    d = re.sub(r"\D", "", str(m or "")); return "\u2022\u2022\u2022\u2022"+d[-4:] if len(d)>=4 else "(none)"

def find_today_file():
    today = date.today().strftime("%Y-%m-%d")
    p = os.path.join(INBOX_DIR, f"Staff_Action_Today_{today}.xlsx")
    if os.path.exists(p):
        return p
    c = sorted(glob.glob(os.path.join(INBOX_DIR, "Staff_Action_Today_*.xlsx")))
    return c[-1] if c else None

def load_optouts():
    out = set()
    if os.path.exists(OPTOUT_FILE):
        with open(OPTOUT_FILE, newline="", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                if r:
                    m = waba.normalize_mobile(r[0])
                    if m:
                        out.add(m)
    return out

def build_sections(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[CALL_SHEET] if CALL_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = next((i for i,r in enumerate(rows) if r and str(r[0]).strip()=="S.N"), None)
    if hdr is None:
        return {}, 0
    optouts = load_optouts()
    sections = {}
    for r in rows[hdr+1:]:
        if not r:
            continue
        sn = r[COL["sn"]] if len(r)>COL["sn"] else None
        st = r[COL["status"]] if len(r)>COL["status"] else None
        if sn is None or st is None:
            continue
        try:
            int(str(sn).strip())
        except (ValueError, TypeError):
            continue
        g = lambda k: str(r[COL[k]] or "").strip() if len(r)>COL[k] else ""
        bstat = base_status(g("status"))
        if bstat in HARD_EXCLUDE or bstat not in STATUS_TEMPLATE:
            continue
        mobile = waba.normalize_mobile(g("mobile"))
        name = g("name")
        if not mobile or not name:
            continue
        if mobile in optouts:
            continue
        tmpl, kind, ncount = STATUS_TEMPLATE[bstat]
        variables = [name]
        if ncount >= 2:
            variables.append(g("date"))
        if ncount >= 3:
            variables.append(g("od") or "0")
        sections.setdefault(bstat, []).append({
            "kind": kind, "template": tmpl, "mobile": mobile, "name": name,
            "variables": variables, "key": g("key"), "diagnosis": g("diagnosis"),
        })
    total = sum(len(v) for v in sections.values())
    return sections, total

def append_log(rows):
    new = not os.path.exists(SEND_LOG)
    with open(SEND_LOG, "a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=LOG_COLS)
        if new:
            w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in LOG_COLS})

def check_key():
    """OPEN-GATE 1. Return None if OK, or a 403 Response."""
    if not APPROVE_KEY:
        return Response("Server not configured (no WA_APPROVE_KEY set).", 500)
    if request.args.get("k", "") != APPROVE_KEY:
        return Response("403 — key required.", 403)
    return None


# ── the page ──────────────────────────────────────────────────────────────────
PAGE_BUILD = "wa_approve v1.0 (S64)"

def render_page(sections, total, msg=""):
    k = html.escape(request.args.get("k", ""))
    parts = []
    parts.append(f"""<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WABA Follow-up Approve</title>
<style>
 body{{font-family:system-ui,Arial,sans-serif;max-width:820px;margin:0 auto;padding:14px;color:#111}}
 h1{{font-size:19px;margin:.2em 0}} .sub{{color:#666;font-size:13px}}
 .sec{{border:1px solid #ddd;border-radius:8px;margin:12px 0;padding:8px 10px}}
 .sechead{{font-weight:600;font-size:15px}}
 .tmpl{{color:#0a7;font-size:12px}}
 .row{{font-size:13px;padding:2px 0 2px 22px;border-top:1px solid #f2f2f2}}
 .bar{{position:sticky;bottom:0;background:#fff;border-top:2px solid #111;padding:10px;margin-top:16px}}
 .test{{background:#e8f6ff;border:1px solid #7bc}} .live{{background:#ffecec;border:1px solid #e88}}
 button{{font-size:15px;padding:8px 16px;border-radius:6px;border:1px solid #888;cursor:pointer}}
 .go{{background:#0a7;color:#fff;border-color:#087}} label{{cursor:pointer}}
 .msg{{background:#fffceb;border:1px solid #e5d98a;padding:8px;border-radius:6px;margin:8px 0}}
</style>
<script>
 function toggleSec(cb){{document.querySelectorAll('.p_'+cb.dataset.sec).forEach(x=>x.checked=cb.checked);}}
 function onLive(cb){{document.getElementById('modebox').className='bar '+(cb.checked?'live':'test');
   document.getElementById('capwrap').style.display=cb.checked?'block':'none';}}
</script>
</head><body>
<h1>WABA Follow-up — approve &amp; send</h1>
<div class="sub">{PAGE_BUILD} · file: {html.escape(os.path.basename(find_today_file() or '(none)'))} · eligible: {total}</div>
""")
    if msg:
        parts.append(f'<div class="msg">{msg}</div>')

    base = request.path.rstrip('/')
    # strip a trailing '/send' so the result page doesn't build '/send/send'
    if base.endswith('/send'):
        base = base[:-len('/send')]
    send_action = (base + '/send') if base else '/send'
    parts.append('<form method="post" action="'+send_action+'?k='+k+'">')
    ordered = [s for s in SECTION_ORDER if s in sections] + [s for s in sections if s not in SECTION_ORDER]
    for stat in ordered:
        entries = sections[stat]
        secid = re.sub(r"\W","",stat)
        tmpl = entries[0]["template"]
        parts.append(f'<div class="sec"><div class="sechead">'
                     f'<label><input type="checkbox" checked data-sec="{secid}" onclick="toggleSec(this)"> '
                     f'{html.escape(stat)}</label> <span class="tmpl">({len(entries)} · {tmpl})</span></div>')
        for e in entries:
            vid = html.escape(f'{e["kind"]}|{e["key"]}|{e["mobile"]}')
            parts.append(f'<div class="row"><label>'
                         f'<input type="checkbox" name="pick" value="{vid}" class="p_{secid}" checked> '
                         f'{mask(e["mobile"])} &nbsp; {html.escape(e["name"])} '
                         f'<span class="sub">— {html.escape(e["diagnosis"][:30])}</span></label></div>')
        parts.append('</div>')

    parts.append(f"""
<div id="modebox" class="bar test">
 <label><input type="checkbox" name="live" id="live" onclick="onLive(this)"> <b>LIVE</b> — send to real patients
   (unticked = TEST MODE: everything goes only to your number {mask(TEST_NUMBER)})</label>
 <div id="capwrap" style="display:none;margin-top:6px">
   <label><input type="checkbox" name="override"> override daily cap ({DAILY_CAP}) if exceeded</label>
 </div>
 <div style="margin-top:8px"><button type="submit" class="go">Send ticked patients</button></div>
</div></form></body></html>""")
    return "".join(parts)


@app.route("/")
@app.route("/wa-approve")
def index():
    bad = check_key()
    if bad:
        return bad
    path = find_today_file()
    if not path:
        return Response("No Staff_Action_Today_*.xlsx found in "+INBOX_DIR, 404)
    sections, total = build_sections(path)
    return render_page(sections, total)


@app.route("/send", methods=["POST"])
@app.route("/wa-approve/send", methods=["POST"])
def send():
    bad = check_key()
    if bad:
        return bad
    path = find_today_file()
    if not path:
        return Response("No file.", 404)
    sections, total = build_sections(path)

    # index eligible entries by their vid so a tampered POST can't inject numbers
    index = {}
    for entries in sections.values():
        for e in entries:
            index[f'{e["kind"]}|{e["key"]}|{e["mobile"]}'] = e
    picked = [index[v] for v in request.form.getlist("pick") if v in index]

    live = request.form.get("live") == "on"
    override = request.form.get("override") == "on"
    mode = "LIVE" if live else "TEST"

    if not picked:
        return redirect_with_msg("Nothing ticked — nothing sent.")

    # LIVE-GATE 2: cap check
    if live and len(picked) > DAILY_CAP and not override:
        return redirect_with_msg(
            f"PAUSED — {len(picked)} ticked exceeds daily cap {DAILY_CAP}. "
            f"Deselect some, or tick 'override daily cap' and resend. Nothing sent.")

    # config sanity before sending
    try:
        waba.check_config()
    except Exception as e:
        return redirect_with_msg(f"Send config error: {html.escape(str(e))}. Nothing sent.")
    if live and not TEST_NUMBER and False:
        pass
    if not live and not TEST_NUMBER:
        return redirect_with_msg("TEST mode but WA_TEST_NUMBER is not set. Nothing sent.")

    log_rows, sent, failed = [], 0, 0
    for e in picked:
        target = e["mobile"] if live else TEST_NUMBER
        ref = f'{e["kind"]}-{e["key"]}-{uuid.uuid4().hex[:8]}'
        try:
            res = waba.send_template(target, e["template"], e["variables"], myop_ref_id=ref)
        except waba.WabaFatalError as fe:
            log_rows.append(_logrow(mode, e, target, ref, {"ok":False,"code":fe.code,"message":str(fe)}))
            failed += 1
            append_log(log_rows)
            return redirect_with_msg(
                f"HALTED — {html.escape(str(fe))}. Sent {sent} before halt; rest not sent.")
        log_rows.append(_logrow(mode, e, target, ref, res))
        if res.get("ok"):
            sent += 1
        else:
            failed += 1
    append_log(log_rows)
    return redirect_with_msg(f"{mode} send done. Sent: {sent}  Failed: {failed}  "
                             f"(logged to {os.path.basename(SEND_LOG)})")


def _logrow(mode, e, target, ref, res):
    v = (e["variables"] + ["","",""])[:3]
    return {"Timestamp": datetime.now().isoformat(timespec="seconds"),
            "Mode": mode, "Kind": e["kind"], "Key": e["key"], "Name": e["name"],
            "Mobile": target, "Template": e["template"],
            "Var1": v[0], "Var2": v[1], "Var3": v[2], "Myop_Ref_Id": ref,
            "Message_ID": res.get("message_id",""), "Conversation_ID": res.get("conversation_id",""),
            "Status": "SENT" if res.get("ok") else "FAILED",
            "Error": "" if res.get("ok") else f'{res.get("code")} {res.get("message")}'}


def redirect_with_msg(msg):
    # re-render the page with a banner (simplest: rebuild and inject)
    bad = check_key()
    if bad:
        return bad
    path = find_today_file()
    sections, total = build_sections(path) if path else ({}, 0)
    return render_page(sections, total, msg=html.escape(msg))


if __name__ == "__main__":
    app.run(host=HOST, port=PORT)
