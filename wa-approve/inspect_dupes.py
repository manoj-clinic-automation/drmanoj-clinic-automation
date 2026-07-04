#!/usr/bin/env python3
"""
inspect_dupes.py  —  READ-ONLY.  Sends nothing, writes nothing.

Finds rows in the Staff Action sheet that share the SAME name AND the SAME
mobile number (the pattern Dr Manoj flagged as a bug), and prints every column
of each such row side by side so we can see whether they are:
  (a) truly identical duplicates (upstream bug), or
  (b) two genuinely different follow-up obligations for one patient.

It also, separately, lists same-number-but-DIFFERENT-name groups (families
sharing a phone) purely as an FYI count — those are legitimate, not touched.
"""
from __future__ import annotations
import argparse, glob, os, re, sys
from datetime import date

CALL_SHEET = "Call Sheet"
INBOX_DIR = "/root/wa/followup-inbox"
COL = {"sn":0,"pr":1,"name":2,"mobile":3,"diagnosis":4,"date":5,"od":6,"status":7,"key":12}

try:
    import openpyxl
except Exception as e:
    print("ERROR: openpyxl not available:", e); sys.exit(2)


def norm_num(raw):
    d = re.sub(r"\D", "", str(raw or ""))
    if d.startswith("91") and len(d) == 12: d = d[2:]
    elif d.startswith("0") and len(d) == 11: d = d[1:]
    return d[-10:] if len(d) >= 10 else d

def norm_name(raw):
    return re.sub(r"\s+", " ", str(raw or "").strip()).upper()

def mask(m):
    d = re.sub(r"\D", "", str(m or "")); return "••••"+d[-4:] if len(d)>=4 else "(none)"

def find_file(explicit):
    if explicit: return explicit if os.path.exists(explicit) else None
    today = date.today().strftime("%Y-%m-%d")
    p = os.path.join(INBOX_DIR, f"Staff_Action_Today_{today}.xlsx")
    if os.path.exists(p): return p
    c = sorted(glob.glob(os.path.join(INBOX_DIR, "Staff_Action_Today_*.xlsx")))
    return c[-1] if c else None

def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[CALL_SHEET] if CALL_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = next((i for i,r in enumerate(rows) if r and str(r[0]).strip()=="S.N"), None)
    if hdr is None: raise SystemExit("no header row (S.N) found")
    out=[]
    for r in rows[hdr+1:]:
        if not r: continue
        sn = r[COL["sn"]] if len(r)>COL["sn"] else None
        st = r[COL["status"]] if len(r)>COL["status"] else None
        if sn is None or st is None: continue
        try: int(str(sn).strip())
        except (ValueError,TypeError): continue
        g = lambda k: str(r[COL[k]] or "").strip() if len(r)>COL[k] else ""
        out.append({k:g(k) for k in COL})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=None)
    args = ap.parse_args()
    path = find_file(args.file)
    if not path: print("no Staff_Action file found"); sys.exit(1)
    print("="*70); print(" READ-ONLY DUPLICATE INSPECTOR   File:", os.path.basename(path)); print("="*70)

    rows = load(path)

    by_namenum = {}
    by_num = {}
    for r in rows:
        num = norm_num(r["mobile"]); nm = norm_name(r["name"])
        if not num: continue
        by_namenum.setdefault((nm,num), []).append(r)
        by_num.setdefault(num, set()).add(nm)

    same_nn = {k:v for k,v in by_namenum.items() if len(v)>1}
    print(f"\n[A] SAME name + SAME number groups (the flagged bug): {len(same_nn)} group(s)\n")
    for (nm,num),grp in sorted(same_nn.items(), key=lambda x:-len(x[1])):
        print(f"  ── {grp[0]['name']}  {mask(num)}   ({len(grp)} rows)")
        for r in grp:
            print(f"       key={r['key']:<9} status={r['status']:<28} date={r['date']:<7} od={r['od']:<3} diag={r['diagnosis'][:26]}")
        # are they truly identical on the fields that matter?
        sigs = {(r['status'], r['date'], r['od'], r['diagnosis']) for r in grp}
        keys = {r['key'] for r in grp}
        verdict = "IDENTICAL (pure dup)" if len(sigs)==1 else "DIFFERENT obligations (status/date/diag differ)"
        print(f"       → keys differ: {len(keys)>1}   fields identical: {len(sigs)==1}   VERDICT: {verdict}\n")

    fam = {num:names for num,names in by_num.items() if len(names)>1}
    print(f"[B] SAME number + DIFFERENT name (families sharing a phone — LEGIT, untouched): {len(fam)} group(s)")
    for num,names in list(fam.items())[:10]:
        print(f"     {mask(num)}  →  {', '.join(sorted(names))}")
    if len(fam)>10: print(f"     … and {len(fam)-10} more")
    print("\n"+"="*70)
    print(" READ-ONLY. Nothing sent, nothing written.")
    print("="*70)


if __name__ == "__main__":
    main()
