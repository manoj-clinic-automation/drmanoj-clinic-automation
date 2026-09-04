#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
push_salts.py -- S225 §8 item 6 (manojz leg): Amir's salt WORK LIST goes server-side, once.

Reads D:\Downloads\Sanjeevni_Salt_Fix_for_Amir.xlsx -- the S207 work list built from Dr Manoj's answers of 28-Aug
(sheets: 1 Rename salts · 2 Create salts · 3 Change items · 5 Waiting · 6 Cleanups; sheet 4 "Already correct" is
not a task and is not sent) -- plus the salt names from D:\Downloads\margsync\_analysis\Sanjeevni_Salt_Corrections.xlsx
sheet 2 (for the doctor's dropdown on the 7 waiting rows), and POSTs them through the /vendors machine door (the
one the finance front gate already opens to the manojz token). A DONE tick or an answer already on the server is
never overwritten by a later push.

Run on manojz:   python -B push_salts.py        (--dry-run to look only)
Needs openpyxl (on manojz since S198) and the S224_MARG_PURCHASES kit beside this one (token + door).
"""
import hashlib, io, os, sys, urllib.error

HERE = os.path.dirname(os.path.abspath(__file__))
KITS = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(KITS, "S224_MARG_PURCHASES"))
DEF_XLSX = os.environ.get("SALT_FIX_XLSX", r"D:\Downloads\Sanjeevni_Salt_Fix_for_Amir.xlsx")
DEF_NAMES = os.environ.get("SALT_NAMES_XLSX", r"D:\Downloads\margsync\_analysis\Sanjeevni_Salt_Corrections.xlsx")
DEF_BASE = "https://followup.dr-manoj.in/finance/purchase/api"
DEF_REFUSED = os.environ.get("MARG_REFUSED", r"D:\Downloads\margsync\MargArchive\_REFUSED")
SHEETS = (("1 Rename salts", "rename", "Old name"), ("2 Create salts", "create", "Salt name to create"),
          ("3 Change items", "change", "Item"), ("5 Waiting", "waiting", "Item"), ("6 Cleanups", "cleanup", "What"))


def read_tasks(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tasks = []
    for sheet, section, first_hdr in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        rows = [r for r in wb[sheet].iter_rows(values_only=True)]
        hi = next((i for i, r in enumerate(rows) if r and r[0] and str(r[0]).strip().startswith(first_hdr)), None)
        if hi is None:
            continue
        seq = 0
        for r in rows[hi + 1:]:
            if not r or r[0] is None or not str(r[0]).strip():
                continue
            if section in ("rename", "create", "change") and not (len(r) > 1 and r[1] is not None and str(r[1]).strip()):
                continue                                          # a note line under the table, not a task
            seq += 1
            tasks.append(dict(section=section, seq=seq, a=str(r[0]).strip(), b=str(r[1] or "").strip() if len(r) > 1 else "",
                              c=str(r[2] or "").strip() if len(r) > 2 and section != "waiting" else ""))
    return tasks


def read_names(path):
    if not os.path.exists(path):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if len(wb.sheetnames) < 2:
        return []
    out = []
    for r in wb[wb.sheetnames[1]].iter_rows(values_only=True):
        if r and r[0] and not str(r[0]).startswith("Every salt"):
            s = str(r[0]).strip()
            if s and s not in out:
                out.append(s)
    return out


def find_marg_salt_list(folder=DEF_REFUSED):
    """The newest Marg SALT WISE ITEM LIST in the archive's _REFUSED folder (the router knows no signature for it
    yet, so that is where it lands). Recognised by its own shape: header 'S.No. DESCRIPTION | PACKING | P.RATE'."""
    import glob, xlrd
    hits = []
    for f in glob.glob(os.path.join(folder, "*.XLS")) + glob.glob(os.path.join(folder, "*.xls")):
        try:
            sh = xlrd.open_workbook(f).sheet_by_index(0)
            if sh.nrows > 3 and str(sh.cell_value(2, 0)).strip().startswith("S.No. DESCRIPTION") and str(sh.cell_value(2, 1)).strip() == "PACKING":
                hits.append(f)
        except Exception:                                   # noqa: BLE001
            continue
    return sorted(hits, key=os.path.getmtime)[-1] if hits else None


def read_marg_salt_list(path):
    """item -> salt from the salt-wise report: a salt is a line with only column A filled and no leading number;
    an item is 'N     NAME' under it. Returns (items, as_on) -- as_on from the capture stamp in the file name."""
    import re, xlrd
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    cur, out = None, []
    for r in range(3, sh.nrows):
        a = str(sh.cell_value(r, 0)).strip(); pk = str(sh.cell_value(r, 1)).strip()
        if not a or a == "1.0":
            continue
        m = re.match(r"^(\d+)\s{2,}(.+)$", a)
        if m and cur:
            out.append(dict(item=m.group(2).strip(), salt=cur))
        elif not pk and not re.match(r"^\d", a) and a.upper() != "SALT WISE ITEM LIST":
            cur = a.upper()
    m = re.search(r"(\d{4})(\d{2})(\d{2})-\d{6}", os.path.basename(path))
    as_on = "%s-%s-%s" % (m.group(1), m.group(2), m.group(3)) if m else None
    return out, as_on


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    dry = "--dry-run" in argv
    if not os.path.exists(DEF_XLSX):
        print("salts: work list not found: %s" % DEF_XLSX)
        return 2
    tasks = read_tasks(DEF_XLSX)
    names = read_names(DEF_NAMES)
    md5 = hashlib.md5(io.open(DEF_XLSX, "rb").read()).hexdigest()
    marg_path = next((a.split("=", 1)[1] for a in argv if a.startswith("--marg=")), None) or find_marg_salt_list()
    marg_items, marg_as_on, marg_md5 = [], None, ""
    if marg_path and os.path.exists(marg_path):
        marg_items, marg_as_on = read_marg_salt_list(marg_path)
        marg_md5 = hashlib.md5(io.open(marg_path, "rb").read()).hexdigest()
        print("salts: Marg's own salt-wise list found: %s -- %d items, %d salts, as on %s" % (os.path.basename(marg_path), len(marg_items), len({i["salt"] for i in marg_items}), marg_as_on))
    else:
        print("salts: no Marg salt-wise list found in %s -- the 'Marg says' column stays empty until one is exported" % DEF_REFUSED)
    by = {}
    for t in tasks:
        by[t["section"]] = by.get(t["section"], 0) + 1
    print("salts: %d task(s) -- %s; %d salt names for the dropdown; work list %s" % (len(tasks), ", ".join("%s %d" % kv for kv in sorted(by.items())), len(names), md5[:8]))
    if not tasks:
        return 2
    if dry:
        print("salts: dry run -- nothing sent")
        return 0
    import push_purchases as PP
    tok, where = PP.read_token()
    if not tok:
        print("salts: no token available -- nothing sent")
        return 2
    try:
        st, body = PP._post(DEF_BASE + "/vendors", {"pairs": {}, "salt_tasks": tasks, "salts": names, "source": os.path.basename(DEF_XLSX),
                                                    "source_md5": md5, "host": "manojz", "marg_items": marg_items,
                                                    "marg_as_on": marg_as_on, "marg_md5": marg_md5}, tok)
    except urllib.error.HTTPError as e:
        print("salts: server said %s -- nothing recorded" % e.code)
        return 1
    except Exception as e:                                     # noqa: BLE001
        print("salts: could not reach the server (%s)" % e.__class__.__name__)
        return 2
    print("salts: sent (token from %s), server %s: %s" % (where, st, body))
    return 0


if __name__ == "__main__":
    sys.exit(main())
