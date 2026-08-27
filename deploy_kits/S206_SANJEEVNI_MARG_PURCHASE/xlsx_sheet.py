#!/usr/bin/python3
"""
xlsx_sheet.py — let the existing .xls reader read .xlsx, without touching it.

WHY
    `marg_report.py` is live and reads legacy .xls through xlrd. xlrd 2.x
    dropped .xlsx entirely. Half of this practice's financial-year sale
    reports are .xlsx (the April–July ones), so without this they are simply
    invisible — which is exactly how a whole financial year of sale data came
    to be reported as "6.8% coverage".

WHAT IT DOES
    Presents an openpyxl worksheet behind the three attributes xlrd's sheet
    exposes and marg_report actually uses: nrows, ncols, cell_value(r, c).
    `marg_report.py` IS NOT MODIFIED — `_open_sheet` is swapped at runtime by
    the caller, and only for this analysis.
"""


class XlsxSheet(object):
    """An openpyxl worksheet wearing xlrd's clothes."""

    def __init__(self, ws):
        self._rows = [list(r) for r in ws.iter_rows(values_only=True)]
        self.nrows = len(self._rows)
        self.ncols = max((len(r) for r in self._rows), default=0)

    def cell_value(self, r, c):
        if r >= self.nrows:
            return ""
        row = self._rows[r]
        if c >= len(row):
            return ""
        v = row[c]
        if v is None:
            return ""
        # xlrd hands back floats for numbers and str for text; match that, so
        # every downstream `isinstance(v, float)` branch behaves identically.
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, int):
            return float(v)
        if isinstance(v, float):
            return v
        return str(v)


def open_sheet_any(path):
    """Open .xls via xlrd or .xlsx via openpyxl, returning an xlrd-shaped sheet."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return XlsxSheet(wb[wb.sheetnames[0]])
        finally:
            wb.close()
    import xlrd
    return xlrd.open_workbook(path).sheet_by_index(0)
