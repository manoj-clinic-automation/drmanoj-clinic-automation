#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docterz_ingest.py -- S223: read the clinic's Day Revenue sheets from Google Drive into finance.db.

WHY DRIVE AND NOT THE PC. The tracker already writes `Staff_Action_Today_<date>.xlsx` and Drive
already syncs it, continuously, since 04-Jul-2026. So Phase 1 needs no new export, no new push,
and NOTHING that breaks when the tracker PC is switched off. Proved at the S223 open: the box's own
service account lists 147 files in that folder.

THE DATE CONVENTION. The file dated D carries the sheet for the LAST BUSINESS DAY BEFORE D --
usually D-1, but D-2 across a Sunday with no file. This module never guesses: it reads the date off
the sheet's own A1 banner. The filename is used for nothing but logging.

TWO EXPORTS OF ONE DAY. If two files carry the same business date, the one Drive modified LATER
wins -- the S221 rule, applied here.

WHAT IT STORES. Two tables. `clinic_day_revenue` -- one row per business day, counts and money.
`clinic_day_line` -- the itemised lines, added at S223 when the owner asked for the per-patient
table: section, serial, patient, clinic ID, amount, mode, shift. That is his own recorded ruling
for this screen, **clinic ID + NAME on the view, no mobile**.

NO MOBILE NUMBER EXISTS IN THE SOURCE SHEET and none is derived. Nothing identifying is ever
printed to the log, and F-185 still governs the repository absolutely: no identity of any kind in
a kit file, a test fixture, or an evidence file. It lives in finance.db, behind the same SSO gate
as every other screen that already names a patient.

    /root/wa/venv/bin/python3 -B /root/finance/docterz_ingest.py [--db PATH] [--all] [--dry-run]

  --all      re-read every file, not just the ones whose Drive mtime has moved
  --dry-run  parse and report, write nothing
"""
import argparse
import datetime as dt
import glob
import io
import json
import os
import sqlite3
import sys

FOLDER_ID = "1Tls3EsrsJRqjWUY2ZtMcSlObOLP1BgQE"
PREFIX = "Staff_Action_Today_"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
DB_DEFAULT = "/root/finance/finance.db"
TABLE = "clinic_day_revenue"

SCHEMA = """
CREATE TABLE IF NOT EXISTS clinic_day_revenue (
  business_date   TEXT PRIMARY KEY,
  source_file     TEXT NOT NULL,
  source_id       TEXT NOT NULL,
  source_mtime    TEXT NOT NULL,
  taken_at        TEXT NOT NULL,
  cons_count      INTEGER NOT NULL,
  cons_amount_p   INTEGER NOT NULL,
  xray_count      INTEGER NOT NULL,
  xray_amount_p   INTEGER NOT NULL,
  proc_count      INTEGER NOT NULL,
  proc_amount_p   INTEGER NOT NULL,
  total_count     INTEGER NOT NULL,
  total_amount_p  INTEGER NOT NULL,
  morning         INTEGER,
  evening         INTEGER,
  free_revisits   INTEGER NOT NULL,
  free_concession INTEGER NOT NULL,
  f93_phantom_rows INTEGER NOT NULL,
  tender_json     TEXT NOT NULL,
  sheet_total_p   INTEGER,
  sheet_cash_p    INTEGER,
  sheet_online_p  INTEGER,
  variance_note   TEXT NOT NULL DEFAULT ''
);

CREATE TABLE IF NOT EXISTS clinic_day_line (
  business_date TEXT    NOT NULL,
  section       TEXT    NOT NULL,
  sn            INTEGER NOT NULL,
  patient       TEXT    NOT NULL DEFAULT '',
  clinic_id     TEXT    NOT NULL DEFAULT '',
  amount_p      INTEGER NOT NULL DEFAULT 0,
  mode          TEXT    NOT NULL DEFAULT '',
  shift         TEXT    NOT NULL DEFAULT '',
  PRIMARY KEY (business_date, section, sn)
);
CREATE INDEX IF NOT EXISTS ix_clinic_day_line_date ON clinic_day_line(business_date);

CREATE TABLE IF NOT EXISTS clinic_day_tender (
  business_date TEXT    NOT NULL,
  clinic_id     TEXT    NOT NULL DEFAULT '',
  invoice_no    TEXT    NOT NULL DEFAULT '',
  tender        TEXT    NOT NULL,
  amount_p      INTEGER NOT NULL,
  source_file   TEXT    NOT NULL DEFAULT ''
);
CREATE INDEX IF NOT EXISTS ix_clinic_day_tender ON clinic_day_tender(business_date, clinic_id);
"""

TENDER_FILE = "Day_Tenders.csv"


def _find_sa_key():
    for d in ("/root/wa", "/root/wa/keys", "/root"):
        for path in sorted(glob.glob(os.path.join(d, "*.json"))):
            try:
                if b"service_account" in open(path, "rb").read():
                    return path
            except OSError:
                continue
    return ""


def _drive():
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request
    import requests
    sa = _find_sa_key()
    if not sa:
        sys.exit("REFUSING: no service-account key found under /root/wa, /root/wa/keys or /root")
    cred = service_account.Credentials.from_service_account_file(sa, scopes=SCOPES)
    cred.refresh(Request())
    return cred, requests


def find_tender_file():
    """The split-payment legs, written by push_day_tenders.py on the clinic PC into the very
    folder Drive already syncs. Absent is NORMAL -- it simply means that pass has not been run,
    and every split line then reads 'Split Payment' with no breakup, exactly as before."""
    cred, requests = _drive()
    q = "'%s' in parents and trashed=false and name='%s'" % (FOLDER_ID, TENDER_FILE)
    r = requests.get("https://www.googleapis.com/drive/v3/files", params={
        "q": q, "fields": "files(id,name,modifiedTime)", "pageSize": 10},
        headers={"Authorization": "Bearer " + cred.token}, timeout=60)
    if r.status_code != 200:
        return None
    fs = r.json().get("files", [])
    if not fs:
        return None
    return sorted(fs, key=lambda f: f["modifiedTime"])[-1]


def ingest_tenders(con, dry_run=False):
    """Replace the legs table wholesale from the CSV. That file is the whole truth about splits,
    so a partial merge could only ever leave a stale leg behind."""
    import csv
    import io as _io
    f = find_tender_file()
    if not f:
        print("split legs: no %s in the folder -- skipped (this is normal until the PC-side pass "
              "has run)" % TENDER_FILE)
        return 0
    raw = fetch_bytes(f["id"]).decode("utf-8-sig", "replace")
    rd = csv.DictReader(_io.StringIO(raw))
    rows, bad = [], 0
    for r in rd:
        try:
            rows.append((r["business_date"].strip(), (r.get("clinic_id") or "").strip(),
                         (r.get("invoice_no") or "").strip(), r["tender"].strip(),
                         int(r["amount_p"]), (r.get("source_file") or "").strip()))
        except (KeyError, ValueError, TypeError):
            bad += 1
    if bad:
        print("split legs: %d unreadable row(s) in %s -- NOT guessed at" % (bad, TENDER_FILE))
    if dry_run:
        print("split legs: %d rows in %s (DRY RUN -- nothing written)" % (len(rows), TENDER_FILE))
        return len(rows)
    con.execute("DELETE FROM clinic_day_tender")
    con.executemany("INSERT INTO clinic_day_tender (business_date, clinic_id, invoice_no, tender, "
                    "amount_p, source_file) VALUES (?,?,?,?,?,?)", rows)
    con.commit()
    days = len({r[0] for r in rows})
    print("split legs: %d legs across %d days, from %s (modified %s)"
          % (len(rows), days, TENDER_FILE, f["modifiedTime"][:16].replace("T", " ")))
    return len(rows)


def list_day_files():
    """[(title, id, modifiedTime)] for every Staff_Action_Today_*.xlsx in the folder."""
    cred, requests = _drive()
    q = "'%s' in parents and trashed=false" % FOLDER_ID
    out, token = [], None
    while True:
        p = {"q": q, "fields": "nextPageToken,files(id,name,modifiedTime)", "pageSize": 1000}
        if token:
            p["pageToken"] = token
        r = requests.get("https://www.googleapis.com/drive/v3/files", params=p,
                         headers={"Authorization": "Bearer " + cred.token}, timeout=60)
        if r.status_code != 200:
            sys.exit("REFUSING: Drive list returned HTTP %s -- %s"
                     % (r.status_code, r.json().get("error", {}).get("message", "")))
        d = r.json()
        for f in d.get("files", []):
            if f["name"].startswith(PREFIX) and f["name"].endswith(".xlsx"):
                out.append((f["name"], f["id"], f["modifiedTime"]))
        token = d.get("nextPageToken")
        if not token:
            return out


def fetch_bytes(file_id):
    cred, requests = _drive()
    r = requests.get("https://www.googleapis.com/drive/v3/files/" + file_id,
                     params={"alt": "media"},
                     headers={"Authorization": "Bearer " + cred.token}, timeout=120)
    if r.status_code != 200:
        raise RuntimeError("download HTTP %s for %s" % (r.status_code, file_id))
    return r.content


def parse_bytes(raw):
    import openpyxl
    from docterz_day import parse_day_revenue
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    names = [s for s in wb.sheetnames if s.strip() == "Day Revenue"]
    if not names:
        raise RuntimeError("no 'Day Revenue' sheet (sheets: %s)" % ", ".join(wb.sheetnames))
    return parse_day_revenue(wb[names[0]])


def upsert_lines(con, d):
    """Replace the day's lines wholesale. A re-read of the same date must not leave a stale line
    behind, and of two exports of one date the later one has already won by the time we get here."""
    con.execute("DELETE FROM clinic_day_line WHERE business_date=?", (d["business_date"],))
    con.executemany(
        "INSERT INTO clinic_day_line (business_date, section, sn, patient, clinic_id, amount_p, "
        "mode, shift) VALUES (?,?,?,?,?,?,?,?)",
        [(d["business_date"], l["section"], l["sn"], l["patient"], l["clinic_id"],
          l["amount_p"], l["mode"], l["shift"]) for l in d.get("lines", [])])


def upsert(con, d, title, fid, mtime):
    cols = ("business_date source_file source_id source_mtime taken_at cons_count cons_amount_p "
            "xray_count xray_amount_p proc_count proc_amount_p total_count total_amount_p "
            "morning evening free_revisits free_concession f93_phantom_rows tender_json "
            "sheet_total_p sheet_cash_p sheet_online_p variance_note").split()
    vals = (d["business_date"], title, fid, mtime,
            dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            d["cons_count"], d["cons_amount_p"], d["xray_count"], d["xray_amount_p"],
            d["proc_count"], d["proc_amount_p"], d["total_count"], d["total_amount_p"],
            d["morning"], d["evening"], d["free_revisits"], d["free_concession"],
            d["f93_phantom_rows"], json.dumps(d["tender"], sort_keys=True),
            d["sheet_total_p"], d["sheet_cash_p"], d["sheet_online_p"], d["variance_note"])
    con.execute("INSERT INTO %s (%s) VALUES (%s) ON CONFLICT(business_date) DO UPDATE SET %s"
                % (TABLE, ",".join(cols), ",".join("?" * len(cols)),
                   ",".join("%s=excluded.%s" % (c, c) for c in cols if c != "business_date")), vals)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=DB_DEFAULT)
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    try:
        import openpyxl                                    # noqa: F401
    except ImportError:
        sys.exit("REFUSING: openpyxl is not installed in this interpreter. "
                 "Install it before running:  /root/wa/venv/bin/pip install openpyxl")
    con = sqlite3.connect(a.db)
    con.row_factory = sqlite3.Row
    con.executescript(SCHEMA)
    known = {r["source_id"]: r["source_mtime"]
             for r in con.execute("SELECT source_id, source_mtime FROM %s" % TABLE)}
    files = list_day_files()
    print("Drive: %d day files in the folder" % len(files))
    # of two files for one business date, the LATER-modified wins: process oldest first.
    files.sort(key=lambda t: t[2])
    done = skipped = failed = 0
    for title, fid, mtime in files:
        if not a.all and known.get(fid) == mtime:
            skipped += 1
            continue
        try:
            d = parse_bytes(fetch_bytes(fid))
        except Exception as e:                              # noqa: BLE001
            failed += 1
            print("  FAILED  %-38s %s" % (title, e))
            continue
        line = ("  %-38s -> %s  Rs %-7s  %2d lines" %
                (title, d["business_date"], d["total_amount_p"] // 100, d["total_count"]))
        if d["variance_note"]:
            line += "   [%s]" % d["variance_note"]
        print(line)
        if not a.dry_run:
            upsert(con, d, title, fid, mtime)
            upsert_lines(con, d)
        done += 1
    if not a.dry_run:
        con.commit()
    try:
        ingest_tenders(con, a.dry_run)
    except Exception as e:                                      # noqa: BLE001
        print("split legs: skipped (%s) -- the day figures above are unaffected" % e)
    n = con.execute("SELECT COUNT(*) c FROM %s" % TABLE).fetchone()["c"]
    ln = con.execute("SELECT COUNT(*) c FROM clinic_day_line").fetchone()["c"]
    print("read %d, skipped %d unchanged, failed %d; %s now holds %d days and %d lines%s"
          % (done, skipped, failed, TABLE, n, ln,
             "  (DRY RUN - nothing written)" if a.dry_run else ""))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
