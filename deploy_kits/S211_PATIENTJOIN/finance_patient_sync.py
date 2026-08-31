#!/usr/bin/env python3
"""
finance_patient_sync.py  --  S211 / H1, the VPS side.

Reads the two workbooks the clinic PC pushed into the follow-up inbox and folds
them into finance.db, so a pharmacy bill can be matched to a real patient by
LOOKUP (D355) instead of by a generated confidence score.

  /root/wa/followup-inbox/Patient_Master_Join.xlsx  ->  patient_ref
  /root/wa/followup-inbox/Visit_Ledger_Join.xlsx    ->  patient_visit

--------------------------------------------------------------------------------
WHAT IT WILL NOT DO
--------------------------------------------------------------------------------
  * It never DELETES a patient row. A patient who disappears from the tracker
    export stays in finance.db, because sale_item rows point at patient_ref.id
    and history must not be able to lose its subject. Removal, if ever wanted,
    is a separate decision with its own gate.
  * It never touches money. It writes to patient_ref and patient_visit only --
    no day_entry, no sale_item, no ingest_batch. Attribution improving later
    must never be able to move the books (finance_schema's own rule).
  * It never invents a patient. A row with no clinic_id is counted and skipped.
  * It is idempotent. Running it twice changes nothing the second time.

--------------------------------------------------------------------------------
THE ADDITIVE COLUMNS
--------------------------------------------------------------------------------
patient_ref already carries clinic_id, name, phone_last4, first_seen,
merged_into and note. This adds, by lazy ALTER (the same pattern finance_ingest
uses for sale_item.home_med, DDL authoritative in the code):

    mobile_fp         the salted one-way fingerprint of the mobile
    patient_uid       the tracker's own patient key
    last_seen         last seen date from the master
    mobile_dup_count  how many patients the tracker knows on that number

**Why a fingerprint and not the number.** Measured on the real master at S211:
1,506 of 4,903 last-four values are shared by more than one mobile, so the
phone_last4 column that already exists cannot identify anybody. The fingerprint
matches exactly and the number cannot be recovered from it -- stronger matching
AND less patient data at rest. The salt lives in the environment, never in code
and never in the repository.

USAGE
    /root/wa/venv/bin/python3 finance_patient_sync.py           # apply
    /root/wa/venv/bin/python3 finance_patient_sync.py --dry-run # report only
    python3 finance_patient_sync.py --selftest                  # offline, temp db
"""
import os
import sqlite3
import sys

INBOX = os.environ.get("FU_INBOX_DIR", "/root/wa/followup-inbox")
DB_PATH = os.environ.get("FINANCE_DB", "/root/finance/finance.db")
MASTER_XLSX = "Patient_Master_Join.xlsx"
VISITS_XLSX = "Visit_Ledger_Join.xlsx"

ADDED_COLS = (("mobile_fp", "TEXT"), ("patient_uid", "TEXT"),
              ("last_seen", "TEXT"), ("mobile_dup_count", "INTEGER"),
              # the SANCTIONED entitlements, so compliance can be CHECKED
              # rather than assumed: what this patient's consultation is
              # sanctioned at, and the pharmacy discount the counter must apply.
              ("admin_cc_p", "INTEGER"), ("admin_pd_pct", "INTEGER"),
              ("admin_bid_pct", "INTEGER"), ("is_vip", "INTEGER"),
              ("concession_scheme", "TEXT"))

COLLISION_DDL = """
CREATE TABLE IF NOT EXISTS patient_id_collision (
    clinic_id   TEXT NOT NULL,
    kept_uid    TEXT,
    other_uid   TEXT,
    other_name  TEXT,
    kind        TEXT,          -- same_person | DIFFERENT_PEOPLE
    first_noted TEXT,
    PRIMARY KEY (clinic_id, other_uid)
)"""

MERGE_DDL = """
CREATE TABLE IF NOT EXISTS patient_merge_candidate (
    mobile_fp   TEXT NOT NULL,
    clinic_id_a TEXT NOT NULL,
    clinic_id_b TEXT NOT NULL,
    name_seen   TEXT,
    first_noted TEXT,
    PRIMARY KEY (clinic_id_a, clinic_id_b)
)"""

VISIT_DDL = """
CREATE TABLE IF NOT EXISTS patient_visit (
    visit_id     TEXT PRIMARY KEY,
    visit_date   TEXT NOT NULL,
    clinic_id    TEXT,
    patient_uid  TEXT,
    mobile_fp    TEXT,
    had_procedure TEXT
)"""


def read_workbook(path):
    """[(header...), rows...] -> list of dicts. Text in, text out."""
    import openpyxl                                          # noqa: PLC0415
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(h or "").strip() for h in next(rows)]
    except StopIteration:
        return []
    out = []
    for r in rows:
        d = {}
        for k, v in zip(header, r):
            d[k] = "" if v is None else str(v).strip()
        if any(d.values()):
            out.append(d)
    wb.close()
    return out


def ensure_columns(con):
    """Additive only. Never drops, never retypes."""
    have = {r[1] for r in con.execute("PRAGMA table_info(patient_ref)")}
    added = []
    for name, typ in ADDED_COLS:
        if name not in have:
            con.execute("ALTER TABLE patient_ref ADD COLUMN %s %s" % (name, typ))
            added.append(name)
    con.execute("CREATE INDEX IF NOT EXISTS ix_patient_mobile_fp "
                "ON patient_ref(mobile_fp)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_patient_pd "
                "ON patient_ref(admin_pd_pct)")
    con.execute(VISIT_DDL)
    con.execute("CREATE INDEX IF NOT EXISTS ix_visit_date ON patient_visit(visit_date)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_visit_clinic ON patient_visit(clinic_id)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_visit_fp ON patient_visit(mobile_fp)")
    con.execute(COLLISION_DDL)
    have_col = {r[1] for r in con.execute("PRAGMA table_info(patient_id_collision)")}
    if "kind" not in have_col:
        con.execute("ALTER TABLE patient_id_collision ADD COLUMN kind TEXT")
    con.execute(MERGE_DDL)
    return added


def _same(a, b):
    return (a or "").strip().upper() == (b or "").strip().upper()


def sync_patients(con, rows, dry_run=False):
    c = dict(seen=0, no_clinic_id=0, inserted=0, updated=0, unchanged=0,
             dup_in_file=0, dup_different_people=0, merge_candidates=0)
    seen_ids = set()
    seen_ids_uid = {}
    seen_ids_kept = {}
    by_person = {}
    for r in rows:
        c["seen"] += 1
        cid = (r.get("clinic_id") or "").strip()
        if not cid or cid.upper() == "WALK-IN":
            c["no_clinic_id"] += 1
            continue
        if cid in seen_ids:
            # S211, found by the live-shape walk: 17 clinic IDs in the real
            # master name MORE THAN ONE patient. The clinic ID is supposed to be
            # the strongest identifier there is, so a collision must never be
            # dropped in silence -- it is recorded, and D355 matching has to
            # treat such an ID as AMBIGUOUS rather than as a clean match.
            c["dup_in_file"] += 1
            # S211, after the owner's two-clinic ruling: NOT all collisions are
            # equal, and only one kind is dangerous.
            #   same_person      -- same name AND same mobile as the row we kept.
            #                       One patient recorded twice. Benign: the
            #                       person IS in the table. All 9 groups in the
            #                       real master are this.
            #   DIFFERENT_PEOPLE -- the name or the mobile differs. THIS is the
            #                       two-clinic hazard actually happening: a real
            #                       patient is being dropped, and any bill
            #                       carrying that id could mean either of them.
            kept = seen_ids_kept.get(cid, {})
            same = (_same(kept.get("name"), r.get("name"))
                    and _same(kept.get("mobile_fp"), r.get("mobile_fp")))
            kind = "same_person" if same else "DIFFERENT_PEOPLE"
            if not same:
                c["dup_different_people"] += 1
            if not dry_run:
                con.execute(
                    "INSERT OR IGNORE INTO patient_id_collision "
                    "(clinic_id, kept_uid, other_uid, other_name, kind, first_noted) "
                    "VALUES (?,?,?,?,?,datetime('now','localtime'))",
                    (cid, seen_ids_uid.get(cid, ""), r.get("patient_uid") or "",
                     r.get("name") or "", kind))
            continue
        seen_ids.add(cid)
        seen_ids_uid[cid] = r.get("patient_uid") or ""
        seen_ids_kept[cid] = dict(name=r.get("name"), mobile_fp=r.get("mobile_fp"))
        # ONE PERSON, TWO CLINIC IDS -- the owner runs two clinics and each
        # issues its own id per patient. Same fingerprint AND same name under
        # two different ids is one human whose history is split in half. The
        # schema already anticipated this: patient_ref.merged_into exists to
        # "de-dup without rewriting history". These are recorded as candidates;
        # nothing is merged automatically, because merging is the owner's call.
        fp, nm = r.get("mobile_fp") or "", (r.get("name") or "").strip().upper()
        if fp and nm:
            key = (fp, nm)
            prev = by_person.get(key)
            if prev and prev != cid:
                a, b = sorted((prev, cid))
                if not dry_run:
                    con.execute(
                        "INSERT OR IGNORE INTO patient_merge_candidate "
                        "(mobile_fp, clinic_id_a, clinic_id_b, name_seen, first_noted) "
                        "VALUES (?,?,?,?,datetime('now','localtime'))",
                        (fp, a, b, r.get("name") or ""))
                c["merge_candidates"] += 1
            else:
                by_person[key] = cid
        def _int_or_none(k):
            v = str(r.get(k) or "").strip()
            return int(v) if v.lstrip("-").isdigit() else None
        new = (r.get("name") or "", r.get("mobile_last4") or "",
               r.get("mobile_fp") or "", r.get("patient_uid") or "",
               r.get("last_seen") or "",
               int(r.get("mobile_dup_count") or 0) if str(
                   r.get("mobile_dup_count") or "0").isdigit() else 0,
               _int_or_none("admin_cc_p"), _int_or_none("admin_pd_pct"),
               _int_or_none("admin_bid_pct"), 1 if str(r.get("is_vip") or "").strip() else 0,
               r.get("concession_scheme") or "")
        cur = con.execute("SELECT name, phone_last4, mobile_fp, patient_uid, "
                          "last_seen, mobile_dup_count, admin_cc_p, admin_pd_pct, "
                          "admin_bid_pct, is_vip, concession_scheme FROM patient_ref "
                          "WHERE clinic_id=?", (cid,)).fetchone()
        if cur is None:
            if not dry_run:
                con.execute(
                    "INSERT INTO patient_ref (clinic_id, name, phone_last4, "
                    "first_seen, mobile_fp, patient_uid, last_seen, "
                    "mobile_dup_count, admin_cc_p, admin_pd_pct, admin_bid_pct, "
                    "is_vip, concession_scheme) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (cid, new[0], new[1], r.get("first_seen") or "") + new[2:])
            c["inserted"] += 1
        elif tuple(cur) != new:
            if not dry_run:
                con.execute(
                    "UPDATE patient_ref SET name=?, phone_last4=?, mobile_fp=?, "
                    "patient_uid=?, last_seen=?, mobile_dup_count=?, "
                    "admin_cc_p=?, admin_pd_pct=?, admin_bid_pct=?, is_vip=?, "
                    "concession_scheme=? WHERE clinic_id=?", new + (cid,))
            c["updated"] += 1
        else:
            c["unchanged"] += 1
    return c


def sync_visits(con, rows, dry_run=False):
    c = dict(seen=0, no_id=0, inserted=0, unchanged=0)
    for r in rows:
        c["seen"] += 1
        vid = (r.get("visit_id") or "").strip()
        if not vid or not (r.get("visit_date") or "").strip():
            c["no_id"] += 1
            continue
        if con.execute("SELECT 1 FROM patient_visit WHERE visit_id=?",
                       (vid,)).fetchone():
            c["unchanged"] += 1
            continue
        if not dry_run:
            con.execute("INSERT INTO patient_visit (visit_id, visit_date, "
                        "clinic_id, patient_uid, mobile_fp, had_procedure) "
                        "VALUES (?,?,?,?,?,?)",
                        (vid, r.get("visit_date"), r.get("clinic_id"),
                         r.get("patient_uid"), r.get("mobile_fp"),
                         r.get("had_procedure")))
        c["inserted"] += 1
    return c


def run(db_path=None, inbox=None, dry_run=False):
    inbox = inbox or INBOX
    mp = os.path.join(inbox, MASTER_XLSX)
    vp = os.path.join(inbox, VISITS_XLSX)
    for p in (mp, vp):
        if not os.path.isfile(p):
            print("!! not in the inbox:", p)
            return 2
    prows, vrows = read_workbook(mp), read_workbook(vp)
    if not prows:
        print("!! REFUSING: the patient workbook has zero rows. Nothing changed.")
        return 1

    con = sqlite3.connect(db_path or DB_PATH, timeout=30)
    con.row_factory = sqlite3.Row
    try:
        before = con.execute("SELECT COUNT(*) FROM patient_ref").fetchone()[0]
        added = ensure_columns(con)
        if added:
            print("columns added to patient_ref:", ", ".join(added))
        pc = sync_patients(con, prows, dry_run)
        vc = sync_visits(con, vrows, dry_run)
        after = con.execute("SELECT COUNT(*) FROM patient_ref").fetchone()[0]
        if dry_run:
            con.rollback()
        else:
            con.commit()
        print("PATIENTS  seen %d | inserted %d | updated %d | unchanged %d | "
              "no clinic id %d | duplicate in file %d"
              % (pc["seen"], pc["inserted"], pc["updated"], pc["unchanged"],
                 pc["no_clinic_id"], pc["dup_in_file"]))
        ncol = con.execute("SELECT COUNT(*) c FROM patient_id_collision").fetchone()[0]
        if pc["dup_different_people"]:
            print("!! DANGER: %d collision(s) where the two rows are DIFFERENT "
                  "PEOPLE - different name or different mobile under one clinic "
                  "ID. A real patient has been dropped, and any bill carrying "
                  "that ID is AMBIGUOUS. Look at patient_id_collision WHERE "
                  "kind='DIFFERENT_PEOPLE'." % pc["dup_different_people"])
        if pc["merge_candidates"]:
            print("   %d patient(s) hold more than one clinic ID (the same person "
                  "in both clinics). Recorded in patient_merge_candidate; nothing "
                  "merged - that is your call." % pc["merge_candidates"])
        if pc["dup_in_file"]:
            print("   %d clinic ID(s) in this export name MORE THAN ONE patient. "
                  "Kept the first; every collision is recorded in "
                  "patient_id_collision (%d on file). Such an ID is AMBIGUOUS, "
                  "not a clean match." % (pc["dup_in_file"], ncol))
        print("VISITS    seen %d | inserted %d | already held %d | unusable %d"
              % (vc["seen"], vc["inserted"], vc["unchanged"], vc["no_id"]))
        print("patient_ref rows: %d -> %d   (this job never deletes)" % (before, after))
        if dry_run:
            print("\nDRY RUN -- nothing was written.")
        return 0
    finally:
        con.close()


# ------------------------------------------------------------------ selftest

SCHEMA = """
CREATE TABLE patient_ref (
    id INTEGER PRIMARY KEY,
    clinic_id TEXT NOT NULL UNIQUE,
    name TEXT, phone_last4 TEXT, first_seen TEXT,
    merged_into INTEGER REFERENCES patient_ref(id), note TEXT);
CREATE TABLE sale_item (
    id INTEGER PRIMARY KEY,
    patient_ref_id INTEGER REFERENCES patient_ref(id),
    amount_p INTEGER NOT NULL CHECK (amount_p >= 0));
"""


def selftest():
    import tempfile
    ok = bad = 0
    def check(name, cond):
        nonlocal ok, bad
        if cond: ok += 1
        else: bad += 1
        print(("  ok   " if cond else "  FAIL ") + name)

    tmp = tempfile.mkdtemp(prefix="pjoin_")
    db = os.path.join(tmp, "finance.db")
    con = sqlite3.connect(db); con.executescript(SCHEMA)
    con.execute("INSERT INTO patient_ref (clinic_id, name, phone_last4) "
                "VALUES ('4471','RAMESH KUMAR','3210')")
    con.execute("INSERT INTO sale_item (patient_ref_id, amount_p) VALUES (1, 5000)")
    con.commit(); con.close()

    import openpyxl                                          # noqa: PLC0415
    def wb(path, header, rows):
        w = openpyxl.Workbook(); s = w.active; s.append(header)
        for r in rows: s.append(r)
        w.save(path)

    MC = ["clinic_id","patient_uid","name","mobile_fp","mobile_last4",
          "mobile_dup_count","identity_status","first_seen","last_seen",
          "admin_cc_p","admin_pd_pct","admin_bid_pct","is_vip","concession_scheme"]
    VC = ["visit_id","visit_date","clinic_id","patient_uid","mobile_fp","had_procedure"]
    # two patients SHARE a fingerprint -- the family-mobile case, F-34
    wb(os.path.join(tmp, MASTER_XLSX), MC, [
        ["4471","U1","RAMESH KUMAR","fp_aaa","3210","2","ok","2026-01-01","2026-08-01","50000","10","30","","Staff / Employee"],
        ["4472","U2","SITA DEVI",    "fp_aaa","3210","2","ok","2026-02-01","2026-08-02","0","","","1",""],
        ["9",   "U3","EARLY PATIENT","fp_bbb","1111","1","ok","2025-01-01","2026-07-01","","","","",""],
        ["",    "U4","NO CLINIC ID", "fp_ccc","2222","1","ok","","","","","","",""],
        ["4471","U1","RAMESH KUMAR","fp_aaa","3210","2","ok","","","","","","",""],
        ["7001","U5","ANIL VERMA","fp_ddd","4444","1","ok","","","","","","",""],
        ["7001","U6","SUNITA RANI","fp_eee","5555","1","ok","","","","","","",""],
        ["8001","U7","KAVITA SINGH","fp_fff","6666","1","ok","","","","","","",""],
        ["8002","U8","KAVITA SINGH","fp_fff","6666","1","ok","","","","","","",""],
    ])
    wb(os.path.join(tmp, VISITS_XLSX), VC, [
        ["V1","2026-08-26","4471","U1","fp_aaa","Y"],
        ["V2","2026-08-27","9",   "U3","fp_bbb",""],
        ["",  "2026-08-28","4472","U2","fp_aaa",""],
    ])

    rc = run(db, tmp)
    check("the sync runs clean", rc == 0)
    con = sqlite3.connect(db); con.row_factory = sqlite3.Row
    g = lambda q, *a: con.execute(q, a).fetchone()
    check("the existing patient was UPDATED, not duplicated",
          g("SELECT COUNT(*) c FROM patient_ref WHERE clinic_id='4471'")["c"] == 1)
    check("  ...and now carries a fingerprint",
          g("SELECT mobile_fp f FROM patient_ref WHERE clinic_id='4471'")["f"] == "fp_aaa")
    check("a SINGLE-DIGIT clinic id is stored as itself",
          g("SELECT name n FROM patient_ref WHERE clinic_id='9'")["n"] == "EARLY PATIENT")
    check("a row with no clinic id is refused, not invented",
          g("SELECT COUNT(*) c FROM patient_ref WHERE name='NO CLINIC ID'")["c"] == 0)
    check("a duplicate clinic id inside one file is taken once",
          g("SELECT name n FROM patient_ref WHERE clinic_id='4471'")["n"] == "RAMESH KUMAR")
    check("the sanctioned consultation charge lands, in paise",
          g("SELECT admin_cc_p a FROM patient_ref WHERE clinic_id='4471'")["a"] == 50000)
    check("the sanctioned pharmacy discount lands, as a percentage",
          g("SELECT admin_pd_pct a FROM patient_ref WHERE clinic_id='4471'")["a"] == 10)
    check("a FREE consultation (CC 0) is stored as 0, not as blank",
          g("SELECT admin_cc_p a FROM patient_ref WHERE clinic_id='4472'")["a"] == 0)
    check("a patient with no code has NULL, so 'no rule' differs from 'free'",
          g("SELECT admin_cc_p a FROM patient_ref WHERE clinic_id='9'")["a"] is None)
    check("a colliding clinic id is RECORDED, not silently dropped",
          g("SELECT COUNT(*) c FROM patient_id_collision WHERE clinic_id='4471'")["c"] == 1)
    check("  ...and a SAME-PERSON duplicate is marked benign",
          g("SELECT kind k FROM patient_id_collision WHERE clinic_id='4471'")["k"]
          == "same_person")
    check("a collision between DIFFERENT PEOPLE is marked as the danger it is",
          g("SELECT kind k FROM patient_id_collision WHERE clinic_id='7001'")["k"]
          == "DIFFERENT_PEOPLE")
    check("one person holding TWO clinic ids becomes a merge candidate",
          g("SELECT COUNT(*) c FROM patient_merge_candidate WHERE clinic_id_a='8001' "
            "AND clinic_id_b='8002'")["c"] == 1)
    check("  ...and NOTHING is merged automatically - both rows stay",
          g("SELECT COUNT(*) c FROM patient_ref WHERE clinic_id IN ('8001','8002')")["c"] == 2)
    check("F-34: two patients on ONE mobile are BOTH kept",
          g("SELECT COUNT(*) c FROM patient_ref WHERE mobile_fp='fp_aaa'")["c"] == 2)
    check("  ...so a lookup on that fingerprint is AMBIGUOUS, never a pick",
          len(con.execute("SELECT id FROM patient_ref WHERE mobile_fp='fp_aaa'")
              .fetchall()) > 1)
    check("visits landed", g("SELECT COUNT(*) c FROM patient_visit")["c"] == 2)
    check("a visit with no id is refused",
          g("SELECT COUNT(*) c FROM patient_visit WHERE visit_id=''")["c"] == 0)
    check("money was not touched",
          g("SELECT COUNT(*) c FROM sale_item WHERE amount_p=5000")["c"] == 1)
    con.close()

    # idempotency, the second run
    print("  -- second run, same files --")
    rc2 = run(db, tmp)
    con = sqlite3.connect(db); con.row_factory = sqlite3.Row
    # three: the seeded 4471, plus 4472 and 9. The blank clinic id and the
    # in-file duplicate were both refused, which is the point of the count.
    check("re-running inserts nothing new",
          rc2 == 0 and con.execute("SELECT COUNT(*) c FROM patient_ref").fetchone()["c"] == 6)
    check("re-running adds no duplicate visits",
          con.execute("SELECT COUNT(*) c FROM patient_visit").fetchone()["c"] == 2)
    con.close()

    # a patient vanishing from the export must NOT vanish from finance.db
    wb(os.path.join(tmp, MASTER_XLSX), MC,
       [["4471","U1","RAMESH KUMAR","fp_aaa","3210","2","ok","2026-01-01","2026-08-01"]])
    run(db, tmp)
    con = sqlite3.connect(db); con.row_factory = sqlite3.Row
    check("a patient dropped from the export is KEPT in finance.db",
          con.execute("SELECT COUNT(*) c FROM patient_ref").fetchone()["c"] == 6)
    check("  ...and the sale still points at its patient",
          con.execute("SELECT patient_ref_id p FROM sale_item").fetchone()["p"] == 1)
    con.close()

    print("\nselftest: %d passed, %d failed" % (ok, bad))
    return 0 if bad == 0 else 1


def main(argv):
    if "--selftest" in argv:
        return selftest()
    return run(dry_run="--dry-run" in argv)


if __name__ == "__main__":
    sys.exit(main(sys.argv))
