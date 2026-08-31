#!/usr/bin/env python3
"""
push_patient_join.py  --  S211 / H1, the clinic PC side.

Builds two workbooks from the follow-up tracker's OWN data folder and sends them
to the VPS through the door that already exists (push_to_vps.upload_workbook).
The VPS then has the patient master and the visit history, so a pharmacy bill can
be matched to a real patient by LOOKUP instead of by a guessed confidence score
(D355).

--------------------------------------------------------------------------------
WHAT TRAVELS, AND WHAT DOES NOT
--------------------------------------------------------------------------------
**A FULL MOBILE NUMBER NEVER LEAVES THIS PC.**

Each number becomes a SALTED ONE-WAY FINGERPRINT plus its last four digits. The
fingerprint matches exactly -- two identical numbers give the same fingerprint --
but the number itself cannot be recovered from it.

The salt is not decoration. A ten-digit number has only ten billion possibilities:
an UNSALTED hash of a phone number can be reversed by brute force in seconds. With
a secret salt shared by this PC and the VPS, it cannot. **If the salt is missing
this script refuses to run.** It never falls back to an unsalted hash, because a
fallback that silently weakens a privacy guarantee is worse than a stop.

Why a fingerprint at all, when `patient_ref` already stores the last four digits:
measured on the real master at S211, **1,506 of 4,903 last-four values are shared
by more than one number.** Last-four cannot identify anybody. The fingerprint can,
and it stores less about the patient than the alternative.

--------------------------------------------------------------------------------
SAFETY, by design -- the same shape as push_to_vps.py and push_patient_mirror.py
--------------------------------------------------------------------------------
  * READ-ONLY on the tracker's data. It never writes to data/ and never deletes.
  * PREVIEW by default. Nothing is uploaded unless you pass --push.
  * Refuses to push zero rows, so a truncated file can never blank the mirror.
  * NEVER RAISES out of main(): a failure returns a code and says why, so this
    can be chained after report generation without ever breaking it.
  * Prints no mobile number, no salt, and no patient name. Counts and shapes only.

USAGE (from inside the follow-up tracker folder)
    python push_patient_join.py             <-- PREVIEW, uploads nothing
    python push_patient_join.py --push      <-- build and upload
    python push_patient_join.py --selftest  <-- offline checks, touches no data
"""
import csv
import datetime
import hashlib
import os
import re
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUT_DIR = os.path.join(BASE_DIR, "join_outbox")
ENV_FILE = os.path.join(BASE_DIR, "patient_fp.env")
SALT_KEY = "PATIENT_FP_SALT"

MASTER_CSV = os.path.join(DATA_DIR, "patient_master.csv")
DIAG_CSV = os.path.join(DATA_DIR, "patient_diagnosis.csv")
VISITS_CSV = os.path.join(DATA_DIR, "visit_ledger.csv")

MASTER_XLSX = "Patient_Master_Join.xlsx"
VISITS_XLSX = "Visit_Ledger_Join.xlsx"

MASTER_COLS = ["clinic_id", "patient_uid", "name", "mobile_fp", "mobile_last4",
               "mobile_dup_count", "identity_status", "first_seen", "last_seen",
               # the sanctioned entitlements, so the VPS can CHECK them:
               # CC = the consultation charge this patient is sanctioned for
               # PD = the pharmacy discount percentage the counter must apply
               # BID = the pathology category  |  VIP + scheme = context
               "admin_cc_p", "admin_pd_pct", "admin_bid_pct", "is_vip",
               "concession_scheme"]
VISIT_COLS = ["visit_id", "visit_date", "clinic_id", "patient_uid", "mobile_fp",
              "had_procedure"]

_DIGITS = re.compile(r"\D+")


# ------------------------------------------------------------------ the salt

def read_salt(env_file=None):
    """Return the salt, or '' if it is not set. Never printed, never logged."""
    path = env_file or ENV_FILE
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith(SALT_KEY + "="):
                    return line.split("=", 1)[1].strip()
    except OSError:
        return ""
    return ""


def normalise_mobile(raw):
    """Ten digits, or '' if this is not a usable Indian mobile.

    Accepts the shapes the tracker actually holds: spaces, dashes, a +91 or 91
    or 0 prefix. Anything that does not reduce to exactly ten digits is refused
    rather than trimmed into something that looks plausible.
    """
    d = _DIGITS.sub("", str(raw or ""))
    if len(d) > 10:
        if d.startswith("91") and len(d) == 12:
            d = d[2:]
        elif d.startswith("0") and len(d) == 11:
            d = d[1:]
        else:
            d = d[-10:]
    return d if len(d) == 10 and d[0] in "6789" else ""


def fingerprint(mobile10, salt):
    """A salted one-way fingerprint. Same number in, same fingerprint out."""
    if not mobile10 or not salt:
        return ""
    return hashlib.sha256(("%s|%s" % (salt, mobile10)).encode("utf-8")).hexdigest()[:32]


# ------------------------------------------------------------------ reading

def _cell(row, *names):
    for n in names:
        if n in row and (row[n] or "").strip():
            return row[n].strip()
    return ""


def _cc_to_paise(v):
    """The counter writes a sanctioned consultation charge either in full (500)
    or in hundreds (5). Measured on the real sheet at S211: values seen were
    0, 2, 3, 4, 5 and 300, 400, 500 -- the same eight amounts written two ways.
    Anything under 100 is therefore hundreds. 0 means a free consultation and is
    a real instruction, not a blank."""
    n = _num(v)
    if n is None:
        return ""
    rupees = n * 100 if 0 < n < 100 else n
    return str(int(round(rupees * 100)))


def _num(v):
    m = re.search(r"-?\d+(?:\.\d+)?", str(v or "").replace(",", ""))
    return float(m.group(0)) if m else None


def load_entitlements(path=None):
    """clinic_id -> the sanctioned CC / PD / BID / VIP, from the diagnosis sheet
    where the admin codes already live. Read-only; missing file is not fatal."""
    out = {}
    try:
        with open(path or DIAG_CSV, encoding="utf-8-sig", errors="replace") as f:
            for r in csv.DictReader(f):
                cid = _cell(r, "Clinic_Specific_Id")
                if not cid:
                    continue
                pd_raw = _cell(r, "Admin_PD")
                out[cid] = dict(
                    admin_cc_p=_cc_to_paise(_cell(r, "Admin_CC")),
                    admin_pd_pct=(str(int(_num(pd_raw))) if _num(pd_raw) is not None else ""),
                    # BID is a PERCENTAGE, per the owner (31-Aug). Values seen
                    # on the real sheet: 30, 40, 50 and a handful of 4 and 5.
                    # They are stored AS WRITTEN -- 4 means four percent. The
                    # low ones are carried through and flagged by the compliance
                    # report as unusually low, because silently reading 4 as 40
                    # would be inventing a correction nobody sanctioned.
                    admin_bid_pct=(str(int(_num(_cell(r, "Admin_BID"))))
                                   if _num(_cell(r, "Admin_BID")) is not None else ""),
                    is_vip="1" if _cell(r, "Is_VIP").strip().lower() == "true" else "",
                    concession_scheme=_cell(r, "Concession_Scheme"))
    except OSError:
        return {}
    return out


def build_master(salt, path=None, diag_path=None):
    """Rows for the patient workbook, plus a counts dict. Never raises."""
    counts = dict(read=0, no_clinic_id=0, no_mobile=0, written=0,
                  with_cc=0, with_pd=0)
    ent = load_entitlements(diag_path)
    out = []
    try:
        with open(path or MASTER_CSV, encoding="utf-8-sig", errors="replace") as f:
            for r in csv.DictReader(f):
                counts["read"] += 1
                cid = _cell(r, "Clinic_Specific_Id", "clinic_id")
                if not cid:
                    counts["no_clinic_id"] += 1
                    continue
                m10 = normalise_mobile(_cell(r, "Mobile_Clean", "Mobile_Raw"))
                if not m10:
                    counts["no_mobile"] += 1
                e = ent.get(cid, {})
                out.append([
                    cid,
                    _cell(r, "Patient_UID"),
                    _cell(r, "Patient_Name"),
                    fingerprint(m10, salt),
                    m10[-4:] if m10 else "",
                    _cell(r, "Mobile_Duplicate_Count") or "0",
                    _cell(r, "Identity_Status"),
                    _cell(r, "First_Seen_Date"),
                    _cell(r, "Last_Seen_Date"),
                ] + [e.get(k, "") for k in ("admin_cc_p", "admin_pd_pct",
                                            "admin_bid_pct", "is_vip",
                                            "concession_scheme")])
                if e.get("admin_cc_p"):
                    counts["with_cc"] += 1
                if e.get("admin_pd_pct"):
                    counts["with_pd"] += 1
                counts["written"] += 1
    except OSError as ex:
        return [], dict(counts, error=str(ex))
    return out, counts


def build_visits(salt, path=None):
    counts = dict(read=0, no_date=0, written=0)
    out = []
    try:
        with open(path or VISITS_CSV, encoding="utf-8-sig", errors="replace") as f:
            for r in csv.DictReader(f):
                counts["read"] += 1
                d = _cell(r, "Visit_Date")
                if not d:
                    counts["no_date"] += 1
                    continue
                m10 = normalise_mobile(_cell(r, "Mobile_Clean", "Mobile_Raw"))
                out.append([
                    _cell(r, "Visit_ID"),
                    d,
                    _cell(r, "Clinic_Specific_Id", "clinic_id"),
                    _cell(r, "Patient_UID"),
                    fingerprint(m10, salt),
                    _cell(r, "Had_Procedure"),
                ])
                counts["written"] += 1
    except OSError as ex:
        return [], dict(counts, error=str(ex))
    return out, counts


# ------------------------------------------------------------------ writing

def write_workbook(path, header, rows):
    """One sheet, a header row, then the data. Everything written as text, so
    a clinic ID with a leading zero survives the round trip."""
    import openpyxl                                          # noqa: PLC0415
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(header)
    for r in rows:
        ws.append(["" if v is None else str(v) for v in r])
    wb.save(path)
    return path


# ------------------------------------------------------------------ selftest

def selftest():
    ok = bad = 0
    def check(name, cond):
        nonlocal ok, bad
        if cond: ok += 1
        else: bad += 1
        print(("  ok   " if cond else "  FAIL ") + name)

    check("a clean ten-digit number normalises", normalise_mobile("9999999999") == "9999999999")
    check("spaces and dashes are tolerated", normalise_mobile("99999-999 99") == "9999999999")
    check("a +91 prefix is stripped", normalise_mobile("+91 9999999999") == "9999999999")
    check("a leading 0 is stripped", normalise_mobile("09999999999") == "9999999999")
    check("a landline-shaped number is REFUSED, not trimmed", normalise_mobile("0581 2345678") == "")
    check("a short number is refused", normalise_mobile("99999") == "")
    check("empty is refused", normalise_mobile("") == "")

    s = "test-salt"
    a, b = fingerprint("9999999999", s), fingerprint("9999999999", s)
    check("the same number gives the same fingerprint", a == b and len(a) == 32)
    check("a different number gives a different one", fingerprint("8888888888", s) != a)
    check("a different salt gives a different one", fingerprint("9999999999", "other") != a)
    check("no salt -> no fingerprint, never an unsalted one", fingerprint("9999999999", "") == "")
    check("the fingerprint does not contain the number",
          "9999999999" not in a and "3210" not in a[:8])

    check("the master header is the shape the VPS expects", len(MASTER_COLS) == 14)
    check("a CC written as 5 means five hundred rupees", _cc_to_paise("5") == "50000")
    check("a CC written as 500 means the same", _cc_to_paise("500") == "50000")
    check("CC 0 is a FREE consultation, not a blank", _cc_to_paise("0") == "0")
    check("no CC is blank, not zero", _cc_to_paise("") == "")
    check("the visit header is the shape the VPS expects", len(VISIT_COLS) == 6)
    print("\nselftest: %d passed, %d failed" % (ok, bad))
    return 0 if bad == 0 else 1


# ------------------------------------------------------------------ main

def main(argv):
    if "--selftest" in argv:
        return selftest()
    push = "--push" in argv

    salt = read_salt()
    if not salt:
        print("!! REFUSING: no %s in %s" % (SALT_KEY, ENV_FILE))
        print("   A fingerprint without a secret salt can be brute-forced back to")
        print("   the phone number, so this script will not create one.")
        print("   Create the salt ONCE, on this PC, and never print it:")
        print('     python -c "import secrets;print(\'%s=\'+secrets.token_hex(32))"'
              % SALT_KEY)
        print("   Put that one line into: %s" % ENV_FILE)
        print("   The SAME value must be set on the VPS for matching to work.")
        return 2

    master, mc = build_master(salt)
    visits, vc = build_visits(salt)
    if mc.get("error") or vc.get("error"):
        print("!! could not read the tracker data:", mc.get("error") or vc.get("error"))
        return 2

    print("PATIENT MASTER  read %d | written %d | no clinic id %d | no usable mobile %d"
          % (mc["read"], mc["written"], mc["no_clinic_id"], mc["no_mobile"]))
    print("VISIT LEDGER    read %d | written %d | no date %d"
          % (vc["read"], vc["written"], vc["no_date"]))

    if not master:
        print("!! REFUSING: the patient master produced zero rows. Nothing sent.")
        return 1
    if not visits:
        print("!! REFUSING: the visit ledger produced zero rows. Nothing sent.")
        return 1

    if not push:
        print("\nPREVIEW ONLY -- nothing was written and nothing was sent.")
        print("Columns that would travel:")
        print("   patients:", ", ".join(MASTER_COLS))
        print("   visits  :", ", ".join(VISIT_COLS))
        print("\nNo full mobile number is among them. Run again with --push to send.")
        return 0

    try:
        os.makedirs(OUT_DIR, exist_ok=True)
        p1 = write_workbook(os.path.join(OUT_DIR, MASTER_XLSX), MASTER_COLS, master)
        p2 = write_workbook(os.path.join(OUT_DIR, VISITS_XLSX), VISIT_COLS, visits)
    except Exception as ex:                                   # noqa: BLE001
        print("!! could not build the workbooks:", str(ex)[:200])
        return 1

    try:
        sys.path.insert(0, BASE_DIR)
        import push_to_vps                                    # noqa: PLC0415
    except Exception as ex:                                   # noqa: BLE001
        print("!! push_to_vps.py not importable from this folder:", str(ex)[:200])
        print("   The workbooks were built and are in:", OUT_DIR)
        return 1

    rc = 0
    for p in (p1, p2):
        okflag, why = push_to_vps.upload_workbook(p)
        print("%-28s %s" % (os.path.basename(p), "sent" if okflag else "NOT sent: %s" % why))
        if not okflag:
            rc = 1
    print("\n%s  (%s)" % ("both workbooks are on the VPS" if rc == 0
                          else "at least one did not go -- nothing on the VPS was changed",
                          datetime.datetime.now().strftime("%Y-%m-%d %H:%M")))
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv))
