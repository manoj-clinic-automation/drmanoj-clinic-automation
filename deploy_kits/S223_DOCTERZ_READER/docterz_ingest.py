#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
docterz_ingest.py -- S223: read the clinic's Day Revenue sheets from Google Drive into finance.db.

WHY DRIVE AND NOT THE PC. The tracker already writes `Staff_Action_Today_<date>.xlsx` and Drive
already syncs it, continuously, since 04-Jul-2026. So Phase 1 needs no new export, no new push,
and NOTHING that breaks when the tracker PC is switched off. Proved at the S223 open: the box's own
service account lists 147 files in that folder.

THE DATE CONVENTION. The file dated D carries the sheet for the LAST BUSINESS DAY BEFORE D --
usually D-1, but D-2 across a Sunday with no file. This module never guesses: it reads the date off
the sheet's own A1 banner. The filename is used for nothing but logging.

TWO EXPORTS OF ONE DAY. If two files carry the same business date, the one Drive modified LATER
wins -- the S221 rule, applied here.

WHAT IT STORES. Counts and money only. No patient name, no clinic ID, no mobile ever reaches the
database or the log (F-185).

    /root/wa/venv/bin/python3 -B /root/finance/docterz_ingest.py [--db PATH] [--all] [--dry-run]

  --all      re-read every file, not just the ones whose Drive mtime has moved
  --dry-run  parse and report, write nothing
"""
import argparse
import datetime as dt
import glob
import io
import json
import os
import sqlite3
import sys

FOLDER_ID = "1Tls3EsrsJRqjWUY2ZtMcSlObOLP1BgQE"
PREFIX = "Staff_Action_Today_"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
DB_DEFAULT = "/root/finance/finance.db"
TABLE = "clinic_day_revenue"

SCHEMA = """
CREATE TABLE IF NOT EXISTS clinic_day_revenue (
  business_date   TEXT PRIMARY KEY,
  source_file     TEXT NOT NULL,
  source_id       TEXT NOT NULL,
  source_mtime    TEXT NOT NULL,
  taken_at        TEXT NOT NULL,
  cons_count      INTEGER NOT NULL,
  cons_amount_p   INTEGER NOT NULL,
  xray_count      INTEGER NOT NULL,
  xray_amount_p   INTEGER NOT NULL,
  proc_count      INTEGER NOT NULL,
  proc_amount_p   INTEGER NOT NULL,
  total_count     INTEGER NOT NULL,
  total_amount_p  INTEGER NOT NULL,
  morning         INTEGER,
  evening         INTEGER,
  free_revisits   INTEGER NOT NULL,
  free_concession INTEGER NOT NULL,
  f93_phantom_rows INTEGER NOT NULL,
  tender_json     TEXT NOT NULL,
  sheet_total_p   INTEGER,
  sheet_cash_p    INTEGER,
  sheet_online_p  INTEGER,
  variance_note   TEXT NOT NULL DEFAULT ''
);
"""


def _find_sa_key():
    for d in ("/root/wa", "/root/wa/keys", "/root"):
        for path in sorted(glob.glob(os.path.join(d, "*.json"))):
            try:
                if b"service_account" in open(path, "rb").read():
                    return path
            except OSError:
                continue
    return ""


def _drive():
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request
    import requests
    sa = _find_sa_key()
    if not sa:
        sys.exit("REFUSING: no service-account key found under /root/wa, /root/wa/keys or /root")
    cred = service_account.Credentials.from_service_account_file(sa, scopes=SCOPES)
    cred.refresh(Request())
    return cred, requests


def list_day_files():
    """[(title, id, modifiedTime)] for every Staff_Action_Today_*.xlsx in the folder."""
    cred, requests = _drive()
    q = "'%s' in parents and trashed=false" % FOLDER_ID
    out, token = [], None
    while True:
        p = {"q": q, "fields": "nextPageToken,files(id,name,modifiedTime)", "pageSize": 1000}
        if token:
            p["pageToken"] = token
        r = requests.get("https://www.googleapis.com/drive/v3/files", params=p,
                         headers={"Authorization": "Bearer " + cred.token}, timeout=60)
        if r.status_code != 200:
            sys.exit("REFUSING: Drive list returned HTTP %s -- %s"
                     % (r.status_code, r.json().get("error", {}).get("message", "")))
        d = r.json()
        for f in d.get("files", []):
            if f["name"].startswith(PREFIX) and f["name"].endswith(".xlsx"):
                out.append((f["name"], f["id"], f["modifiedTime"]))
        token = d.get("nextPageToken")
        if not token:
            return out


def fetch_bytes(file_id):
    cred, requests = _drive()
    r = requests.get("https://www.googleapis.com/drive/v3/files/" + file_id,
                     params={"alt": "media"},
                     headers={"Authorization": "Bearer " + cred.token}, timeout=120)
    if r.status_code != 200:
        raise RuntimeError("download HTTP %s for %s" % (r.status_code, file_id))
    return r.content


def parse_bytes(raw):
    import openpyxl
    from docterz_day import parse_day_revenue
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    names = [s for s in wb.sheetnames if s.strip() == "Day Revenue"]
    if not names:
        raise RuntimeError("no 'Day Revenue' sheet (sheets: %s)" % ", ".join(wb.sheetnames))
    return parse_day_revenue(wb[names[0]])


def upsert(con, d, title, fid, mtime):
    cols = ("business_date source_file source_id source_mtime taken_at cons_count cons_amount_p "
            "xray_count xray_amount_p proc_count proc_amount_p total_count total_amount_p "
            "morning evening free_revisits free_concession f93_phantom_rows tender_json "
            "sheet_total_p sheet_cash_p sheet_online_p variance_note").split()
    vals = (d["business_date"], title, fid, mtime,
            dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            d["cons_count"], d["cons_amount_p"], d["xray_count"], d["xray_amount_p"],
            d["proc_count"], d["proc_amount_p"], d["total_count"], d["total_amount_p"],
            d["morning"], d["evening"], d["free_revisits"], d["free_concession"],
            d["f93_phantom_rows"], json.dumps(d["tender"], sort_keys=True),
            d["sheet_total_p"], d["sheet_cash_p"], d["sheet_online_p"], d["variance_note"])
    con.execute("INSERT INTO %s (%s) VALUES (%s) ON CONFLICT(business_date) DO UPDATE SET %s"
                % (TABLE, ",".join(cols), ",".join("?" * len(cols)),
                   ",".join("%s=excluded.%s" % (c, c) for c in cols if c != "business_date")), vals)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=DB_DEFAULT)
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    try:
        import openpyxl                                    # noqa: F401
    except ImportError:
        sys.exit("REFUSING: openpyxl is not installed in this interpreter. "
                 "Install it before running:  /root/wa/venv/bin/pip install openpyxl")
    con = sqlite3.connect(a.db)
    con.row_factory = sqlite3.Row
    con.executescript(SCHEMA)
    known = {r["source_id"]: r["source_mtime"]
             for r in con.execute("SELECT source_id, source_mtime FROM %s" % TABLE)}
    files = list_day_files()
    print("Drive: %d day files in the folder" % len(files))
    # of two files for one business date, the LATER-modified wins: process oldest first.
    files.sort(key=lambda t: t[2])
    done = skipped = failed = 0
    for title, fid, mtime in files:
        if not a.all and known.get(fid) == mtime:
            skipped += 1
            continue
        try:
            d = parse_bytes(fetch_bytes(fid))
        except Exception as e:                              # noqa: BLE001
            failed += 1
            print("  FAILED  %-38s %s" % (title, e))
            continue
        line = ("  %-38s -> %s  Rs %-7s  %2d lines" %
                (title, d["business_date"], d["total_amount_p"] // 100, d["total_count"]))
        if d["variance_note"]:
            line += "   [%s]" % d["variance_note"]
        print(line)
        if not a.dry_run:
            upsert(con, d, title, fid, mtime)
        done += 1
    if not a.dry_run:
        con.commit()
    n = con.execute("SELECT COUNT(*) c FROM %s" % TABLE).fetchone()["c"]
    print("read %d, skipped %d unchanged, failed %d; %s now holds %d days%s"
          % (done, skipped, failed, TABLE, n, "  (DRY RUN - nothing written)" if a.dry_run else ""))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
