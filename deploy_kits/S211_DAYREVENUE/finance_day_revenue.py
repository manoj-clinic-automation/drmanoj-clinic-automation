#!/usr/bin/env python3
"""
finance_day_revenue.py  --  S211 / K1: the clinic's day-income sheet, read on the VPS.

THE FINDING THAT MADE THIS SMALL
  The owner asked for the Docterz day-income sheet to be pushed to the VPS so
  Shavez, Alisha, Shivani and he could work from it. It is ALREADY THERE.
  `Staff_Action_Today_<date>.xlsx` carries a **"Day Revenue"** worksheet, and
  `push_to_vps.py` has been uploading that workbook to /root/wa/followup-inbox
  every day. Nothing was reading it. His words for it were "this buried data",
  and that turned out to be literal.

  So this is a READER, not a pipeline. No new push, no new endpoint, no new
  secret, and nothing to change on the clinic PC.

WHAT IT READS
  SUMMARY block   -- Consultations (paid), X-ray, Procedures, Grand Total,
                     Cash, Online/UPI, and the morning/evening shift split.
  DETAIL table    -- one row per paid consultation: patient, clinic id, amount,
                     mode, shift, notes.

  Row positions are NEVER hardcoded. The sheet is laid out for a human, so the
  blocks are found by their LABELS. A sheet that grows a row must not silently
  shift what this reads -- that is the F-107 shape, a check pointed at the wrong
  place and still reporting green.

READ-ONLY. It opens the workbook, returns a dict, and writes nothing anywhere.
"""
import datetime
import glob
import os
import re

INBOX = os.environ.get("FU_INBOX_DIR", "/root/wa/followup-inbox")
SHEET = "Day Revenue"

# the summary labels we understand, mapped to stable keys
SUMMARY_KEYS = (
    ("consultations", re.compile(r"consultation", re.I)),
    ("xray", re.compile(r"x-?ray", re.I)),
    ("procedures", re.compile(r"procedure", re.I)),
    ("grand_total", re.compile(r"grand\s*total", re.I)),
    ("cash", re.compile(r"^\s*cash\s*$", re.I)),
    ("online", re.compile(r"online|upi", re.I)),
)
_MONEY = re.compile(r"-?\d+(?:\.\d+)?")


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = _MONEY.search(str(v).replace(",", ""))
    return float(m.group(0)) if m else None


def latest_workbook(inbox=None, pattern="Staff_Action_Today_*.xlsx"):
    """The most recent staff workbook in the inbox, or None."""
    files = glob.glob(os.path.join(inbox or INBOX, pattern))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def read_day_revenue(path):
    """Return dict(date_hint, summary, shift_note, rows, warnings). Never raises
    on a sheet whose shape has drifted -- it says what it could not find."""
    import openpyxl                                          # noqa: PLC0415
    warnings = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        return dict(summary={}, rows=[], shift_note="", date_hint="",
                    warnings=["the workbook has no '%s' sheet" % SHEET])
    ws = wb[SHEET]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    summary, shift_note = {}, ""
    detail_header_at = None
    for i, row in enumerate(grid):
        cells = ["" if c is None else str(c).strip() for c in row]
        joined = " | ".join(cells)
        for key, rx in SUMMARY_KEYS:
            if key in summary:
                continue
            for j, c in enumerate(cells):
                if c and rx.search(c):
                    nums = [_num(x) for x in cells[j + 1:]]
                    nums = [n for n in nums if n is not None]
                    if nums:
                        summary[key] = dict(
                            count=int(nums[0]) if len(nums) > 1 else None,
                            amount=nums[-1])
                    break
        if not shift_note and re.search(r"shift\s*split", joined, re.I):
            shift_note = joined.strip(" |")
        if detail_header_at is None and any(
                re.fullmatch(r"clinic\s*id", c, re.I) for c in cells):
            detail_header_at = i

    rows = []
    if detail_header_at is None:
        warnings.append("no detail header row containing 'Clinic ID' was found - "
                        "the sheet's layout may have changed")
    else:
        hdr = ["" if c is None else str(c).strip() for c in grid[detail_header_at]]
        for row in grid[detail_header_at + 1:]:
            cells = ["" if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue
            d = {}
            for k, v in zip(hdr, cells):
                if k:
                    d[k] = v
            # a row is a real consultation only if it carries a clinic id
            cid = next((v for k, v in d.items()
                        if re.fullmatch(r"clinic\s*id", k, re.I)), "")
            if not cid:
                continue
            rows.append(d)

    missing = [k for k, _ in SUMMARY_KEYS if k not in summary]
    if missing:
        warnings.append("summary lines not found: %s" % ", ".join(missing))

    # THE DATE COMES FROM THE SHEET, NEVER FROM THE FILENAME.
    # Measured across every real workbook at S211: the file generated on day N
    # carries day N-1's revenue -- the last COMPLETED day, which is correct
    # behaviour. Reading the date off the filename would have mislabelled every
    # day's money by one day, silently, forever.
    date_in_sheet = ""
    title = ""
    for row in grid[:3]:
        for c in row:
            if c and re.search(r"day\s*revenue", str(c), re.I):
                title = str(c)
                break
        if title:
            break
    m = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", title)
    if m:
        try:
            date_in_sheet = datetime.datetime.strptime(
                m.group(1), "%d-%b-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    date_in_filename = ""
    fm = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
    if fm:
        date_in_filename = fm.group(1)
    if not date_in_sheet:
        warnings.append("no date found in the sheet title - the filename date is "
                        "NOT a safe substitute, so this day is unlabelled")
    return dict(business_date=date_in_sheet,
                date_in_sheet=date_in_sheet, date_in_filename=date_in_filename,
                generated_on_is_ahead=bool(date_in_sheet and date_in_filename
                                           and date_in_filename > date_in_sheet),
                summary=summary, shift_note=shift_note,
                rows=rows, warnings=warnings, source=os.path.basename(path))


def cross_check(day):
    """Does the sheet agree with itself? Cash + Online against Grand Total, and
    the detail rows against the consultation line. Reported, never corrected."""
    out = []
    s = day.get("summary", {})
    gt = (s.get("grand_total") or {}).get("amount")
    cash = (s.get("cash") or {}).get("amount")
    online = (s.get("online") or {}).get("amount")
    if gt is not None and cash is not None and online is not None:
        diff = round(gt - (cash + online), 2)
        out.append(dict(check="cash + online vs grand total",
                        agrees=abs(diff) < 0.01, difference=diff))
    # Deliberately NOT checked: the detail table against the summary counts.
    # The detail table lists more rows than the summary counts paid heads, and
    # its money column does not sum to the consultation line. Until the owner
    # says what that column means, asserting a relationship would be inventing
    # one -- and a gate built on a guess is worse than no gate (fault i).
    cons = (s.get("consultations") or {}).get("count")
    if cons is not None:
        out.append(dict(check="paid consultations counted", agrees=None,
                        count=cons, detail_rows=len(day.get("rows", [])),
                        note="reported, not asserted"))
    return out
