#!/usr/bin/env python3
# =============================================================================
#  finance_upi.py  ·  ICICI Merchant (MPR) statement — parse + reconcile
#  Session 179 · B5
#
#  SOURCE OF TRUTH RULING (owner, S179): the BANK is the arbiter for UPI.
#  Marg's payment-mode field is unreliable (operator habit: everything rung as
#  CASH), and the typed figure is only a claim. This module turns the daily
#  ICICI statement into the settled truth, per unit, per business day.
#
#  FILE SHAPE (verified on a real statement, 15-AUG-2026, MID …312505):
#    sheet 'CD_1' · header row of 22 columns · one row per settled transaction
#    (TXN DT = business day · Transaction Amount · RRN · timestamp) · trailer
#    rows 'TID Subtotal' / 'MID Subtotal' / 'Grand Total'.
#    NOTE: the sheet's declared dimension LIES (says 1 row) — the loader must
#    not trust it, hence load_workbook without read_only.
#
#  SELF-CHECK: the parsed transactions must sum exactly to the file's own
#  Grand Total row, or the whole file is REJECTED. We never half-ingest a
#  statement.
#
#  Money is INTEGER PAISE. Requires openpyxl (installer checks).
# =============================================================================

import datetime as dt
import hashlib
import io
import json
import os
import re
import sqlite3

MIDS = {
    "100000000312505": "medical",
    "100000000306941": "clinic",
    "100000000319164": "lab",
}


class StatementRejected(Exception):
    """The file failed its own internal total check, or is not an MPR file."""


def _p(v):
    """float rupees (as the bank writes) -> integer paise, exactly."""
    if v is None:
        return None
    try:
        return int(round(float(str(v).replace(",", "").strip()) * 100))
    except (TypeError, ValueError):
        return None


def _txn_date(v):
    """TXN DT is '14-AUG-26'; Transaction Date is '14-08-2026 20:33:53'.
    Never slice — parse (F-78)."""
    s = str(v or "").strip().split(" ")[0]
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            d = dt.datetime.strptime(s.upper() if "%b" in fmt else s, fmt).date()
            if d.year >= 2000:
                return d
        except ValueError:
            continue
    return None


def parse_mpr(blob):
    """Parse one ICICI MPR xlsx. Returns
       {mid, unit, days: {iso: {'total_p','count','modes':{mode: p}}},
        grand_total_p, txn_sum_p, rows}
    Raises StatementRejected on shape or total failure."""
    import openpyxl                                          # noqa: PLC0415
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as ex:                                  # noqa: BLE001
        raise StatementRejected("not an xlsx: %s" % ex)

    ws = None
    for name in wb.sheetnames:
        if name.upper().startswith("CD"):
            ws = wb[name]
            break
    ws = ws or wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise StatementRejected("empty sheet")
    hdr = [str(c or "").strip() for c in rows[0]]

    def col(name):
        for i, h in enumerate(hdr):
            if h.lower() == name.lower():
                return i
        return None

    c_mid = col("Merchant ID")
    c_txndt = col("TXN DT")
    c_amt = col("Transaction Amount")
    c_mode = col("Mode of Payment")
    c_rrn = col("RRN")
    c_txndate = col("Transaction Date")
    if c_amt is None or (c_txndt is None and c_txndate is None):
        raise StatementRejected("MPR columns not found — header was: %s" % hdr[:8])

    mid = None
    days = {}
    txn_sum_p = 0
    grand_total_p = None
    kept = []

    for r in rows[1:]:
        cells = ["" if v is None else str(v).strip() for v in r]
        joined = " ".join(cells)
        amt_p = _p(r[c_amt]) if c_amt < len(r) else None

        if "Grand Total" in joined:
            grand_total_p = amt_p
            continue
        if "Subtotal" in joined:
            continue

        d = _txn_date(r[c_txndt]) if (c_txndt is not None and c_txndt < len(r)) else None
        if d is None and c_txndate is not None and c_txndate < len(r):
            d = _txn_date(r[c_txndate])
        rrn = cells[c_rrn] if (c_rrn is not None and c_rrn < len(r)) else ""
        if d is None or amt_p is None or not rrn:
            continue                                         # not a transaction row

        if mid is None and c_mid is not None and c_mid < len(r) and cells[c_mid]:
            mid = cells[c_mid]
        mode = (cells[c_mode] if (c_mode is not None and c_mode < len(r)) else "") or "UPI"

        iso = d.isoformat()
        slot = days.setdefault(iso, {"total_p": 0, "count": 0, "modes": {}})
        slot["total_p"] += amt_p
        slot["count"] += 1
        slot["modes"][mode] = slot["modes"].get(mode, 0) + amt_p
        txn_sum_p += amt_p
        kept.append(dict(date=iso, amount_p=amt_p, rrn=rrn, mode=mode))

    if not kept:
        raise StatementRejected("no transaction rows found")
    if grand_total_p is not None and grand_total_p != txn_sum_p:
        raise StatementRejected(
            "file's own Grand Total (%d p) does not equal the sum of its rows (%d p) "
            "— refusing the whole file" % (grand_total_p, txn_sum_p))

    return dict(mid=mid, unit=MIDS.get(mid or "", None), days=days,
                grand_total_p=grand_total_p, txn_sum_p=txn_sum_p, rows=kept)


# --------------------------------------------------------------- persistence

def ingest_statement(con, filename, blob, store_dir, now=None, source_ref=None):
    """Parse + store one statement. Upserts one upi_statement row per
    (merchant, business day). Returns a summary dict; never half-ingests."""
    now = now or dt.datetime.now().replace(microsecond=0).isoformat()
    parsed = parse_mpr(blob)                                 # raises on any failure
    sha = hashlib.sha256(blob).hexdigest()

    path = None
    if store_dir:
        os.makedirs(store_dir, exist_ok=True)
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", filename or "statement.xlsx")[-120:]
        path = os.path.join(store_dir, "%s_%s" % (sha[:10], safe))
        if not os.path.exists(path):
            with open(path, "wb") as fh:
                fh.write(blob)

    out_days = []
    for iso, slot in sorted(parsed["days"].items()):
        con.execute(
            "INSERT INTO upi_statement (merchant_id, unit, statement_date, source_msg_id, "
            " filename, sha256, parsed_total_p, txn_count, ingested_at) "
            "VALUES (?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT(merchant_id, statement_date) DO UPDATE SET "
            " filename=excluded.filename, sha256=excluded.sha256, "
            " parsed_total_p=excluded.parsed_total_p, txn_count=excluded.txn_count, "
            " ingested_at=excluded.ingested_at",
            (parsed["mid"], parsed["unit"], iso, source_ref, filename, sha,
             slot["total_p"], slot["count"], now))
        out_days.append(dict(date=iso, total_p=slot["total_p"], count=slot["count"]))

    con.commit()
    return dict(ok=True, mid=parsed["mid"], unit=parsed["unit"], days=out_days,
                sha256=sha[:12], stored=path is not None)


def reconcile_upi(con, unit, business_date, now=None, tolerance_p=0):
    """Compare the bank's settled UPI against the day's entered UPI.
    Opens / re-opens / closes the upi_vs_statement exception. Returns a dict,
    or None when there is nothing to compare yet (no statement, or no entry —
    both normal early states, not faults)."""
    now = now or dt.datetime.now().replace(microsecond=0).isoformat()

    st = con.execute("SELECT parsed_total_p, txn_count FROM upi_statement "
                     "WHERE unit=? AND statement_date=?", (unit, business_date)).fetchone()
    day = con.execute(
        "SELECT e.id, COALESCE(SUM(CASE WHEN l.mode='upi' THEN l.amount_p END),0) upi_p "
        "FROM day_entry e LEFT JOIN day_line l ON l.day_entry_id=e.id "
        "WHERE e.unit=? AND e.business_date=? GROUP BY e.id", (unit, business_date)).fetchone()

    if st is None or day is None:
        return None

    bank_p = int(st["parsed_total_p"] or 0)
    entered_p = int(day["upi_p"] or 0)
    diff_p = entered_p - bank_p

    if abs(diff_p) <= tolerance_p:
        con.execute("UPDATE recon_exception SET status='resolved', "
                    "resolution='bank statement agrees (settled %d p, %d txns)', closed_at=? "
                    "WHERE unit=? AND business_date=? AND kind='upi_vs_statement' "
                    "AND status IN ('open','acknowledged')",
                    (now, unit, business_date))
    else:
        con.execute(
            "INSERT INTO recon_exception (unit, business_date, kind, expected_p, actual_p, "
            " diff_p, severity, status, detail, opened_at, shout_count) "
            "VALUES (?,?, 'upi_vs_statement', ?,?,?, 'high', 'open', ?, ?, 0) "
            "ON CONFLICT(unit, business_date, kind) DO UPDATE SET "
            " expected_p=excluded.expected_p, actual_p=excluded.actual_p, "
            " diff_p=excluded.diff_p, status='open', detail=excluded.detail, "
            " resolution=NULL, closed_by=NULL, closed_at=NULL",
            (unit, business_date, bank_p, entered_p, diff_p,
             "bank settled %.2f (%d txns) but the day was entered with UPI %.2f — "
             "difference %.2f. The bank is the arbiter; approve only with acknowledgment."
             % (bank_p / 100.0, st["txn_count"] or 0, entered_p / 100.0, diff_p / 100.0),
             now))
    con.commit()
    return dict(bank_p=bank_p, entered_p=entered_p, diff_p=diff_p,
                match=abs(diff_p) <= tolerance_p)


# ------------------------------------------------------------------ selftest

def _build_test_xlsx(rows, grand_total):
    import openpyxl                                          # noqa: PLC0415
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CD_1"
    hdr = ["Date of Settlement Initiation", "Merchant ID", "Terminal ID", "TXN DT",
           "Batch Number", "RRN", "Mode of Payment", "Scheme Name", "Card Type",
           "Card Category", "Settlement Currency", "Transaction Amount",
           "Transaction Charges", "GST", "Net Settlement Amount", "Transaction Date",
           "Transaction ID", "Transaction Type", "Card Number", "AUTHCODE",
           "Card Program", "UTR Transaction Reference No"]
    ws.append(hdr)
    for d, amt, rrn in rows:
        ws.append(["15-AUG-2026", "100000000312505", "EP162174", d, "000271", rrn,
                   "UPI", "UPI", "", "", "INR", amt, 0, 0, amt,
                   "14-08-2026 20:00:00", "TX" + rrn, "PURCHASE", "", "", "", ""])
    ws.append(["15-AUG-2026", "100000000312505", "TID Subtotal", "", "", "", "", "",
               "", "", "", grand_total, 0, 0, grand_total, "", "", "", "", "", "", ""])
    ws.append(["Grand Total", "", "", "", "", "", "", "", "", "", "",
               grand_total, 0, 0, grand_total, "", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def selftest(db_path="finance.db"):
    ok, fail = 0, []

    def check(name, cond):
        nonlocal ok
        if cond:
            ok += 1
        else:
            fail.append(name)

    blob = _build_test_xlsx([("14-AUG-26", 400.0, "R1"), ("14-AUG-26", 750.0, "R2"),
                             ("13-AUG-26", 1311.0, "R3")], 2461.0)
    p = parse_mpr(blob)
    check("mid read", p["mid"] == "100000000312505")
    check("unit mapped", p["unit"] == "medical")
    check("two business days split", sorted(p["days"]) == ["2026-08-13", "2026-08-14"])
    check("14 Aug total", p["days"]["2026-08-14"]["total_p"] == 115000)
    check("13 Aug total", p["days"]["2026-08-13"]["total_p"] == 131100)
    check("trailer rows not counted as txns", len(p["rows"]) == 3)
    check("grand total verified", p["grand_total_p"] == p["txn_sum_p"] == 246100)

    # a file whose own total disagrees must be rejected whole
    bad = _build_test_xlsx([("14-AUG-26", 400.0, "R1")], 999.0)
    try:
        parse_mpr(bad)
        check("corrupt total rejected", False)
    except StatementRejected as ex:
        check("corrupt total rejected", "Grand Total" in str(ex))

    try:
        parse_mpr(b"this is not an xlsx at all")
        check("garbage rejected", False)
    except StatementRejected:
        check("garbage rejected", True)

    if os.path.exists(db_path):
        import shutil
        import tempfile
        fd, tmp = tempfile.mkstemp(prefix="upi_smoke_", suffix=".db")
        os.close(fd)
        shutil.copyfile(db_path, tmp)
        con = sqlite3.connect(tmp)
        con.row_factory = sqlite3.Row

        res = ingest_statement(con, "test_MPR.xlsx", blob, store_dir=None,
                               now="2026-08-15T09:00:00")
        check("ingest ok", res["ok"] and len(res["days"]) == 2)

        # entered 10565 on 13 Aug (the legacy day) vs bank 1311 -> mismatch opens
        r = reconcile_upi(con, "medical", "2026-08-13", now="2026-08-15T09:00:01")
        check("mismatch detected", r is not None and r["match"] is False)
        n = con.execute("SELECT COUNT(*) FROM recon_exception WHERE unit='medical' "
                        "AND business_date='2026-08-13' AND kind='upi_vs_statement' "
                        "AND status='open'").fetchone()[0]
        check("exception opened", n == 1)

        # re-ingest (the .zip duplicate an hour later) must not duplicate anything
        res2 = ingest_statement(con, "test_MPR_again.xlsx", blob, store_dir=None,
                                now="2026-08-15T10:00:00")
        n = con.execute("SELECT COUNT(*) FROM upi_statement WHERE merchant_id=?",
                        ("100000000312505",)).fetchone()[0]
        check("duplicate statement upserts, not duplicates", n == 2 and res2["ok"])

        # no statement for a day -> None, silently (normal early state)
        check("no-statement day is not a fault",
              reconcile_upi(con, "medical", "2026-07-01") is None)

        con.close()
        os.remove(tmp)
    else:
        print("  (db tests skipped — no finance.db)")

    print("UPI %d/%d passed" % (ok, ok + len(fail)))
    for f in fail:
        print("  FAIL:", f)
    return 0 if not fail else 1


if __name__ == "__main__":
    import sys
    sys.exit(selftest(sys.argv[1] if len(sys.argv) > 1 else "finance.db"))
