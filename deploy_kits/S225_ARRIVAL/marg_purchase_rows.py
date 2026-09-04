#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
marg_purchase_rows.py -- S224: the four Marg purchase exports as CONTRACT rows.

SHARED by both legs of S224_PURCHASE_PUSH_CONTRACT.md: the manojz sender
(push_purchases.py) builds its POST body with `payload()`, and the VPS selftest
(selftest_purchase_app.py) feeds the very same rows into purchase_app.py. One
parser, two callers -- the two legs cannot disagree on a row's shape.

DEPENDENCY-FREE except xlrd (legacy .xls) or openpyxl (.xlsx). ITEMWISE is the
S206 marg_purchase.read_purchase() report; hand that function in through
`payload(path, "ITEMWISE", read_purchase=...)` -- this module never imports
across kits. norm() and billno() are copies of packmap.norm / push_expected._billno
(S206/S208), reproduced here with attribution so the key is identical everywhere.

Money leaves here as integer PAISE. Dates leave as ISO yyyy-mm-dd.
"""
import datetime as dt
import hashlib
import os
import re

TYPES = ("ITEMWISE", "BILLWISE", "SUPPLIERWISE", "BILLITEMWISE")
NAME_RE = re.compile(r"__(\d{4}-\d{2}-\d{2})_to_(\d{4}-\d{2}-\d{2})__(\d{8}-\d{6})__")
TITLE_RE = re.compile(r"PURCHASE STATEMENT FROM (\d{2}-\d{2}-\d{4}) TO (\d{2}-\d{2}-\d{4})", re.I)
EXPIRY_TAIL_RE = re.compile(r"(\d{1,2}/\d{2})\s*$")
# Page furniture. Marg reprints the shop header, the title AND the column header
# after every page break, so none of these may be read as a supplier or a bill.
FURNITURE_RE = re.compile(
    r"^(SANJEEVNI|35G/15B|Phone\s*:|PAGE\b|C/F\b|Print\s+HEALTH|MARG\s+ERP|"
    r"[A-Z/ ]*PURCHASE STATEMENT|BILL NO\b|SUPPLIER NAME\b)", re.I)


class Refused(Exception):
    """The file cannot be trusted. Never returned as data."""


# ----------------------------------------------------------- the keys
def norm(s):
    """packmap.norm (S206): case, inner spacing and trailing dots -- never content."""
    s = re.sub(r"\s+", " ", (s or "").upper()).strip()
    return re.sub(r"[.\s]+$", "", s)


def billno(s):
    """push_expected._billno (S208): '232.0' and '232' are the same bill."""
    s = str(s if s is not None else "").strip().strip("'")
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


_CITY_TAILS = ("BAREILLY", "BAREILL", "BAREIL", "BAREI", "BARE", "BAR", "BA")
_DATE_OVERFLOW_RE = re.compile(r"^[A-Z]{0,8}(\d{2}[-/]\d{2}[-/]\d{4})$")


def supplier_key(s):
    """THE bill identity key across the four reports.

    push_expected's finding, confirmed on the August 2026 files: SUPPLIERWISE prints the
    bare name ('JUBILEE AGENCIES'); BILLWISE and ITEMWISE print name + city
    ('JUBILEE AGENCIES          BAREILLY'); and when the name is long the city overflows
    into the DATE column as 'BA03-08-2026'. norm() alone therefore cannot join the two
    reports, so the trailing city token -- whole or truncated -- is dropped from the key.
    The printed name is kept separately for display.
    """
    k = norm(s)
    parts = k.split(" ")
    while len(parts) > 1 and parts[-1] in _CITY_TAILS:
        parts.pop()
    return " ".join(parts)


def iso_date(s):
    """'27-08-2026', '27/08/2026' or '2026-08-27' -> 'yyyy-mm-dd'; else None.
    A city overflow fused onto the front ('BA03-08-2026') is stripped first."""
    s = str(s or "").strip()
    m = _DATE_OVERFLOW_RE.match(s)
    if m:
        s = m.group(1)
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def paise(v):
    """A Marg money cell -> integer paise. '  -' and blank are 0. None when unreadable."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(float(v) * 100))
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return 0
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def _p(v):
    """float rupees or None -> paise or None (line-level fields)."""
    return None if v is None else int(round(v * 100))


# ----------------------------------------------------------- the sheet
class _XlsxSheet(object):
    """An openpyxl worksheet wearing xlrd's clothes (xlsx_sheet.py, S206)."""

    def __init__(self, ws):
        self._rows = [list(r) for r in ws.iter_rows(values_only=True)]
        self.nrows = len(self._rows)
        self.ncols = max((len(r) for r in self._rows), default=0)

    def cell_value(self, r, c):
        if r >= self.nrows:
            return ""
        row = self._rows[r]
        if c >= len(row) or row[c] is None:
            return ""
        v = row[c]
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, int):
            return float(v)
        return v if isinstance(v, float) else str(v)


def open_sheet(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return _XlsxSheet(wb[wb.sheetnames[0]])
        finally:
            wb.close()
    import xlrd
    return xlrd.open_workbook(path).sheet_by_index(0)


def _rows(sh):
    for r in range(sh.nrows):
        yield [sh.cell_value(r, c) for c in range(sh.ncols)]


def _period(sh):
    for r in range(min(30, sh.nrows)):
        m = TITLE_RE.search(_txt(sh.cell_value(r, 0)))
        if m:
            return iso_date(m.group(1)), iso_date(m.group(2))
    return None, None


# ----------------------------------------------------------- the file name
def name_parts(path):
    """(period_from, period_to, export_stamp) from the archive's own file name."""
    m = NAME_RE.search(os.path.basename(path))
    return (m.group(1), m.group(2), m.group(3)) if m else (None, None, None)


def md5_of(path):
    h = hashlib.md5()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ----------------------------------------------------------- BILLWISE
def read_billwise(path):
    """BILL NO. | PARTY NAME | CASH | CREDIT under dd-mm-yyyy date-group rows.
    Nothing above the header is a bill; TOTAL closes the report."""
    sh = open_sheet(path)
    pf, pt = _period(sh)
    rows, grand, seen_hdr, date = [], None, False, None
    for c in _rows(sh):
        first = _txt(c[0]) if c else ""
        if first.upper().startswith("BILL NO"):
            seen_hdr = True
            continue
        if not seen_hdr or not first or FURNITURE_RE.match(first):
            continue
        if first.upper().startswith("TOTAL"):
            grand = paise(c[1]) if len(c) > 1 else None
            date = None
            continue
        rest = [_txt(x) for x in c[1:]]
        d = iso_date(first)
        if d and not any(rest):
            date = d
            continue
        if date is None or not rest or not rest[0]:
            continue
        cash, credit = paise(c[2] if len(c) > 2 else 0), paise(c[3] if len(c) > 3 else 0)
        if cash is None or credit is None:
            raise Refused("unreadable money at bill %s" % first)
        rows.append({"bill_date": date, "bill_no": billno(first), "supplier": rest[0].strip(),
                     "cash_p": cash, "credit_p": credit})
    if not seen_hdr:
        raise Refused("no 'BILL NO.' header -- not a BILL WISE PURCHASE STATEMENT")
    return {"type": "BILLWISE", "period_from": pf, "period_to": pt, "rows": rows,
            "grand_amount_p": grand}


# ----------------------------------------------------------- SUPPLIERWISE
def read_supplierwise(path):
    """SUPPLIER NAME | DATE | BILL NO. | CASH | CREDIT; the supplier is printed once
    and carried down; TOTAL rows are furniture; GRAND TOTAL closes."""
    sh = open_sheet(path)
    pf, pt = _period(sh)
    rows, grand, seen_hdr, supplier = [], None, False, None
    for c in _rows(sh):
        first = _txt(c[0]) if c else ""
        if first.upper().startswith("SUPPLIER NAME"):
            seen_hdr = True
            continue
        if not seen_hdr or FURNITURE_RE.match(first):
            continue
        if first.upper().startswith("GRAND TOTAL"):
            grand = paise(c[2]) if len(c) > 2 else None
            continue
        if first:
            supplier = first.strip()
        second = _txt(c[1]) if len(c) > 1 else ""
        if second.strip().upper().startswith("TOTAL"):
            continue
        d = iso_date(second)
        bill = _txt(c[2]) if len(c) > 2 else ""
        if not (supplier and d and bill):
            continue
        cash, credit = paise(c[3] if len(c) > 3 else 0), paise(c[4] if len(c) > 4 else 0)
        if cash is None or credit is None:
            raise Refused("unreadable money at bill %s" % bill)
        rows.append({"supplier": supplier, "bill_date": d, "bill_no": billno(bill),
                     "cash_p": cash, "credit_p": credit})
    if not seen_hdr:
        raise Refused("no 'SUPPLIER NAME' header -- not a SUPPLIER WISE PURCHASE STATEMENT")
    return {"type": "SUPPLIERWISE", "period_from": pf, "period_to": pt, "rows": rows,
            "grand_amount_p": grand}


# ----------------------------------------------------------- BILLITEMWISE
def _split_pair(cell):
    nums = [n for n in (_num(p) for p in _txt(cell).split()) if n is not None]
    if not nums:
        return None, None
    return nums[0], (nums[1] if len(nums) > 1 else None)


def _pack_batch_exp(c2, c3):
    c2, c3 = _txt(c2), _txt(c3)
    if c3 and re.fullmatch(r"\d{1,2}/\d{2}", c3):
        parts = re.split(r"\s{2,}", c2, maxsplit=1)
        return parts[0].strip() or None, (parts[1].strip() if len(parts) > 1 else None), c3
    m = EXPIRY_TAIL_RE.search(c3)
    if m:
        return c2.strip() or None, (c3[:m.start(1)].strip() or None), m.group(1)
    parts = re.split(r"\s{2,}", c2, maxsplit=1)
    return (parts[0].strip() or None,
            parts[1].strip() if len(parts) > 1 else (c3.strip() or None), None)


def read_billitemwise(path):
    """BILL | ITEM DESCRIPTION | PACKING BATCH | EXP. | TAX | QTY. | FREE | RATE | DIS. |
    AMOUNT NET RATE | LOOS PURC. | AMOUNT, under dd-mm-yyyy date-group rows. There is
    no supplier column in this report: supplier is None and the server finds it."""
    sh = open_sheet(path)
    pf, pt = _period(sh)
    rows, grand, seen_hdr, date = [], None, False, None
    for c in _rows(sh):
        first = _txt(c[0]) if c else ""
        if first.upper() == "BILL" and _txt(c[1]).upper() == "ITEM DESCRIPTION":
            seen_hdr = True
            continue
        if not seen_hdr or not first or FURNITURE_RE.match(first):
            continue
        if first.upper().startswith("TOTAL"):
            grand = paise(c[11]) if len(c) > 11 else None
            date = None
            continue
        rest = [_txt(x) for x in c[1:]]
        d = iso_date(first)
        if d and not any(rest):
            date = d
            continue
        if date is None or not rest[0]:
            continue
        packing, batch, expiry = _pack_batch_exp(c[2], c[3])
        amount, net_rate = _split_pair(c[9])
        loose, purc = _split_pair(c[10])
        rows.append({"bill_no": billno(first), "bill_date": date, "supplier": None,
                     "item": rest[0].strip(), "packing": packing, "batch": batch,
                     "expiry": expiry, "tax": _num(c[4]), "qty": _num(c[5]),
                     "free": _num(c[6]), "rate_p": _p(_num(c[7])),
                     "discount_pct": _num(c[8]), "amount_p": _p(_num(c[11])),
                     "net_rate_p": _p(net_rate), "net_amount_p": _p(amount),
                     "loose_qty": loose, "purchase_rate_p": _p(purc),
                     "direction": "PURCHASE"})
    if not seen_hdr:
        raise Refused("no 'BILL | ITEM DESCRIPTION' header -- not a BILL/ITEM WISE report")
    return {"type": "BILLITEMWISE", "period_from": pf, "period_to": pt, "rows": rows,
            "grand_amount_p": grand}


# ----------------------------------------------------------- ITEMWISE
def itemwise_rows(report):
    """marg_purchase.read_purchase() output -> contract rows (paise, null date)."""
    out = []
    for r in report["rows"]:
        out.append({"bill_no": billno(r.get("bill")), "bill_date": None,
                    "supplier": (r.get("supplier") or "").strip() or None,
                    "item": (r.get("item") or "").strip(), "packing": r.get("packing"),
                    "batch": r.get("batch"), "expiry": r.get("expiry"), "tax": r.get("tax"),
                    "qty": r.get("qty"), "free": r.get("free"), "rate_p": _p(r.get("rate")),
                    "discount_pct": r.get("discount_pct"), "amount_p": _p(r.get("amount")),
                    "net_rate_p": _p(r.get("net_rate")), "net_amount_p": _p(r.get("net_amount")),
                    "loose_qty": r.get("loose_qty"),
                    "purchase_rate_p": _p(r.get("purchase_rate")),
                    "direction": r.get("direction") or "PURCHASE"})
    return out


# ----------------------------------------------------------- the push body
def payload(path, type_, read_purchase=None):
    """The exact JSON body for POST /finance/purchase/api/push, for one archived file.

    type_ is one of TYPES. For ITEMWISE pass marg_purchase.read_purchase (S206 kit).
    """
    if type_ not in TYPES:
        raise Refused("unknown type %r" % type_)
    if type_ == "ITEMWISE":
        if read_purchase is None:
            raise Refused("ITEMWISE needs marg_purchase.read_purchase")
        rep = read_purchase(path)
        rows = itemwise_rows(rep)
        pf, pt = rep.get("period") or (None, None)
        grand = _p(rep.get("grand_amount"))
    else:
        rep = {"BILLWISE": read_billwise, "SUPPLIERWISE": read_supplierwise,
               "BILLITEMWISE": read_billitemwise}[type_](path)
        rows, pf, pt, grand = rep["rows"], rep["period_from"], rep["period_to"], rep["grand_amount_p"]
    nf, nt, stamp = name_parts(path)
    return {"type": type_, "md5": md5_of(path), "file": os.path.basename(path),
            "period_from": pf or nf, "period_to": pt or nt,
            "export_stamp": stamp or dt.datetime.fromtimestamp(
                os.path.getmtime(path)).strftime("%Y%m%d-%H%M%S"),
            "n_rows": len(rows), "grand_amount_p": grand, "rows": rows}
