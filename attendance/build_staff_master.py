#!/usr/bin/env python3
"""
build_staff_master.py  —  rebuild staff_master.csv from the salary workbook.

The salary workbook (Salary_System_2026.xlsx, "Staff Master" sheet) stays the
single source of truth for names, timings, Sunday timings, base salary,
allowed offs, Sunday roster group and the minutes-exempt flag.  Run this
whenever you add or change an employee there, then copy the new
staff_master.csv up to the VPS.

v2 (S154): carries the two roster columns added at S153 —
    sunday_group    (col I, header row 2)  values: A / B / C / ARJ
    minutes_exempt  (col J, header row 2)  values: Y / N
The script REFUSES to run if either column is missing or any value is
invalid, so a stale workbook can never silently produce a CSV that drops
the roster (the failure that this version exists to prevent).

Usage:
    python build_staff_master.py
    python build_staff_master.py "C:\\path\\to\\Salary_System_2026.xlsx"

Output:  staff_master.csv  (next to this script)
"""
import sys
import re
import csv
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "Salary_System_2026.xlsx"
OUT = "staff_master.csv"

HEADER_ROW = 2
REQUIRED_NEW = {"sunday_group": ("A", "B", "C", "ARJ"),
                "minutes_exempt": ("Y", "N")}


def die(msg):
    print("ERROR: " + msg)
    print("staff_master.csv was NOT written.")
    sys.exit(1)


def split_timing(s):
    """'08:00-16:00' or '09:30-15:30 + 18:00-21:00' -> (start, end, raw_note)."""
    if not s:
        return "", "", ""
    times = re.findall(r"\d{1,2}:\d{2}", str(s))
    if not times:
        return "", "", ""
    start = times[0]
    end = times[-1]
    note = str(s).strip() if "+" in str(s) else ""   # keep raw only for split shifts
    return start, end, note


def find_new_columns(ws):
    """Locate sunday_group / minutes_exempt by header name in HEADER_ROW.
    Matching is case/space-insensitive so 'Sunday Group' also works."""
    found = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(HEADER_ROW, c).value
        if h is None:
            continue
        key = str(h).strip().lower().replace(" ", "_")
        if key in REQUIRED_NEW:
            found[key] = c
    missing = [k for k in REQUIRED_NEW if k not in found]
    if missing:
        die("Staff Master sheet is missing column(s): " + ", ".join(missing)
            + ".\nAdd them in header row " + str(HEADER_ROW)
            + " (I=sunday_group, J=minutes_exempt) and fill every staff row, "
              "then run again.\nA rebuild without these columns would break "
              "the Sunday roster and Arjun's exemption on the VPS.")
    return found


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if "Staff Master" not in wb.sheetnames:
        die('sheet "Staff Master" not found in ' + SRC)
    ws = wb["Staff Master"]
    newcols = find_new_columns(ws)

    rows = []
    problems = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code in (None, ""):
            continue  # no Emp Code -> not on the biometric device (salary-only)
        name = ws.cell(r, 2).value or f"#{code}"
        dept = ws.cell(r, 3).value or ""
        base = ws.cell(r, 4).value
        offs = ws.cell(r, 5).value
        wd_start, wd_end, wd_note = split_timing(ws.cell(r, 6).value)
        sun_start, sun_end, _ = split_timing(ws.cell(r, 7).value)

        grp = ws.cell(r, newcols["sunday_group"]).value
        exm = ws.cell(r, newcols["minutes_exempt"]).value
        grp = "" if grp is None else str(grp).strip().upper()
        exm = "" if exm is None else str(exm).strip().upper()
        if grp not in REQUIRED_NEW["sunday_group"]:
            problems.append(f"row {r} ({name}): sunday_group is '{grp or 'blank'}'"
                            " — must be A, B, C or ARJ")
        if exm not in REQUIRED_NEW["minutes_exempt"]:
            problems.append(f"row {r} ({name}): minutes_exempt is '{exm or 'blank'}'"
                            " — must be Y or N")

        rows.append({
            "user_id": int(code),
            "name": str(name).strip(),
            "department": str(dept).strip(),
            "base_salary": "" if base in (None, "") else int(base),
            "allowed_offs": "" if offs in (None, "") else offs,
            "wd_start": wd_start,
            "wd_end": wd_end,
            "sun_start": sun_start,
            "sun_end": sun_end,
            "active": "Y",
            "timing_note": wd_note,
            "sunday_group": grp,
            "minutes_exempt": exm,
        })

    if problems:
        die("invalid roster values in the Staff Master sheet:\n  "
            + "\n  ".join(problems))
    if not rows:
        die("no staff rows with an Emp Code were found.")

    cols = ["user_id", "name", "department", "base_salary", "allowed_offs",
            "wd_start", "wd_end", "sun_start", "sun_end", "active",
            "timing_note", "sunday_group", "minutes_exempt"]
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in rows:
            w.writerow(row)

    # eyeball summary — groups and exemptions, so a wrong fill is visible at once
    groups = {}
    exempt = []
    for row in rows:
        groups.setdefault(row["sunday_group"], []).append(row["name"])
        if row["minutes_exempt"] == "Y":
            exempt.append(row["name"])
    print(f"Wrote {OUT} with {len(rows)} staff.")
    for g in ("A", "B", "C", "ARJ"):
        if g in groups:
            print(f"  group {g}: " + ", ".join(groups[g]))
    print("  minutes_exempt: " + (", ".join(exempt) if exempt else "none"))


if __name__ == "__main__":
    main()
